#!/usr/bin/env python3
"""
Enhanced diagnostic script to scan entire Excel file for time patterns
"""

import sys
from pathlib import Path
from openpyxl import load_workbook
import sqlite3
from collections import Counter

def enhanced_time_diagnosis(excel_file_path: str, db_path: str = None):
    """
    Comprehensive time analysis of entire Excel file.
    """
    print("🔍 ENHANCED TIME DIAGNOSIS - FULL FILE SCAN")
    print("=" * 60)
    
    if not Path(excel_file_path).exists():
        print(f"❌ Excel file not found: {excel_file_path}")
        return
    
    # Step 1: Full file time pattern analysis
    print("\n📊 STEP 1: Complete Excel File Analysis")
    print("-" * 50)
    
    try:
        workbook = load_workbook(excel_file_path, data_only=False)
        worksheet = workbook.active
        
        # Find time columns
        header_row = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        time_columns = {}
        
        for col_idx, header in enumerate(header_row):
            if header and str(header).strip().lower() in ['time in', 'time out', 'time_in', 'time_out']:
                time_columns[col_idx] = str(header).strip()
        
        print(f"🕐 Found time columns: {list(time_columns.values())}")
        
        # Scan entire file for time patterns
        time_patterns = Counter()
        twenty_four_patterns = Counter()
        midnight_patterns = Counter()
        sample_rows = {}
        total_rows = 0
        
        print(f"📋 Scanning entire file...")
        
        for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(cell for cell in row):
                continue
                
            total_rows += 1
            
            for col_idx, col_name in time_columns.items():
                if col_idx < len(row) and row[col_idx]:
                    value = str(row[col_idx]).strip()
                    
                    # Count all time patterns
                    time_patterns[value] += 1
                    
                    # Special attention to 24:xx:xx patterns
                    if value.startswith('24:'):
                        twenty_four_patterns[value] += 1
                        if value not in sample_rows:
                            sample_rows[value] = f"Row {row_num}, {col_name}"
                    
                    # Check for midnight patterns (0:00:00 variants)
                    if value in ['0:00:00', '00:00:00', '24:00:00']:
                        midnight_patterns[value] += 1
                        if len(sample_rows) < 20:  # Collect samples
                            key = f"{value}_sample"
                            if key not in sample_rows:
                                sample_rows[key] = f"Row {row_num}, {col_name}: '{value}'"
            
            # Progress indicator for large files
            if total_rows % 10000 == 0:
                print(f"  📈 Processed {total_rows:,} rows...")
        
        workbook.close()
        
        print(f"✅ Scan complete: {total_rows:,} total rows processed")
        
        # Display results
        print(f"\n📊 TIME PATTERN ANALYSIS:")
        print(f"   Total unique time patterns found: {len(time_patterns)}")
        
        if twenty_four_patterns:
            print(f"\n🎯 24:xx:xx PATTERNS FOUND:")
            for pattern, count in twenty_four_patterns.most_common(10):
                sample = sample_rows.get(pattern, "No sample")
                print(f"   '{pattern}': {count:,} occurrences (e.g., {sample})")
        else:
            print(f"\n❌ NO 24:xx:xx patterns found in entire file")
        
        print(f"\n🌙 MIDNIGHT PATTERNS:")
        for pattern, count in midnight_patterns.most_common():
            print(f"   '{pattern}': {count:,} occurrences")
        
        print(f"\n🕐 TOP 15 TIME PATTERNS:")
        for pattern, count in time_patterns.most_common(15):
            percentage = (count / sum(time_patterns.values())) * 100
            print(f"   '{pattern}': {count:,} ({percentage:.1f}%)")
        
        # Show samples if we have them
        if sample_rows:
            print(f"\n📋 SAMPLE LOCATIONS:")
            for key, sample in list(sample_rows.items())[:10]:
                if not key.endswith('_sample'):
                    print(f"   {sample}")
        
    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        return
    
    # Step 2: Compare with database (if provided)
    if db_path and Path(db_path).exists():
        print(f"\n🗄️  STEP 2: Database Comparison")
        print("-" * 50)
        
        try:
            conn = sqlite3.connect(db_path)
            
            # Get time pattern distribution from database
            cursor = conn.execute("""
                SELECT time_out, COUNT(*) as count 
                FROM spots 
                WHERE time_out IS NOT NULL 
                GROUP BY time_out 
                ORDER BY count DESC 
                LIMIT 15
            """)
            
            db_time_patterns = cursor.fetchall()
            
            print(f"📊 Database time_out patterns:")
            for pattern, count in db_time_patterns:
                print(f"   '{pattern}': {count:,} occurrences")
            
            # Check for specific patterns
            cursor = conn.execute("SELECT COUNT(*) FROM spots WHERE time_out = '24:00:00'")
            db_24_count = cursor.fetchone()[0]
            
            cursor = conn.execute("SELECT COUNT(*) FROM spots WHERE time_out = '0:00:00'")
            db_00_count = cursor.fetchone()[0]
            
            print(f"\n🎯 Key pattern comparison:")
            excel_24 = twenty_four_patterns.get('24:00:00', 0)
            excel_00 = midnight_patterns.get('0:00:00', 0) + midnight_patterns.get('00:00:00', 0)
            
            print(f"   24:00:00 - Excel: {excel_24:,}, Database: {db_24_count:,}")
            print(f"   0:00:00 - Excel: {excel_00:,}, Database: {db_00_count:,}")
            
            if excel_24 > 0 and db_24_count == 0 and db_00_count > excel_00:
                print(f"   🚨 CONVERSION DETECTED: Excel 24:00:00 → Database 0:00:00")
            
            conn.close()
            
        except Exception as e:
            print(f"❌ Error reading database: {e}")
    
    # Step 3: File format investigation
    print(f"\n🔬 STEP 3: File Format Analysis")
    print("-" * 50)
    
    try:
        # Check with data_only=True to see Excel's conversion
        workbook_processed = load_workbook(excel_file_path, data_only=True)
        worksheet_processed = workbook_processed.active
        
        processed_patterns = Counter()
        
        # Sample first 1000 rows with data_only=True
        for row_num, row in enumerate(worksheet_processed.iter_rows(min_row=2, max_row=1002, values_only=True), start=2):
            if not any(cell for cell in row):
                continue
                
            for col_idx, col_name in time_columns.items():
                if col_idx < len(row) and row[col_idx]:
                    value = str(row[col_idx]).strip()
                    processed_patterns[value] += 1
        
        workbook_processed.close()
        
        print(f"📊 Time patterns after Excel processing (data_only=True, first 1000 rows):")
        for pattern, count in processed_patterns.most_common(10):
            print(f"   '{pattern}': {count:,} occurrences")
        
        # Compare raw vs processed
        if '24:00:00' in time_patterns and '24:00:00' not in processed_patterns:
            print(f"\n🚨 CONFIRMED: Excel converts 24:00:00 → other format when data_only=True")
        
    except Exception as e:
        print(f"❌ Error in format analysis: {e}")
    
    print(f"\n💡 ENHANCED RECOMMENDATIONS:")
    print("-" * 50)
    
    if twenty_four_patterns:
        print(f"✅ Found {sum(twenty_four_patterns.values()):,} instances of 24:xx:xx patterns")
        print(f"   Most common: {twenty_four_patterns.most_common(1)[0] if twenty_four_patterns else 'None'}")
        print(f"   🔧 Action: Implement enhanced time conversion to preserve these values")
    else:
        print(f"❌ No 24:xx:xx patterns found in this Excel file")
        print(f"   🔍 Suggestions:")
        print(f"     • Check other Excel files that may have been imported")
        print(f"     • Verify the source of database 0:00:00 values")
        print(f"     • Check if 24:00:00 values were converted to other formats")


def main():
    import argparse
    
    parser = argparse.ArgumentParser(description="Enhanced time pattern diagnosis")
    parser.add_argument("excel_file", help="Path to Excel file")
    parser.add_argument("--database", help="Optional database path for comparison")
    
    args = parser.parse_args()
    
    enhanced_time_diagnosis(args.excel_file, args.database)


if __name__ == "__main__":
    main()