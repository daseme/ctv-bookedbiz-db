#!/usr/bin/env python3
"""
Simple, self-contained import test that avoids complex relative imports.
Tests the complete flow: Excel → Database with minimal dependencies.
"""

import sys
import sqlite3
import logging
from pathlib import Path
from datetime import datetime, date
from decimal import Decimal
from typing import Optional, List

# Add src to path
sys.path.insert(0, str(Path(__file__).parent / "src"))

# Import only what we absolutely need, handling import issues gracefully
try:
    from openpyxl import load_workbook
    print("✓ openpyxl available")
except ImportError:
    print("❌ openpyxl not available. Run: uv add openpyxl")
    sys.exit(1)

# Simple data classes (avoiding complex imports)
class SimpleSpot:
    def __init__(self, bill_code, air_date, gross_rate=None, sales_person=None, market_name=None, revenue_type=None):
        self.bill_code = bill_code
        self.air_date = air_date
        self.gross_rate = gross_rate
        self.sales_person = sales_person
        self.market_name = market_name
        self.revenue_type = revenue_type
        self.customer_id = None
        self.agency_id = None
        self.market_id = None

def parse_bill_code(bill_code: str) -> tuple[Optional[str], str]:
    """Simple bill code parser."""
    if not bill_code or not bill_code.strip():
        raise ValueError("Empty bill code")
    
    bill_code = bill_code.strip()
    
    if ':' in bill_code:
        parts = bill_code.split(':', 1)
        agency = parts[0].strip()
        customer = parts[1].strip()
        
        # Remove PRODUCTION suffix
        if customer.endswith(' PRODUCTION'):
            customer = customer[:-len(' PRODUCTION')]
        elif customer.endswith(' Production'):
            customer = customer[:-len(' Production')]
        elif customer.endswith(' PROD'):
            customer = customer[:-len(' PROD')]
        
        return (agency, customer)
    else:
        customer = bill_code
        # Remove PRODUCTION suffix
        if customer.endswith(' PRODUCTION'):
            customer = customer[:-len(' PRODUCTION')]
        elif customer.endswith(' Production'):
            customer = customer[:-len(' Production')]
        elif customer.endswith(' PROD'):
            customer = customer[:-len(' PROD')]
        
        return (None, customer)

def read_excel_simple(file_path: str, limit: Optional[int] = None) -> List[SimpleSpot]:
    """Simple Excel reader that gets just what we need."""
    print(f"Reading Excel file: {file_path}")
    
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook.active
    
    # Get headers
    header_row = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    
    # Find column indexes
    col_mapping = {}
    for i, header in enumerate(header_row):
        if header:
            header = str(header).strip()
            if header == 'Bill Code':
                col_mapping['bill_code'] = i
            elif header == 'Start Date':
                col_mapping['air_date'] = i
            elif header == 'Unit rate Gross':
                col_mapping['gross_rate'] = i
            elif header == 'Sales Person':
                col_mapping['sales_person'] = i
            elif header == 'Market':
                col_mapping['market_name'] = i
            elif header == 'Revenue Type':
                col_mapping['revenue_type'] = i
    
    print(f"Found columns: {list(col_mapping.keys())}")
    
    # Read spots
    spots = []
    for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if limit and len(spots) >= limit:
            break
        
        if not any(cell for cell in row):
            continue
        
        try:
            # Get required fields
            bill_code = row[col_mapping['bill_code']] if 'bill_code' in col_mapping else None
            air_date = row[col_mapping['air_date']] if 'air_date' in col_mapping else None
            
            if not bill_code or not air_date:
                continue
            
            # Convert air_date to date object if needed
            if isinstance(air_date, datetime):
                air_date = air_date.date()
            
            # Get optional fields
            gross_rate = row[col_mapping['gross_rate']] if 'gross_rate' in col_mapping and col_mapping['gross_rate'] < len(row) else None
            sales_person = row[col_mapping['sales_person']] if 'sales_person' in col_mapping and col_mapping['sales_person'] < len(row) else None
            market_name = row[col_mapping['market_name']] if 'market_name' in col_mapping and col_mapping['market_name'] < len(row) else None
            revenue_type = row[col_mapping['revenue_type']] if 'revenue_type' in col_mapping and col_mapping['revenue_type'] < len(row) else None
            
            # Convert gross_rate to Decimal if it's a number
            if isinstance(gross_rate, (int, float)):
                gross_rate = float(gross_rate)  # SQLite works with float, not Decimal
            
            spot = SimpleSpot(
                bill_code=str(bill_code),
                air_date=air_date,
                gross_rate=gross_rate,
                sales_person=str(sales_person) if sales_person else None,
                market_name=str(market_name) if market_name else None,
                revenue_type=str(revenue_type) if revenue_type else None
            )
            
            spots.append(spot)
            
        except Exception as e:
            print(f"Warning: Error reading row {row_num}: {e}")
            continue
    
    workbook.close()
    print(f"Read {len(spots)} spots from Excel")
    return spots

def get_or_create_agency(db_conn, agency_name: str) -> int:
    """Get or create agency, return agency_id."""
    cursor = db_conn.execute("SELECT agency_id FROM agencies WHERE agency_name = ?", (agency_name,))
    row = cursor.fetchone()
    
    if row:
        return row[0]
    
    # Create new agency
    cursor = db_conn.execute("INSERT INTO agencies (agency_name) VALUES (?)", (agency_name,))
    return cursor.lastrowid

def get_or_create_customer(db_conn, customer_name: str) -> int:
    """Get or create customer, return customer_id."""
    cursor = db_conn.execute("SELECT customer_id FROM customers WHERE normalized_name = ?", (customer_name,))
    row = cursor.fetchone()
    
    if row:
        return row[0]
    
    # Create new customer
    cursor = db_conn.execute("INSERT INTO customers (normalized_name) VALUES (?)", (customer_name,))
    return cursor.lastrowid

def get_market_id(db_conn, market_name: str) -> Optional[int]:
    """Get market_id for market name."""
    cursor = db_conn.execute("SELECT market_id FROM markets WHERE LOWER(market_name) = LOWER(?)", (market_name,))
    row = cursor.fetchone()
    return row[0] if row else None

def test_import(excel_file: str, database_file: str, limit: int = 10):
    """Test the complete import process."""
    print(f"Testing import: {limit} spots from {excel_file} to {database_file}")
    print("=" * 60)
    
    # Check files exist
    if not Path(excel_file).exists():
        print(f"❌ Excel file not found: {excel_file}")
        return False
    
    if not Path(database_file).exists():
        print(f"❌ Database not found: {database_file}")
        print("Run: python scripts/setup_database.py --db-path {database_file}")
        return False
    
    # Step 1: Read Excel
    spots = read_excel_simple(excel_file, limit)
    if not spots:
        print("❌ No spots read from Excel")
        return False
    
    # Step 2: Process and import
    db_conn = sqlite3.connect(database_file)
    db_conn.execute("PRAGMA foreign_keys = ON")
    
    stats = {
        'spots_processed': 0,
        'spots_imported': 0,
        'new_agencies': 0,
        'new_customers': 0,
        'customers_normalized': 0,
        'errors': [],  # Track detailed errors
        'error_types': {}  # Count error types
    }
    
    try:
        db_conn.execute("BEGIN")
        
        for i, spot in enumerate(spots):
            print(f"\nProcessing spot {i+1}: {spot.bill_code}")
            
            try:
                # Parse bill code
                agency_name, customer_name = parse_bill_code(spot.bill_code)
                print(f"  Parsed → Agency: {agency_name}, Customer: {customer_name}")
                
                # Track normalization
                if spot.bill_code != (f"{agency_name}:{customer_name}" if agency_name else customer_name):
                    stats['customers_normalized'] += 1
                    print(f"  Normalized customer name (removed PRODUCTION)")
                
                # Handle agency
                if agency_name:
                    # Check if agency exists
                    cursor = db_conn.execute("SELECT agency_id FROM agencies WHERE agency_name = ?", (agency_name,))
                    if not cursor.fetchone():
                        stats['new_agencies'] += 1
                        print(f"  Creating new agency: {agency_name}")
                    
                    spot.agency_id = get_or_create_agency(db_conn, agency_name)
                
                # Handle customer
                cursor = db_conn.execute("SELECT customer_id FROM customers WHERE normalized_name = ?", (customer_name,))
                if not cursor.fetchone():
                    stats['new_customers'] += 1
                    print(f"  Creating new customer: {customer_name}")
                
                spot.customer_id = get_or_create_customer(db_conn, customer_name)
                
                # Handle market
                if spot.market_name:
                    spot.market_id = get_market_id(db_conn, spot.market_name)
                    if spot.market_id:
                        print(f"  Mapped market: {spot.market_name}")
                
                # Insert spot (allow negative gross_rate for broker fees/adjustments)
                db_conn.execute("""
                INSERT INTO spots (
                    bill_code, air_date, gross_rate, sales_person, market_name, revenue_type,
                    customer_id, agency_id, market_id, source_file
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    spot.bill_code, spot.air_date, spot.gross_rate, spot.sales_person, 
                    spot.market_name, spot.revenue_type, spot.customer_id, spot.agency_id, 
                    spot.market_id, Path(excel_file).name
                ))
                
                stats['spots_imported'] += 1
                stats['spots_processed'] += 1
                print(f"  ✓ Imported successfully")
                
            except Exception as e:
                error_msg = f"Row {i+1}: {str(e)}"
                print(f"  ✗ Error: {e}")
                stats['spots_processed'] += 1
                stats['errors'].append(error_msg)
                
                # Count error types
                error_type = type(e).__name__
                stats['error_types'][error_type] = stats['error_types'].get(error_type, 0) + 1
        
        db_conn.commit()
        print(f"\n✅ Transaction committed successfully")
        
    except Exception as e:
        db_conn.rollback()
        print(f"\n❌ Transaction rolled back: {e}")
        return False
    finally:
        db_conn.close()
    
    # Print results
    spots_skipped = stats['spots_processed'] - stats['spots_imported']
    print(f"\nImport Results:")
    print(f"  Spots processed: {stats['spots_processed']}")
    print(f"  Spots imported: {stats['spots_imported']}")
    print(f"  Spots skipped: {spots_skipped}")
    print(f"  New agencies created: {stats['new_agencies']}")
    print(f"  New customers created: {stats['new_customers']}")
    print(f"  Customers normalized: {stats['customers_normalized']}")
    
    if spots_skipped > 0:
        print(f"\n⚠️  Analysis of {spots_skipped} skipped spots:")
        
        # Show error type breakdown
        if stats['error_types']:
            print(f"  Error types:")
            for error_type, count in sorted(stats['error_types'].items(), key=lambda x: x[1], reverse=True):
                percentage = (count / spots_skipped) * 100
                print(f"    {error_type}: {count} spots ({percentage:.1f}%)")
        
        # Show first few detailed errors as examples
        print(f"\n  Sample errors:")
        for i, error in enumerate(stats['errors'][:10]):  # Show first 10 errors
            print(f"    {error}")
        
        if len(stats['errors']) > 10:
            print(f"    ... and {len(stats['errors']) - 10} more errors")
        
        print(f"\n  💡 Common causes:")
        print(f"    - Empty/invalid bill codes")
        print(f"    - Missing air dates") 
        print(f"    - Malformed data in Excel rows")
        print(f"    - Date conversion issues")
    
    return stats['spots_imported'] > 0

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Simple import test")
    parser.add_argument("excel_file", help="Path to Excel file")
    parser.add_argument("--database", default="data/database/test.db", help="Database path")
    parser.add_argument("--limit", type=int, default=None, help="Number of spots to import (default: import all)")
    
    args = parser.parse_args()
    
    success = test_import(args.excel_file, args.database, args.limit)
    
    if success:
        print(f"\n🎉 Import test completed successfully!")
    else:
        print(f"\n💥 Import test failed!")
        sys.exit(1)