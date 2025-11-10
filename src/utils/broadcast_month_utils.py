#!/usr/bin/env python3
"""
Broadcast month utilities for handling Excel date format conversion and month extraction.
Converts Excel dates (11/15/2024) to broadcast month format (Nov-24).
Enhanced with clean output, tqdm progress bars, and suppression support.
"""

import sys
import logging
from pathlib import Path
from datetime import datetime, date
from typing import Set, List, Optional, Union
import re

# Add src to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

try:
    from openpyxl import load_workbook
except ImportError:
    print("❌ openpyxl not available. Run: uv add openpyxl")
    sys.exit(1)

try:
    from tqdm import tqdm

    TQDM_AVAILABLE = True
except ImportError:
    TQDM_AVAILABLE = False

    # Fallback class for when tqdm is not available
    class tqdm:
        def __init__(
            self, iterable=None, total=None, desc=None, disable=False, **kwargs
        ):
            self.iterable = iterable
            self.total = total
            self.desc = desc
            self.disable = disable
            self.n = 0

        def __iter__(self):
            for item in self.iterable:
                yield item
                self.n += 1

        def __enter__(self):
            return self

        def __exit__(self, *args):
            pass

        def update(self, n=1):
            self.n += n

        def set_description(self, desc):
            self.desc = desc


logger = logging.getLogger(__name__)


class BroadcastMonthParseError(Exception):
    """Raised when there's an error parsing broadcast month data."""

    pass


class BroadcastMonthParser:
    """Handles conversion between Excel dates and broadcast month format."""

    # Month name mapping for consistent format
    MONTH_NAMES = {
        1: "Jan",
        2: "Feb",
        3: "Mar",
        4: "Apr",
        5: "May",
        6: "Jun",
        7: "Jul",
        8: "Aug",
        9: "Sep",
        10: "Oct",
        11: "Nov",
        12: "Dec",
    }

    # Reverse mapping for parsing
    MONTH_NAME_TO_NUM = {name: num for num, name in MONTH_NAMES.items()}

    def __init__(self, verbose: bool = False):
        """Initialize the broadcast month parser."""
        self.verbose = verbose  # Control verbose output
        self.parse_stats = {
            "total_parsed": 0,
            "successful_conversions": 0,
            "errors": 0,
            "unique_months_found": set(),
        }

    def parse_excel_date_to_broadcast_month(
        self, excel_date: Union[datetime, date, str]
    ) -> str:
        """
        Convert Excel date to broadcast month format.

        Examples:
        - datetime(2024, 11, 15) -> 'Nov-24'
        - date(2024, 5, 1) -> 'May-24'
        - '11/15/2024' -> 'Nov-24'
        - '2024-11-15' -> 'Nov-24'

        Args:
            excel_date: Date from Excel (datetime, date, or string)

        Returns:
            Broadcast month in 'Mmm-YY' format

        Raises:
            BroadcastMonthParseError: If date cannot be parsed
        """
        self.parse_stats["total_parsed"] += 1

        try:
            # Handle datetime and date objects
            if isinstance(excel_date, (datetime, date)):
                target_date = excel_date

            # Handle string dates
            elif isinstance(excel_date, str):
                target_date = self._parse_date_string(excel_date.strip())

            # Handle None or other types
            else:
                raise BroadcastMonthParseError(
                    f"Unsupported date type: {type(excel_date)}"
                )

            # Convert to broadcast month format
            month_name = self.MONTH_NAMES[target_date.month]
            year_suffix = str(target_date.year)[2:]  # Last 2 digits
            broadcast_month = f"{month_name}-{year_suffix}"

            self.parse_stats["successful_conversions"] += 1
            self.parse_stats["unique_months_found"].add(broadcast_month)

            # Only log debug info if explicitly requested and at debug level
            if self.verbose and logger.isEnabledFor(logging.DEBUG):
                logger.debug(f"Converted {excel_date} -> {broadcast_month}")

            return broadcast_month

        except Exception as e:
            self.parse_stats["errors"] += 1
            error_msg = f"Failed to convert '{excel_date}' to broadcast month: {str(e)}"

            # Only log warnings if explicitly verbose and at warning level
            if self.verbose and logger.isEnabledFor(logging.WARNING):
                logger.warning(error_msg)

            raise BroadcastMonthParseError(error_msg)

    def _parse_date_string(self, date_str: str) -> date:
        """Parse various string date formats to date object."""
        if not date_str:
            raise ValueError("Empty date string")

        # Try common date formats
        formats_to_try = [
            "%Y-%m-%d %H:%M:%S",  # Handle datetime strings like '2024-11-15 00:00:00'
            "%Y-%m-%d",  # 2024-11-15
            "%m/%d/%Y",  # 11/15/2024
            "%m/%d/%y",  # 11/15/24
            "%d/%m/%Y",  # 15/11/2024 (international)
            "%Y/%m/%d",  # 2024/11/15
            "%m-%d-%Y",  # 11-15-2024
            "%m-%d-%y",  # 11-15-24
        ]

        for fmt in formats_to_try:
            try:
                parsed_date = datetime.strptime(date_str, fmt).date()

                # Only log successful parsing if verbose and debug level
                if self.verbose and logger.isEnabledFor(logging.DEBUG):
                    logger.debug(f"Parsed '{date_str}' using format '{fmt}'")

                return parsed_date
            except ValueError:
                continue

        # If all formats fail
        raise ValueError(
            f"Could not parse date string '{date_str}' with any known format"
        )

    def validate_broadcast_month_format(self, broadcast_month: str) -> bool:
        """
        Validate if a string matches broadcast month format.

        Args:
            broadcast_month: String to validate

        Returns:
            True if valid format (Mmm-YY), False otherwise
        """
        if not isinstance(broadcast_month, str):
            return False

        # Pattern: 3-letter month name, dash, 2-digit year
        pattern = r"^[A-Z][a-z]{2}-\d{2}$"

        if not re.match(pattern, broadcast_month):
            return False

        # Check if month name is valid
        month_part = broadcast_month.split("-")[0]
        return month_part in self.MONTH_NAME_TO_NUM

    def extract_year_from_broadcast_month(self, broadcast_month: str) -> int:
        """Extract 4-digit year from broadcast month."""
        if not self.validate_broadcast_month_format(broadcast_month):
            raise BroadcastMonthParseError(
                f"Invalid broadcast month format: '{broadcast_month}'"
            )

        year_suffix = broadcast_month.split("-")[1]
        year_num = int(year_suffix)

        # Convert 2-digit year to 4-digit year
        if year_num <= 30:
            return 2000 + year_num
        else:
            return 1900 + year_num

    def get_broadcast_months_in_year(self, year: int) -> List[str]:
        """Get all 12 broadcast months for a given year."""
        year_suffix = str(year)[2:]  # Last 2 digits
        return [
            f"{month_name}-{year_suffix}" for month_name in self.MONTH_NAMES.values()
        ]

    def format_broadcast_month_for_display(self, broadcast_month: str) -> str:
        """Format broadcast month for user-friendly display."""
        if not self.validate_broadcast_month_format(broadcast_month):
            return broadcast_month  # Return as-is if invalid

        month_part, year_part = broadcast_month.split("-")

        # Full month names
        full_month_names = {
            "Jan": "January",
            "Feb": "February",
            "Mar": "March",
            "Apr": "April",
            "May": "May",
            "Jun": "June",
            "Jul": "July",
            "Aug": "August",
            "Sep": "September",
            "Oct": "October",
            "Nov": "November",
            "Dec": "December",
        }

        full_month = full_month_names.get(month_part, month_part)
        full_year = self.extract_year_from_broadcast_month(broadcast_month)

        return f"{full_month} {full_year}"

    def get_statistics(self) -> dict:
        """Get parsing statistics."""
        stats = self.parse_stats.copy()
        stats["unique_months_found"] = sorted(list(stats["unique_months_found"]))
        if stats["total_parsed"] > 0:
            stats["success_rate"] = (
                stats["successful_conversions"] / stats["total_parsed"]
            ) * 100
        else:
            stats["success_rate"] = 0
        return stats

    def reset_statistics(self):
        """Reset parsing statistics."""
        self.parse_stats = {
            "total_parsed": 0,
            "successful_conversions": 0,
            "errors": 0,
            "unique_months_found": set(),
        }


def extract_broadcast_months_from_excel(
    excel_file_path: str,
    month_column: str = "Month",
    limit: Optional[int] = None,
    return_display_format: bool = True,
    verbose: bool = False,
    show_progress: bool = True,
) -> Set[str]:
    """
    Extract all unique broadcast months from an Excel file.

    Args:
        excel_file_path: Path to Excel file
        month_column: Name of column containing month data (default: 'Month')
        limit: Optional limit on rows to process (for testing)
        return_display_format: If True, return 'Nov-24' format; if False, return datetime format
        verbose: If True, show detailed messages; if False, work more quietly
        show_progress: If True, show progress bar during processing

    Returns:
        Set of unique broadcast months found in file

    Raises:
        BroadcastMonthParseError: If file cannot be read or processed
    """
    if verbose:
        print(f"📊 Processing Excel file: {Path(excel_file_path).name}")

    if not Path(excel_file_path).exists():
        raise BroadcastMonthParseError(f"Excel file not found: {excel_file_path}")

    parser = BroadcastMonthParser(verbose=verbose)
    broadcast_months = set()

    try:
        workbook = load_workbook(excel_file_path, read_only=True, data_only=True)
        worksheet = workbook['Data']

        # Find the month column
        header_row = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[
            0
        ]
        month_col_index = None

        for i, header in enumerate(header_row):
            if header and str(header).strip() == month_column:
                month_col_index = i
                break

        if month_col_index is None:
            raise BroadcastMonthParseError(
                f"Column '{month_column}' not found in Excel file"
            )

        if verbose:
            print(f"✓ Found '{month_column}' column at position {month_col_index + 1}")

        # Get total row count for progress bar
        total_rows = worksheet.max_row - 1  # Subtract header row
        if limit:
            total_rows = min(total_rows, limit)

        # Process rows with progress bar
        processed_count = 0
        error_count = 0

        # Create progress bar description
        desc = f"Processing {month_column} data"
        if limit:
            desc += f" (limited to {limit} rows)"

        # Set up progress bar
        disable_progress = not show_progress or not TQDM_AVAILABLE or verbose

        with tqdm(
            total=total_rows, desc=desc, disable=disable_progress, unit="rows", ncols=80
        ) as pbar:
            for row_num, row in enumerate(
                worksheet.iter_rows(min_row=2, values_only=True), start=2
            ):
                if limit and processed_count >= limit:
                    break

                if not any(cell for cell in row):
                    pbar.update(1)
                    continue  # Skip empty rows

                if month_col_index < len(row):
                    month_value = row[month_col_index]

                    if month_value is not None:
                        try:
                            if return_display_format:
                                broadcast_month = (
                                    parser.parse_excel_date_to_broadcast_month(
                                        month_value
                                    )
                                )
                            else:
                                # Return the raw datetime value (for database compatibility)
                                if isinstance(month_value, str):
                                    broadcast_month = month_value
                                else:
                                    broadcast_month = str(month_value)

                            broadcast_months.add(broadcast_month)
                            processed_count += 1

                            # Update progress bar description with current count
                            if (
                                processed_count % 100 == 0
                            ):  # Update every 100 rows to avoid too frequent updates
                                pbar.set_description(
                                    f"{desc} - {len(broadcast_months)} unique months found"
                                )

                        except BroadcastMonthParseError as e:
                            error_count += 1
                            if verbose and logger.isEnabledFor(logging.WARNING):
                                logger.warning(f"Row {row_num}: {e}")

                pbar.update(1)

        workbook.close()

        # Final summary - only show if verbose or if there were issues
        stats = parser.get_statistics()

        if verbose or error_count > 0:
            success_rate = stats["success_rate"]
            print(f"✅ Processing complete:")
            print(f"   • Processed: {stats['total_parsed']} month values")
            print(f"   • Success rate: {success_rate:.1f}%")
            print(f"   • Unique months found: {len(broadcast_months)}")

            if error_count > 0:
                print(f"   ⚠️  Parsing errors: {error_count}")
        elif show_progress and TQDM_AVAILABLE:
            # Just a simple completion message if not verbose
            print(f"✅ Found {len(broadcast_months)} unique broadcast months")

        if verbose and return_display_format and broadcast_months:
            print(f"📅 Months: {', '.join(sorted(broadcast_months))}")

        return broadcast_months

    except Exception as e:
        raise BroadcastMonthParseError(f"Failed to process Excel file: {str(e)}")


def validate_broadcast_months_for_year(
    broadcast_months: Set[str], expected_year: int
) -> tuple[Set[str], Set[str]]:
    """
    Validate that broadcast months belong to expected year.

    Args:
        broadcast_months: Set of broadcast months to validate
        expected_year: Expected 4-digit year

    Returns:
        Tuple of (matching_months, mismatched_months)
    """
    parser = BroadcastMonthParser(verbose=False)  # Non-verbose for utility function
    matching_months = set()
    mismatched_months = set()

    for broadcast_month in broadcast_months:
        try:
            year = parser.extract_year_from_broadcast_month(broadcast_month)
            if year == expected_year:
                matching_months.add(broadcast_month)
            else:
                mismatched_months.add(broadcast_month)
        except BroadcastMonthParseError:
            mismatched_months.add(broadcast_month)  # Invalid format

    return matching_months, mismatched_months


# Convenience functions for simple usage
def parse_excel_date(
    excel_date: Union[datetime, date, str], verbose: bool = False
) -> str:
    """Simple function to parse a single Excel date."""
    parser = BroadcastMonthParser(verbose=verbose)
    return parser.parse_excel_date_to_broadcast_month(excel_date)


def is_valid_broadcast_month(broadcast_month: str) -> bool:
    """Simple function to validate broadcast month format."""
    parser = BroadcastMonthParser(verbose=False)
    return parser.validate_broadcast_month_format(broadcast_month)


def normalize_broadcast_day(dt: Union[datetime, str]) -> Union[datetime, str]:
    """
    Normalize broadcast date to either 1st or 15th.
    Enhanced to handle both datetime objects and strings.

    • Day 1–15 → 1st
    • Day 16–31 → 15th
    """
    if isinstance(dt, str):
        # Return string as-is if it's already a string
        return dt
    elif isinstance(dt, datetime):
        return dt.replace(day=1) if dt.day <= 15 else dt.replace(day=15)
    else:
        # Handle other types gracefully
        return dt


# Test and example usage
if __name__ == "__main__":
    import argparse

    parser_cli = argparse.ArgumentParser(description="Test broadcast month utilities")
    parser_cli.add_argument(
        "--test-dates", action="store_true", help="Test date conversion"
    )
    parser_cli.add_argument(
        "--extract-from-excel", help="Excel file to extract months from"
    )
    parser_cli.add_argument(
        "--limit", type=int, help="Limit rows processed (for testing)"
    )
    parser_cli.add_argument(
        "--verbose", action="store_true", help="Enable verbose logging"
    )
    parser_cli.add_argument(
        "--no-progress", action="store_true", help="Disable progress bar"
    )

    args = parser_cli.parse_args()

    # Setup logging
    level = (
        logging.DEBUG if args.verbose else logging.WARNING
    )  # Changed default to WARNING to reduce noise
    logging.basicConfig(level=level, format="%(levelname)s: %(message)s")

    if not TQDM_AVAILABLE:
        print(
            "⚠️  tqdm not available. Install with: pip install tqdm (progress bars will be disabled)"
        )

    if args.test_dates:
        print("🧪 Testing Broadcast Month Parser:")
        print("=" * 50)

        parser = BroadcastMonthParser(verbose=args.verbose)

        # Test various date formats
        test_dates = [
            datetime(2024, 11, 15),  # datetime object
            date(2024, 5, 1),  # date object
            "11/15/2024",  # MM/DD/YYYY
            "5/1/24",  # M/D/YY
            "2024-11-15",  # ISO format
            "invalid_date",  # Should fail
            None,  # Should fail
        ]

        # Process with progress bar if available and not disabled
        test_iter = tqdm(
            test_dates,
            desc="Testing date formats",
            disable=args.no_progress or not TQDM_AVAILABLE,
        )

        for test_date in test_iter:
            try:
                result = parser.parse_excel_date_to_broadcast_month(test_date)
                print(f"✓ {test_date} -> {result}")
            except BroadcastMonthParseError as e:
                print(f"✗ {test_date} -> Error: {e}")

        print(f"\n📊 Test Statistics:")
        stats = parser.get_statistics()
        for key, value in stats.items():
            if key != "unique_months_found":  # Skip the set for cleaner output
                print(f"   {key}: {value}")

    if args.extract_from_excel:
        try:
            broadcast_months = extract_broadcast_months_from_excel(
                args.extract_from_excel,
                limit=args.limit,
                verbose=args.verbose,
                show_progress=not args.no_progress,
            )

            if broadcast_months:
                if not args.verbose:  # Only show summary if not already shown
                    # Group by year for display
                    years = {}
                    parser = BroadcastMonthParser(verbose=False)
                    for month in sorted(broadcast_months):
                        try:
                            year = parser.extract_year_from_broadcast_month(month)
                            if year not in years:
                                years[year] = []
                            years[year].append(month)
                        except BroadcastMonthParseError:
                            continue

                    print(f"\n📅 Summary by year:")
                    for year in sorted(years.keys()):
                        print(f"   {year}: {', '.join(sorted(years[year]))}")
            else:
                print("❌ No broadcast months found in file")

        except BroadcastMonthParseError as e:
            print(f"❌ Error: {e}")
            sys.exit(1)

    if not args.test_dates and not args.extract_from_excel:
        print("📖 Usage Examples:")
        print("  python broadcast_month_utils.py --test-dates")
        print("  python broadcast_month_utils.py --extract-from-excel data/sample.xlsx")
        print(
            "  python broadcast_month_utils.py --extract-from-excel data/sample.xlsx --limit 1000 --verbose"
        )
        print(
            "  python broadcast_month_utils.py --extract-from-excel data/sample.xlsx --no-progress"
        )
        if not TQDM_AVAILABLE:
            print("\n💡 For better progress tracking, install tqdm: pip install tqdm")
