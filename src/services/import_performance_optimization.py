#!/usr/bin/env python3
"""
Import Performance Optimization - Fixed Version
No long transactions to avoid database locks.
"""

import sys
import logging
from pathlib import Path
from typing import Dict, Set, Optional
from dataclasses import dataclass

# Add src to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from src.database.connection import DatabaseConnection

logger = logging.getLogger(__name__)

@dataclass
class EntityLookupResult:
    """Result of entity lookup with performance tracking"""
    customer_id: Optional[int]
    agency_id: Optional[int]
    used_cache: bool = False
    lookup_method: str = "unknown"

class BatchEntityResolver:
    """Fixed entity resolver - avoids database locks."""
    
    def __init__(self, db_connection: DatabaseConnection):
        self.db = db_connection
        self.entity_cache: Dict[str, EntityLookupResult] = {}
        self.cache_stats = {
            'cache_hits': 0,
            'cache_misses': 0, 
            'batch_resolved': 0,
            'individual_fallbacks': 0
        }
    
    def build_entity_cache_from_excel(self, excel_file: str) -> None:
        """Disabled cache pre-building to avoid database locks - performance still gained through cached lookups during import"""
        logger.info("Entity cache pre-building disabled to avoid database locks")
        logger.info("Performance improvements still active through cached lookups during import")
        
        # Set dummy stats so reporting doesn't break
        self.cache_stats['batch_resolved'] = 0
        
        # The entity_cache dict is still available for caching during individual lookups
        # This provides the main performance benefit without the database conflicts
        return
    
    def _extract_unique_bill_codes(self, excel_file: str) -> Set[str]:
        """Extract all unique bill_codes from Excel file efficiently"""
        from openpyxl import load_workbook
        
        unique_bill_codes = set()
        
        try:
            workbook = load_workbook(excel_file, read_only=True, data_only=True)
            try:
                worksheet = workbook.active
                
                # Find bill_code column (assuming position 0)
                bill_code_col_index = 0
                
                # Extract all bill_codes efficiently
                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    if len(row) > bill_code_col_index and row[bill_code_col_index]:
                        bill_code = str(row[bill_code_col_index]).strip()
                        if bill_code:
                            unique_bill_codes.add(bill_code)
            finally:
                workbook.close()
        
        except Exception as e:
            logger.warning(f"Failed to extract bill_codes from Excel: {e}")
            return set()
        
        return unique_bill_codes
    
    def _batch_resolve_entities(self, bill_codes: Set[str]) -> None:
        """Simple batch resolve without transactions to avoid database locks."""
        if not bill_codes:
            return
        
        logger.info(f"Batch resolving {len(bill_codes)} bill_codes")
        
        batch_resolved_count = 0
        
        # Use simple connection without transaction management
        try:
            conn = self.db.connect()
            
            for bill_code in bill_codes:
                try:
                    # Parse bill code parts
                    if ':' in bill_code:
                        agency_part = bill_code.split(':', 1)[0].strip()
                        customer_part = bill_code.split(':', 1)[1].strip()
                    else:
                        agency_part = None
                        customer_part = bill_code.strip()
                    
                    # Simple direct lookups (no transactions)
                    customer_id = self._lookup_customer_direct(customer_part, conn)
                    agency_id = self._lookup_agency_direct(agency_part, conn) if agency_part else None
                    
                    # Cache the result
                    self.entity_cache[bill_code] = EntityLookupResult(
                        customer_id=customer_id,
                        agency_id=agency_id,
                        used_cache=True,
                        lookup_method='batch_resolve'
                    )
                    
                    if customer_id or agency_id:
                        batch_resolved_count += 1
                        
                except Exception:
                    # Cache as unresolved
                    self.entity_cache[bill_code] = EntityLookupResult(
                        customer_id=None,
                        agency_id=None,
                        used_cache=True,
                        lookup_method='batch_resolve'
                    )
            
            conn.close()
            
        except Exception as e:
            logger.warning(f"Cache building failed: {e}")
        
        self.cache_stats['batch_resolved'] = batch_resolved_count
        logger.info(f"Batch resolved {batch_resolved_count}/{len(bill_codes)} entities")

    def _lookup_customer_direct(self, customer_name: str, conn) -> Optional[int]:
        """Direct customer lookup without transactions"""
        if not customer_name:
            return None
        
        try:
            # Direct lookup
            cursor = conn.execute("""
                SELECT customer_id FROM customers 
                WHERE normalized_name = ? AND is_active = 1
            """, (customer_name,))
            result = cursor.fetchone()
            if result:
                return result[0]
            
            # Alias lookup
            cursor = conn.execute("""
                SELECT ea.target_entity_id FROM entity_aliases ea
                JOIN customers c ON c.customer_id = ea.target_entity_id
                WHERE ea.alias_name = ? 
                AND ea.entity_type = 'customer' 
                AND ea.is_active = 1 
                AND c.is_active = 1
            """, (customer_name,))
            result = cursor.fetchone()
            return result[0] if result else None
        except:
            return None

    def _lookup_agency_direct(self, agency_name: str, conn) -> Optional[int]:
        """Direct agency lookup without transactions"""
        if not agency_name:
            return None
        
        try:
            # Direct lookup
            cursor = conn.execute("""
                SELECT agency_id FROM agencies 
                WHERE agency_name = ? AND is_active = 1
            """, (agency_name,))
            result = cursor.fetchone()
            if result:
                return result[0]
            
            # Alias lookup  
            cursor = conn.execute("""
                SELECT ea.target_entity_id FROM entity_aliases ea
                JOIN agencies a ON a.agency_id = ea.target_entity_id
                WHERE ea.alias_name = ? 
                AND ea.entity_type = 'agency' 
                AND ea.is_active = 1 
                AND a.is_active = 1
            """, (agency_name,))
            result = cursor.fetchone()
            return result[0] if result else None
        except:
            return None
    
    def _lookup_customer_quick(self, customer_name: str, conn) -> Optional[int]:
        """Quick customer lookup"""
        if not customer_name:
            return None
        
        # Try direct lookup first
        cursor = conn.execute("""
            SELECT customer_id FROM customers 
            WHERE normalized_name = ? AND is_active = 1
        """, (customer_name,))
        result = cursor.fetchone()
        if result:
            return result[0]
        
        # Try entity alias lookup
        cursor = conn.execute("""
            SELECT ea.target_entity_id FROM entity_aliases ea
            JOIN customers c ON c.customer_id = ea.target_entity_id
            WHERE ea.alias_name = ? 
              AND ea.entity_type = 'customer' 
              AND ea.is_active = 1 
              AND c.is_active = 1
        """, (customer_name,))
        result = cursor.fetchone()
        return result[0] if result else None
    
    def _lookup_agency_quick(self, agency_name: str, conn) -> Optional[int]:
        """Quick agency lookup"""
        if not agency_name:
            return None
        
        # Try direct lookup first
        cursor = conn.execute("""
            SELECT agency_id FROM agencies 
            WHERE agency_name = ? AND is_active = 1
        """, (agency_name,))
        result = cursor.fetchone()
        if result:
            return result[0]
        
        # Try entity alias lookup
        cursor = conn.execute("""
            SELECT ea.target_entity_id FROM entity_aliases ea
            JOIN agencies a ON a.agency_id = ea.target_entity_id
            WHERE ea.alias_name = ? 
              AND ea.entity_type = 'agency' 
              AND ea.is_active = 1 
              AND a.is_active = 1
        """, (agency_name,))
        result = cursor.fetchone()
        return result[0] if result else None
    
    def lookup_entities_cached(self, bill_code: str, conn=None) -> EntityLookupResult:
        """Cache-first entity lookup"""
        # Cache hit (fastest path)
        if bill_code in self.entity_cache:
            self.cache_stats['cache_hits'] += 1
            return self.entity_cache[bill_code]
        
        # Cache miss - individual fallback
        self.cache_stats['cache_misses'] += 1
        result = self._individual_entity_lookup(bill_code, conn)
        
        self.entity_cache[bill_code] = result
        return result
    
    def _individual_entity_lookup(self, bill_code: str, conn) -> EntityLookupResult:
        """Fallback individual entity lookup"""
        self.cache_stats['individual_fallbacks'] += 1
        
        result = EntityLookupResult(
            customer_id=None,
            agency_id=None,
            used_cache=False,
            lookup_method='individual_fallback'
        )
        
        if not bill_code:
            return result
        
        try:
            if ':' in bill_code:
                agency_part, customer_part = bill_code.split(':', 1)
                result.customer_id = self._lookup_customer_quick(customer_part.strip(), conn)
                result.agency_id = self._lookup_agency_quick(agency_part.strip(), conn)
            else:
                result.customer_id = self._lookup_customer_quick(bill_code.strip(), conn)
        
        except Exception as e:
            logger.warning(f"Individual lookup failed for {bill_code}: {e}")
        
        return result
    
    def get_performance_stats(self) -> Dict[str, int]:
        """Get performance statistics"""
        total_lookups = sum(self.cache_stats.values())
        cache_hit_rate = (self.cache_stats['cache_hits'] / total_lookups * 100) if total_lookups > 0 else 0
        
        return {
            **self.cache_stats,
            'total_lookups': total_lookups,
            'cache_hit_rate_percent': round(cache_hit_rate, 1),
            'cache_size': len(self.entity_cache)
        }
