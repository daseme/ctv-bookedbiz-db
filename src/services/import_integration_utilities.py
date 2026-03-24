#!/usr/bin/env python3
"""
Import integration utilities for Excel file processing.
Enhanced to support both daily commercial log format (Commercials sheet)
and monthly import format (Data sheet).

FIXED: Now handles flexible sheet detection instead of hardcoded "Data" sheet.
"""

import sys
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Generator
from dataclasses import dataclass

logger = logging.getLogger(__name__)

# Add src to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))


@dataclass
class ValidationResult:
    """Result of Excel import validation"""

    is_valid: bool
    error_message: str = ""
    suggested_action: str = ""
    closed_months_found: List[str] = None
    open_months_found: List[str] = None

    def __post_init__(self):
        if self.closed_months_found is None:
            self.closed_months_found = []
        if self.open_months_found is None:
            self.open_months_found = []


def get_excel_worksheet_flexible(excel_file_path: str):
    """
    ENHANCED: Flexible worksheet detection for both daily and monthly imports.

    Tries multiple sheet names in order of preference:
    1. "Data" (monthly import format)
    2. "Commercials" (daily commercial log format)
    3. "Commercial Lines" (alternate daily format)
    4. "Sheet1" (generic fallback)
    5. Active sheet (final fallback)

    Returns:
        tuple: (worksheet, sheet_name_used)
    """
    from openpyxl import load_workbook

    # Sheet names to try in order of preference
    sheet_candidates = [
        "Data",  # Monthly import format (existing)
        "Commercials",  # Daily commercial log format (new)
        "Commercial Lines",  # Alternate daily format
        "Commercial",  # Short form
        "Sheet1",  # Generic Excel default
    ]

    try:
        workbook = load_workbook(excel_file_path, read_only=True, data_only=True)
        available_sheets = workbook.sheetnames

        logger.debug(f"Available sheets in {excel_file_path}: {available_sheets}")

        # Try each candidate sheet name
        for sheet_name in sheet_candidates:
            if sheet_name in available_sheets:
                worksheet = workbook[sheet_name]
                logger.info(
                    f"Using sheet '{sheet_name}' from {Path(excel_file_path).name}"
                )
                return worksheet, sheet_name, workbook

        # Final fallback: use active sheet
        if available_sheets:
            worksheet = workbook.active
            sheet_name = (
                worksheet.title if hasattr(worksheet, "title") else available_sheets[0]
            )
            logger.info(
                f"Using active sheet '{sheet_name}' from {Path(excel_file_path).name}"
            )
            return worksheet, sheet_name, workbook

        # No sheets available
        workbook.close()
        raise ValueError(f"No readable sheets found in {excel_file_path}")

    except Exception as e:
        logger.error(f"Failed to open Excel file {excel_file_path}: {e}")
        raise Exception(f"Failed to process Excel file: {str(e)}")


# Sheets that contain importable spot data in the commercial log
IMPORT_SHEET_NAMES = [
    "Commercials",
    "Worldlink Lines",
    "Add to booked business",
    "Pending",
]


def get_all_import_worksheets(excel_file_path: str):
    """
    Return all importable worksheets from an Excel file.

    Yields (worksheet, sheet_name) tuples for every sheet in
    IMPORT_SHEET_NAMES that exists in the workbook. Falls back to
    get_excel_worksheet_flexible() if none of the known sheets are found
    (e.g. monthly "Data" sheet imports).

    The caller is responsible for closing the returned workbook.

    Returns:
        tuple: (list_of_(worksheet, sheet_name), workbook)
    """
    from openpyxl import load_workbook

    workbook = load_workbook(excel_file_path, read_only=True, data_only=True)
    available = workbook.sheetnames

    sheets = []
    for name in IMPORT_SHEET_NAMES:
        if name in available:
            sheets.append((workbook[name], name))

    if sheets:
        logger.info(
            f"Found {len(sheets)} importable sheets in {Path(excel_file_path).name}: "
            f"{[s[1] for s in sheets]}"
        )
        return sheets, workbook

    # Fallback: not a multi-sheet commercial log (e.g. monthly import)
    workbook.close()
    worksheet, sheet_name, workbook = get_excel_worksheet_flexible(excel_file_path)
    return [(worksheet, sheet_name)], workbook


def extract_display_months_from_excel(excel_file: str) -> Generator[str, None, None]:
    """
    Extract broadcast months from all importable sheets in an Excel file.

    Reads every sheet that contains spot data (Commercials, Worldlink Lines,
    Add to booked business, Pending) so that month analysis covers the full
    dataset — not just the first sheet found.

    Yields:
        str: Broadcast months in display format (e.g., 'Nov-24')
    """
    try:
        sheets, workbook = get_all_import_worksheets(excel_file)

        months = set()
        total_row_count = 0

        for worksheet, sheet_name in sheets:
            month_column_index = _find_month_column(worksheet)

            row_count = 0
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                row_count += 1
                if not row or len(row) <= month_column_index:
                    continue

                month_str = _parse_month_value(row[month_column_index])
                if month_str:
                    months.add(month_str)

            total_row_count += row_count
            logger.debug(
                f"Sheet '{sheet_name}': scanned {row_count:,} rows"
            )

        workbook.close()

        logger.info(
            f"Extracted {len(months)} unique months from {total_row_count:,} rows "
            f"across {len(sheets)} sheet(s)"
        )
        logger.debug(f"Months found: {sorted(months)}")

        for month in sorted(months):
            yield month

    except Exception as e:
        logger.error(f"Failed to extract months from {excel_file}: {e}")
        raise Exception(f"Failed to process Excel file: {str(e)}")


def _find_month_column(worksheet) -> int:
    """Find the broadcast_month column index in a worksheet."""
    month_column_names = [
        "broadcast_month", "Month", "month", "Broadcast Month", "Broadcast_Month",
    ]
    try:
        header_row = next(
            worksheet.iter_rows(min_row=1, max_row=1, values_only=True)
        )
        if header_row:
            for i, header in enumerate(header_row):
                if header and str(header).strip() in month_column_names:
                    return i
    except Exception:
        pass
    return 18  # fallback: known position from EXCEL_COLUMN_POSITIONS


def _parse_month_value(month_value) -> str | None:
    """Parse a month cell value into 'Mmm-YY' display format."""
    if not month_value:
        return None
    try:
        if hasattr(month_value, "strftime"):
            return month_value.strftime("%b-%y")
        if isinstance(month_value, str) and "-" in month_value:
            return month_value.strip()
        if isinstance(month_value, str):
            for fmt in ("%Y-%m-%d", "%m/%d/%Y"):
                try:
                    return datetime.strptime(month_value.strip(), fmt).date().strftime("%b-%y")
                except ValueError:
                    continue
            return None
        if hasattr(month_value, "date"):
            return month_value.date().strftime("%b-%y")
        return datetime.strptime(str(month_value), "%Y-%m-%d").date().strftime("%b-%y")
    except Exception:
        return None


def validate_excel_for_import(
    excel_file: str, import_mode: str, db_path: str
) -> ValidationResult:
    """
    ENHANCED: Validate Excel file for import with flexible sheet detection.

    Args:
        excel_file: Path to Excel file
        import_mode: 'HISTORICAL', 'WEEKLY_UPDATE', or 'MANUAL'
        db_path: Path to database file

    Returns:
        ValidationResult: Detailed validation result
    """
    try:
        # Use enhanced month extraction
        display_months = list(extract_display_months_from_excel(excel_file))

        if not display_months:
            return ValidationResult(
                is_valid=False,
                error_message="No broadcast months found in Excel file",
                suggested_action="Check if the Excel file contains valid month data in the broadcast_month column",
            )

        logger.info(
            f"Found {len(display_months)} months for validation: {display_months}"
        )

        # Import and use existing validation logic
        from src.services.month_closure_service import MonthClosureService
        from src.database.connection import DatabaseConnection

        db = DatabaseConnection(db_path)
        try:
            closure_service = MonthClosureService(db)
            validation_result = closure_service.validate_months_for_import(
                display_months, import_mode
            )

            # Enhance result with month details
            if hasattr(validation_result, "closed_months_found"):
                validation_result.closed_months_found = (
                    validation_result.closed_months_found or []
                )
            if hasattr(validation_result, "open_months_found"):
                validation_result.open_months_found = (
                    validation_result.open_months_found or []
                )

            return validation_result

        finally:
            db.close()

    except Exception as e:
        logger.error(f"Validation failed for {excel_file}: {e}")
        return ValidationResult(
            is_valid=False,
            error_message=f"Failed to validate Excel file: {str(e)}",
            suggested_action="Check if the Excel file is properly formatted and accessible",
        )


def get_excel_import_summary(excel_file: str, db_path: str) -> Dict[str, Any]:
    """
    ENHANCED: Get comprehensive import summary with flexible sheet detection.

    Returns summary of what would be imported including month breakdown,
    existing data analysis, and validation status.
    """
    try:
        # Use enhanced month extraction
        months_in_excel = list(extract_display_months_from_excel(excel_file))

        if not months_in_excel:
            return {
                "months_in_excel": [],
                "total_existing_spots_affected": 0,
                "open_months": [],
                "closed_months": [],
                "validation_status": "failed",
                "error": "No months found in Excel file",
            }

        # Get database info
        from src.database.connection import DatabaseConnection
        from src.services.month_closure_service import MonthClosureService

        db = DatabaseConnection(db_path)
        try:
            # Check closed months
            closure_service = MonthClosureService(db)
            closed_months = closure_service.get_closed_months(months_in_excel)
            open_months = [m for m in months_in_excel if m not in closed_months]

            # Count existing spots that would be affected
            with db.transaction() as conn:
                placeholders = ",".join("?" * len(months_in_excel))
                cursor = conn.execute(
                    f"""
                    SELECT COUNT(*) 
                    FROM spots 
                    WHERE broadcast_month IN ({placeholders})
                """,
                    months_in_excel,
                )

                total_existing_spots = cursor.fetchone()[0]

            # Get row count across all importable sheets
            sheets, workbook = get_all_import_worksheets(excel_file)
            total_rows_in_excel = sum(
                (ws.max_row - 1) for ws, _ in sheets
            )
            sheets_used = [name for _, name in sheets]
            workbook.close()

            return {
                "months_in_excel": sorted(months_in_excel),
                "total_existing_spots_affected": total_existing_spots,
                "total_rows_in_excel": total_rows_in_excel,
                "open_months": sorted(open_months),
                "closed_months": sorted(closed_months),
                "sheet_used": ", ".join(sheets_used),
                "validation_status": "success"
                if len(months_in_excel) > 0
                else "failed",
            }

        finally:
            db.close()

    except Exception as e:
        logger.error(f"Failed to generate import summary for {excel_file}: {e}")
        return {
            "months_in_excel": [],
            "total_existing_spots_affected": 0,
            "open_months": [],
            "closed_months": [],
            "validation_status": "failed",
            "error": str(e),
        }


def analyze_excel_structure(excel_file: str) -> Dict[str, Any]:
    """
    Analyze Excel file structure for debugging purposes.

    Returns detailed information about sheets, columns, and data structure.
    """
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(excel_file, read_only=True, data_only=True)

        analysis = {
            "file_path": excel_file,
            "file_name": Path(excel_file).name,
            "total_sheets": len(workbook.sheetnames),
            "sheet_names": workbook.sheetnames,
            "sheets_analysis": {},
        }

        # Analyze each sheet
        for sheet_name in workbook.sheetnames:
            try:
                worksheet = workbook[sheet_name]

                # Get header row
                header_row = next(
                    worksheet.iter_rows(min_row=1, max_row=1, values_only=True)
                )
                headers = (
                    [str(h) if h else f"Column_{i}" for i, h in enumerate(header_row)]
                    if header_row
                    else []
                )

                # Find potential month column
                month_column_candidates = []
                if header_row:
                    for i, header in enumerate(header_row):
                        if header and any(
                            keyword in str(header).lower()
                            for keyword in ["month", "broadcast"]
                        ):
                            month_column_candidates.append((i, str(header)))

                analysis["sheets_analysis"][sheet_name] = {
                    "total_rows": worksheet.max_row,
                    "total_columns": worksheet.max_column,
                    "data_rows": max(0, worksheet.max_row - 1),
                    "headers": headers[:20],  # First 20 headers
                    "month_column_candidates": month_column_candidates,
                    "is_active_sheet": worksheet == workbook.active,
                }

            except Exception as e:
                analysis["sheets_analysis"][sheet_name] = {"error": str(e)}

        workbook.close()
        return analysis

    except Exception as e:
        return {
            "file_path": excel_file,
            "error": f"Failed to analyze Excel structure: {str(e)}",
        }


# Convenience function for testing
def test_excel_processing(excel_file: str, db_path: str = None):
    """
    Test function to validate Excel processing with enhanced utilities.
    """
    print(f"Testing Excel processing for: {excel_file}")
    print("=" * 60)

    # Test structure analysis
    print("1. EXCEL STRUCTURE ANALYSIS:")
    structure = analyze_excel_structure(excel_file)

    if "error" in structure:
        print(f"   ERROR: {structure['error']}")
        return

    print(f"   File: {structure['file_name']}")
    print(
        f"   Sheets: {structure['total_sheets']} ({', '.join(structure['sheet_names'])})"
    )

    for sheet_name, sheet_info in structure["sheets_analysis"].items():
        if "error" not in sheet_info:
            print(
                f"   {sheet_name}: {sheet_info['data_rows']:,} data rows, {sheet_info['total_columns']} columns"
            )
            if sheet_info["month_column_candidates"]:
                print(f"      Month columns: {sheet_info['month_column_candidates']}")

    # Test month extraction
    print("\n2. MONTH EXTRACTION:")
    try:
        months = list(extract_display_months_from_excel(excel_file))
        print(
            f"   Found {len(months)} months: {', '.join(months) if len(months) <= 10 else f'{len(months)} months'}"
        )
        if len(months) > 10:
            print(f"   Sample: {', '.join(months[:10])}, ...")
    except Exception as e:
        print(f"   ERROR: {e}")
        return

    # Test validation (if database provided)
    if db_path:
        print("\n3. VALIDATION TEST:")
        try:
            validation = validate_excel_for_import(excel_file, "WEEKLY_UPDATE", db_path)
            print(f"   Valid: {validation.is_valid}")
            if not validation.is_valid:
                print(f"   Error: {validation.error_message}")
                print(f"   Solution: {validation.suggested_action}")
            else:
                print(f"   Open months: {len(validation.open_months_found)}")
                print(f"   Closed months: {len(validation.closed_months_found)}")
        except Exception as e:
            print(f"   ERROR: {e}")

    # Test import summary (if database provided)
    if db_path:
        print("\n4. IMPORT SUMMARY:")
        try:
            summary = get_excel_import_summary(excel_file, db_path)
            print(f"   Status: {summary.get('validation_status', 'unknown')}")
            print(f"   Months in Excel: {len(summary.get('months_in_excel', []))}")
            print(
                f"   Existing spots affected: {summary.get('total_existing_spots_affected', 0):,}"
            )
            print(f"   Sheet used: {summary.get('sheet_used', 'unknown')}")
        except Exception as e:
            print(f"   ERROR: {e}")

    print("\n✅ Excel processing test completed")


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Test enhanced Excel import utilities")
    parser.add_argument("excel_file", help="Excel file to test")
    parser.add_argument("--db-path", help="Database path for validation tests")
    parser.add_argument(
        "--analyze-only", action="store_true", help="Only analyze structure"
    )

    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

    if not Path(args.excel_file).exists():
        print(f"Excel file not found: {args.excel_file}")
        sys.exit(1)

    if args.analyze_only:
        structure = analyze_excel_structure(args.excel_file)
        print("Excel Structure Analysis:")
        import json

        print(json.dumps(structure, indent=2, default=str))
    else:
        test_excel_processing(args.excel_file, args.db_path)
