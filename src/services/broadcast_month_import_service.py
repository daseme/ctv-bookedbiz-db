#!/usr/bin/env python3
"""
Broadcast month import service for managing month-based data imports.
Enhanced to support multi-sheet source tracking using source_file field.
Orchestrates the complete import workflow with validation, deletion, and import.
Enhanced with progress bars and clean output.
"""

import re
import sys
import logging
import uuid
import contextlib
import io
from pathlib import Path
from datetime import datetime, date, timedelta
from typing import List, Set, Optional, Dict, Any, Callable
from dataclasses import dataclass

# Add tqdm for progress bars
from tqdm import tqdm

# Add src to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from src.database.connection import DatabaseConnection
from src.services.month_closure_service import (
    MonthClosureService,
    ValidationResult,
    MonthClosureError,
)
from src.services.import_integration_utilities import (
    extract_display_months_from_excel,
    validate_excel_for_import,
)
from src.utils.broadcast_month_utils import (
    BroadcastMonthParser,
    extract_broadcast_months_from_excel,
)
from src.services.base_service import BaseService
from src.utils.broadcast_month_utils import normalize_broadcast_day
from src.services.entity_alias_service import EntityAliasService
from src.services.import_performance_optimization import BatchEntityResolver
from src.services.import_integration_utilities import get_excel_worksheet_flexible

logger = logging.getLogger(__name__)

EXCEL_COLUMN_POSITIONS = {
    0: "bill_code",  # Bill Code
    1: "air_date",  # Start Date
    2: "end_date",  # End Date
    3: "day_of_week",  # Day(s)
    4: "time_in",  # Time In
    5: "time_out",  # Time Out
    6: "length_seconds",  # Length
    7: "media",  # Media/Name/Program (was "program", now renamed)
    8: "comments",  # Comments (if we want to preserve both, otherwise make this `None`)
    9: "language_code",  # Language
    10: "format",  # Format
    11: "sequence_number",  # Units-Spot count
    12: "line_number",  # Line
    13: "spot_type",  # Type
    14: "estimate",  # Agency/Episode# or cut number
    15: "gross_rate",  # Unit rate Gross
    16: "make_good",  # Make Good
    17: "spot_value",  # Spot Value
    18: "broadcast_month",  # Month
    19: "broker_fees",  # Broker Fees
    20: "priority",  # Sales/rep com: revenue sharing
    21: "station_net",  # Station Net
    22: "sales_person",  # Sales Person
    23: "revenue_type",  # Revenue Type
    24: "billing_type",  # Billing Type
    25: "agency_flag",  # Agency?
    26: "affidavit_flag",  # Affidavit?
    27: "contract",  # Notarize?
    28: "market_name",  # Market
    29: "sheet_source",  # NEW: Sheet source tracking (added by commercial log importer)
}


@contextlib.contextmanager
def suppress_verbose_logging():
    """Context manager to suppress verbose logging during import operations."""
    # Get all loggers and save their levels
    root_logger = logging.getLogger()
    original_level = root_logger.level

    # Set root logger to only show WARNINGS and above
    root_logger.setLevel(logging.WARNING)

    # Suppress specific noisy loggers
    noisy_loggers = ["openpyxl", "xlrd", "pandas", "services", "utils", "__main__"]

    original_levels = {}
    for logger_name in noisy_loggers:
        logger_obj = logging.getLogger(logger_name)
        original_levels[logger_name] = logger_obj.level
        logger_obj.setLevel(logging.WARNING)

    try:
        yield
    finally:
        # Restore original logging levels
        root_logger.setLevel(original_level)
        for logger_name, level in original_levels.items():
            logging.getLogger(logger_name).setLevel(level)


@contextlib.contextmanager
def suppress_stdout_stderr():
    """Context manager to suppress stdout/stderr during noisy operations."""
    old_stdout = sys.stdout
    old_stderr = sys.stderr

    try:
        sys.stdout = io.StringIO()
        sys.stderr = io.StringIO()
        yield
    finally:
        sys.stdout = old_stdout
        sys.stderr = old_stderr


@dataclass
class ImportResult:
    """Comprehensive result of import operation."""

    success: bool
    batch_id: str
    import_mode: str
    broadcast_months_affected: List[str]
    records_deleted: int
    records_imported: int
    duration_seconds: float
    error_messages: List[str]
    closed_months: List[str]  # For historical mode

    def __post_init__(self):
        """Ensure lists are not None."""
        if self.error_messages is None:
            self.error_messages = []
        if self.closed_months is None:
            self.closed_months = []


# ============================================================================
# Value Objects for Source Tracking
# ============================================================================


class SourceFileFormatter:
    """Value object for creating standardized source file tracking strings"""

    @staticmethod
    def format_source_file(filename: str, sheet_name: Optional[str]) -> str:
        """
        Create source_file string in format 'filename:sheet_name' or just 'filename'

        Args:
            filename: The Excel file name (e.g., 'Commercial Log 250902.xlsx')
            sheet_name: The sheet name (e.g., 'Commercials', 'WorldLink Lines')

        Returns:
            Formatted source file string for database storage
        """
        if sheet_name and sheet_name.strip():
            return f"{filename}:{sheet_name.strip()}"
        else:
            return filename

    @staticmethod
    def extract_filename_from_path(file_path: str) -> str:
        """Extract just the filename from a full path"""
        return Path(file_path).name


# ============================================================================
# Business Logic Enhancements
# ============================================================================


def build_revenue_type_normalizer() -> Callable[[Optional[str]], Optional[str]]:
    """Pure normalizer: A&O variants → 'Internal Ad Sales'; else passthrough."""

    def _norm(v: Optional[str]) -> Optional[str]:
        if v is None:
            return None
        key = re.sub(r"[\s\./-]+", "", v.lower()).replace("&", "and")
        return "Internal Ad Sales" if key in {"ao", "aando"} else v

    return _norm


def build_spot_type_normalizer() -> Callable[[Optional[str]], str]:
    """Pure normalizer for CHECK constraint on spots.spot_type."""
    allowed = {"AV", "BB", "BNS", "COM", "CRD", "PKG", "PRD", "PRG", "SVC", ""}
    aliases = {
        "SPOT": "COM",
        "COMMERCIAL": "COM",
        "BONUS": "BNS",
        "PKG": "PKG",
        "PROD": "PRD",
        "PROGRAM": "PRG",
        "SERVICE": "SVC",
    }

    def _norm(v: Optional[str]) -> str:
        x = (v or "").strip().upper()
        x = aliases.get(x, x)
        return x if x in allowed else ""

    return _norm


class BroadcastMonthImportError(Exception):
    """Raised when there's an error with import operations."""

    pass


class BroadcastMonthImportService(BaseService):
    def __init__(self, db_connection: DatabaseConnection):
        super().__init__(db_connection)
        self.closure_service = MonthClosureService(db_connection)
        self.parser = BroadcastMonthParser()
        self.alias_service = EntityAliasService(db_connection)
        self.normalize_revenue_type = build_revenue_type_normalizer()
        self.normalize_spot_type = build_spot_type_normalizer()
        self.batch_resolver = BatchEntityResolver(db_connection)

    def validate_import(self, excel_file: str, import_mode: str) -> ValidationResult:
        """
        Validate Excel file for import based on mode.

        Args:
            excel_file: Path to Excel file
            import_mode: 'HISTORICAL', 'WEEKLY_UPDATE', or 'MANUAL'

        Returns:
            ValidationResult with detailed validation info

        Raises:
            BroadcastMonthImportError: If validation cannot be performed
        """
        logger.info(f"Validating {excel_file} for {import_mode} import")

        try:
            # Extract broadcast months from Excel in display format
            with suppress_verbose_logging():
                display_months = list(extract_display_months_from_excel(excel_file))

            if not display_months:
                raise BroadcastMonthImportError(
                    "No broadcast months found in Excel file"
                )

            logger.info(
                f"Found {len(display_months)} months in Excel: {sorted(display_months)}"
            )

            # Validate against closed months
            return self.closure_service.validate_months_for_import(
                display_months, import_mode
            )

        except Exception as e:
            error_msg = f"Failed to validate import: {str(e)}"
            logger.error(error_msg)
            raise BroadcastMonthImportError(error_msg)

    def _lookup_market_id(self, market_name: str, conn) -> Optional[int]:
        if not market_name:
            return None

        # Validate input first
        if len(market_name) > 50 or any(c in market_name for c in [";", "--", "/*"]):
            logger.warning(f"Suspicious market_name rejected: {market_name}")
            return None

        cursor = conn.execute(
            """
            SELECT market_id FROM markets 
            WHERE market_code = ? OR market_name = ?
        """,
            (market_name.strip(), market_name.strip()),
        )

        result = cursor.fetchone()
        return result[0] if result else None

    def _lookup_entities_exact(self, bill_code: str, conn) -> Dict[str, Optional[int]]:
        """
        Phase 1: Exact string matching for agencies and customers.
        Bill code format: "Agency:Customer Name"
        """
        result = {"agency_id": None, "customer_id": None}

        if not bill_code or ":" not in bill_code:
            return result

        try:
            # Split on ':' to get agency and customer
            parts = bill_code.split(":", 1)
            if len(parts) != 2:
                return result

            agency_name = parts[0].strip()
            customer_name = parts[1].strip()

            # Exact agency lookup
            cursor = conn.execute(
                """
                SELECT agency_id FROM agencies 
                WHERE agency_name = ? AND is_active = 1
            """,
                (agency_name,),
            )

            agency_result = cursor.fetchone()
            if agency_result:
                result["agency_id"] = agency_result[0]

            # Exact customer lookup
            cursor = conn.execute(
                """
                SELECT customer_id FROM customers 
                WHERE normalized_name = ? AND is_active = 1
            """,
                (customer_name,),
            )

            customer_result = cursor.fetchone()
            if customer_result:
                result["customer_id"] = customer_result[0]

            return result

        except Exception as e:
            logger.warning(
                f"Failed to lookup entities for bill_code '{bill_code}': {e}"
            )
            return result

    def _lookup_direct_customer(self, bill_code: str, conn) -> Optional[int]:
        """
        Phase 1.5: Handle direct billing customers (no agency in bill_code).
        For entities like "Sacramento County Water Agency" that are customers, not "Agency:Customer".
        """
        if not bill_code:
            return None

        try:
            # Direct exact lookup for customer
            cursor = conn.execute(
                """
                SELECT customer_id FROM customers 
                WHERE normalized_name = ? AND is_active = 1
            """,
                (bill_code.strip(),),
            )

            result = cursor.fetchone()
            return result[0] if result else None

        except Exception as e:
            logger.warning(f"Failed direct customer lookup for '{bill_code}': {e}")
            return None

# ============================================================================
    # Month Preservation Safeguard Methods
    # ============================================================================
# ============================================================================
    # Month Preservation & Filtering Safeguard Methods
    # ============================================================================

    def _get_previous_calendar_month(self, reference_date: date = None) -> str:
        """
        Get the previous calendar month in display format (e.g., 'Dec-24').
        Used to filter out stray previous-month data from daily imports.
        """
        if reference_date is None:
            reference_date = date.today()
        
        # Go to first of current month, then subtract one day to get last day of previous month
        first_of_month = reference_date.replace(day=1)
        last_of_previous = first_of_month - timedelta(days=1)
        
        return last_of_previous.strftime("%b-%y")

    def _get_months_with_data_in_excel(self, excel_file: str) -> Dict[str, int]:
        """
        Get a count of records per broadcast month in the Excel file.
        Used to determine which months actually have data to import.
        """
        from datetime import datetime
        
        month_counts = {}
        
        try:
            with suppress_verbose_logging(), suppress_stdout_stderr():
                worksheet, sheet_name, workbook = get_excel_worksheet_flexible(
                    excel_file
                )
            
            month_col_indices = [
                k for k, v in EXCEL_COLUMN_POSITIONS.items() 
                if v == "broadcast_month"
            ]
            if not month_col_indices:
                workbook.close()
                return month_counts
            month_col = month_col_indices[0]
            
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                
                month_value = row[month_col] if month_col < len(row) else None
                if not month_value:
                    continue
                
                try:
                    if hasattr(month_value, 'date'):
                        bm_date = month_value.date()
                    elif isinstance(month_value, str):
                        parsed = False
                        for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", 
                                    "%Y-%m-%d %H:%M:%S"]:
                            try:
                                bm_date = datetime.strptime(
                                    month_value.strip(), fmt
                                ).date()
                                parsed = True
                                break
                            except:
                                continue
                        if not parsed:
                            continue
                    else:
                        bm_date = month_value
                    
                    broadcast_month_display = bm_date.strftime("%b-%y")
                    month_counts[broadcast_month_display] = (
                        month_counts.get(broadcast_month_display, 0) + 1
                    )
                except:
                    continue
            
            workbook.close()
            
        except Exception as e:
            tqdm.write(f"Warning: Could not count Excel records by month: {e}")
        
        return month_counts

    def _identify_months_to_preserve(
        self, 
        open_months: List[str], 
        excel_month_counts: Dict[str, int],
        conn
    ) -> Dict[str, Dict[str, Any]]:
        """
        Identify open months that should be preserved (not deleted) because:
        1. They have existing data in the database
        2. The Excel file has NO data for that month
        
        Returns dict of month -> {existing_count, existing_revenue, reason}
        """
        months_to_preserve = {}
        
        for month in open_months:
            excel_count = excel_month_counts.get(month, 0)
            
            if excel_count == 0:
                cursor = conn.execute("""
                    SELECT COUNT(*), COALESCE(SUM(gross_rate), 0)
                    FROM spots 
                    WHERE broadcast_month = ?
                """, (month,))
                
                existing_count, existing_revenue = cursor.fetchone()
                
                if existing_count > 0:
                    months_to_preserve[month] = {
                        'existing_count': existing_count,
                        'existing_revenue': existing_revenue,
                        'reason': 'No data in Excel - preserving existing'
                    }
        
        return months_to_preserve

    def execute_month_replacement(
        self,
        excel_file: str,
        import_mode: str,
        closed_by: str = None,
        dry_run: bool = False,
    ) -> ImportResult:
        """
        Execute complete month replacement workflow with progress tracking.
        
        Enhanced with:
        - Month preservation safeguard for unclosed months without Excel data
        - Previous month filtering to skip stray MC operational data
        """
        start_time = datetime.now()
        batch_id = f"{import_mode.lower()}_{int(start_time.timestamp())}"

        result = ImportResult(
            success=False,
            batch_id=batch_id,
            import_mode=import_mode,
            broadcast_months_affected=[],
            records_deleted=0,
            records_imported=0,
            duration_seconds=0.0,
            error_messages=[],
            closed_months=[],
        )

        try:
            # Step 1: Extract months from Excel
            with suppress_verbose_logging():
                tqdm.write(f"Analyzing Excel file: {Path(excel_file).name}")
                display_months = list(
                    extract_display_months_from_excel(excel_file)
                )

            if not display_months:
                raise BroadcastMonthImportError(
                    "No broadcast months found in Excel file"
                )

            tqdm.write(
                f"Found {len(display_months)} months: "
                f"{', '.join(sorted(display_months))}"
            )

            # Step 2: Check which months are closed
            closed_months = self.closure_service.get_closed_months(display_months)
            open_months = [
                month for month in display_months if month not in closed_months
            ]

            if closed_months:
                tqdm.write(
                    f"Closed months: {len(closed_months)} "
                    f"({', '.join(sorted(closed_months))})"
                )
            if open_months:
                tqdm.write(
                    f"Open months: {len(open_months)} "
                    f"({', '.join(sorted(open_months))})"
                )

            # Step 3: Handle different import modes
            if import_mode == "WEEKLY_UPDATE":
                if closed_months:
                    tqdm.write(
                        f"WEEKLY_UPDATE: Auto-skipping "
                        f"{len(closed_months)} closed months"
                    )

                if not open_months:
                    tqdm.write(
                        "No open months to import - all months are closed"
                    )
                    result.success = True
                    return result

                # Step 3.5: Check for months to preserve (no Excel data)
                tqdm.write(
                    "Analyzing Excel data by month for preservation check..."
                )
                excel_month_counts = self._get_months_with_data_in_excel(
                    excel_file
                )
                
                # Log what's in the Excel file
                for month in sorted(open_months):
                    count = excel_month_counts.get(month, 0)
                    if count > 0:
                        tqdm.write(f"   {month}: {count:,} records in Excel")
                    else:
                        tqdm.write(f"   {month}: NO DATA in Excel")
                
                # Identify months to preserve (have DB data, no Excel data)
                with self.safe_connection() as check_conn:
                    months_to_preserve = self._identify_months_to_preserve(
                        open_months, excel_month_counts, check_conn
                    )
                
                if months_to_preserve:
                    tqdm.write(
                        f"⚠️  PRESERVATION: Protecting "
                        f"{len(months_to_preserve)} open month(s) "
                        f"with no Excel data:"
                    )
                    for month, info in sorted(months_to_preserve.items()):
                        tqdm.write(
                            f"   {month}: {info['existing_count']:,} existing "
                            f"records, ${info['existing_revenue']:,.2f} "
                            f"revenue - PRESERVED"
                        )
                    
                    # Remove preserved months from processing list
                    open_months = [
                        m for m in open_months if m not in months_to_preserve
                    ]

                # Step 3.6: Skip previous month (stray MC operational data)
                previous_month = self._get_previous_calendar_month()
                current_month = date.today().strftime("%b-%y")
                
                if previous_month in open_months:
                    prev_month_excel_count = excel_month_counts.get(
                        previous_month, 0
                    )
                    
                    # Check what we have in DB for context
                    with self.safe_connection() as check_conn:
                        cursor = check_conn.execute("""
                            SELECT COUNT(*), COALESCE(SUM(gross_rate), 0)
                            FROM spots 
                            WHERE broadcast_month = ?
                        """, (previous_month,))
                        existing_count, existing_revenue = cursor.fetchone()
                    
                    tqdm.write(
                        f"⚠️  SKIPPING PREVIOUS MONTH: {previous_month}"
                    )
                    tqdm.write(
                        f"   Excel has {prev_month_excel_count:,} records "
                        f"(likely stray MC operational data)"
                    )
                    tqdm.write(
                        f"   Preserving {existing_count:,} existing records, "
                        f"${existing_revenue:,.2f} revenue"
                    )
                    tqdm.write(
                        f"   Reason: Once in {current_month}, previous month "
                        f"data is not authoritative"
                    )
                    
                    open_months = [
                        m for m in open_months if m != previous_month
                    ]
                
                # Final check - any months left to process?
                if not open_months:
                    tqdm.write(
                        "✅ No months to update after filtering "
                        "- existing data protected"
                    )
                    result.success = True
                    return result
                
                tqdm.write(
                    f"Proceeding with {len(open_months)} month(s): "
                    f"{', '.join(sorted(open_months))}"
                )

                result.broadcast_months_affected = open_months

            elif import_mode == "HISTORICAL":
                if not closed_by:
                    raise BroadcastMonthImportError(
                        "HISTORICAL mode requires --closed-by parameter"
                    )
                result.broadcast_months_affected = display_months

            else:  # MANUAL mode
                if closed_months:
                    validation_error = (
                        f"Manual import contains closed months: "
                        f"{closed_months}. Use --force to override "
                        f"or filter the Excel file."
                    )
                    result.error_messages.append(validation_error)
                    tqdm.write(f"{validation_error}")
                    return result
                result.broadcast_months_affected = display_months

            tqdm.write(
                f"Processing {len(result.broadcast_months_affected)} months: "
                f"{', '.join(result.broadcast_months_affected)}"
            )

            if dry_run:
                tqdm.write("DRY RUN - No changes will be made")
                result.success = True
                return result

            # Step 4: Build entity cache
            tqdm.write("🚀 Phase 1: Building high-performance entity cache...")
            self.batch_resolver.build_entity_cache_from_excel(excel_file)
            cache_stats = self.batch_resolver.get_performance_stats()
            tqdm.write(
                f"✅ Cache ready: {cache_stats['cache_size']} entities "
                f"pre-resolved ({cache_stats['batch_resolved']} found)"
            )

            # Step 5: Create import batch record
            batch_record_id = self._create_import_batch(
                batch_id, import_mode, excel_file, 
                result.broadcast_months_affected
            )

            # Step 6: Execute the import in transaction
            try:
                with self.safe_transaction() as conn:
                    # Delete existing data
                    deleted_count = (
                        self._delete_broadcast_month_data_with_progress(
                            result.broadcast_months_affected, conn
                        )
                    )
                    result.records_deleted = deleted_count

                    # Import filtered data
                    imported_count = self._import_excel_data_with_progress(
                        excel_file, batch_id, conn, 
                        result.broadcast_months_affected
                    )
                    result.records_imported = imported_count

                    if import_mode == "WEEKLY_UPDATE" and closed_months:
                        net_change = imported_count - deleted_count
                        tqdm.write(
                            f"Import complete: {imported_count:,} imported, "
                            f"{deleted_count:,} deleted (net: {net_change:+,})"
                        )

                    # Step 7: Validate customer alignment
                    tqdm.write("Validating customer alignment...")
                    validation_result = (
                        self.validate_customer_alignment_post_import(
                            batch_id, conn
                        )
                    )

                    if not validation_result["validation_passed"]:
                        tqdm.write("⚠️ Customer alignment issues detected:")
                        tqdm.write(
                            f"   {validation_result['total_mismatches']} "
                            f"mismatches affecting "
                            f"{validation_result['total_spots_affected']:,} "
                            f"spots"
                        )
                        tqdm.write(
                            f"   Revenue affected: "
                            f"${validation_result['total_revenue_affected']:,.2f}"
                        )

                        tqdm.write(
                            "🔧 Auto-correcting customer_id mismatches..."
                        )
                        self.auto_correct_customer_mismatches(batch_id, conn)

                        validation_result = (
                            self.validate_customer_alignment_post_import(
                                batch_id, conn
                            )
                        )

                        if validation_result["validation_passed"]:
                            tqdm.write(
                                "✅ All customer_id mismatches corrected"
                            )
                        else:
                            error_msg = (
                                f"Failed to correct "
                                f"{validation_result['total_mismatches']} "
                                f"customer alignment issues"
                            )
                            tqdm.write(f"❌ {error_msg}")
                            raise BroadcastMonthImportError(error_msg)
                    else:
                        tqdm.write("✅ Customer alignment validation passed")

                    # Step 8: Handle HISTORICAL mode
                    if import_mode == "HISTORICAL":
                        with tqdm(
                            total=len(result.broadcast_months_affected),
                            desc="Closing months",
                            unit=" months",
                        ) as pbar:
                            for month in result.broadcast_months_affected:
                                try:
                                    self.closure_service \
                                        .close_broadcast_month_with_connection(
                                            month, closed_by, conn
                                        )
                                    result.closed_months.append(month)
                                    pbar.update(1)
                                    pbar.set_description(f"Closed {month}")
                                except MonthClosureError as e:
                                    tqdm.write(
                                        f"Failed to close month {month}: {e}"
                                    )
                                    pbar.update(1)
                        tqdm.write(
                            f"Closed {len(result.closed_months)} months"
                        )

                    # Step 9: Complete the batch record
                    self._complete_import_batch(batch_id, result, conn)

                    result.success = True
                    duration = (datetime.now() - start_time).total_seconds()
                    tqdm.write(
                        f"Import completed successfully in "
                        f"{duration:.1f} seconds"
                    )

            except Exception as transaction_error:
                error_msg = f"Transaction failed: {str(transaction_error)}"
                tqdm.write(f"{error_msg}")
                result.error_messages.append(error_msg)
                self._fail_import_batch(batch_id, error_msg)
                raise BroadcastMonthImportError(error_msg)

        except Exception as e:
            error_msg = f"Import failed: {str(e)}"
            tqdm.write(f"{error_msg}")
            result.error_messages.append(error_msg)
            if "batch_id" in locals():
                self._fail_import_batch(batch_id, error_msg)

        result.duration_seconds = (datetime.now() - start_time).total_seconds()

        return result

    def validate_customer_alignment_post_import(
        self, batch_id: str, conn
    ) -> Dict[str, Any]:
        """
        Validate that all imported spots align with normalization system
        Run this after every import to catch mismatches immediately
        """
        # Check for customer_id mismatches
        cursor = conn.execute(
            """
            SELECT 
                s.bill_code,
                s.customer_id as spots_customer_id,
                audit.customer_id as audit_customer_id,
                COUNT(*) as spot_count,
                SUM(COALESCE(s.gross_rate, 0)) as revenue_affected
            FROM spots s
            LEFT JOIN v_customer_normalization_audit audit ON audit.raw_text = s.bill_code
            WHERE s.import_batch_id = ?
            AND (s.customer_id != audit.customer_id OR s.customer_id IS NULL)
            AND audit.customer_id IS NOT NULL
            GROUP BY s.bill_code, s.customer_id, audit.customer_id
            ORDER BY revenue_affected DESC
        """,
            (batch_id,),
        )

        mismatches = []
        total_spots = 0
        total_revenue = 0

        for row in cursor.fetchall():
            mismatch = {
                "bill_code": row[0],
                "spots_customer_id": row[1],
                "audit_customer_id": row[2],
                "spot_count": row[3],
                "revenue_affected": row[4],
            }
            mismatches.append(mismatch)
            total_spots += row[3]
            total_revenue += row[4]

        return {
            "mismatches": mismatches,
            "total_mismatches": len(mismatches),
            "total_spots_affected": total_spots,
            "total_revenue_affected": total_revenue,
            "validation_passed": len(mismatches) == 0,
        }

    def auto_correct_customer_mismatches(self, batch_id: str, conn) -> int:
        """
        Automatically correct customer_id mismatches after import
        Returns number of corrections made
        """
        cursor = conn.execute(
            """
            UPDATE spots 
            SET customer_id = (
                SELECT audit.customer_id 
                FROM v_customer_normalization_audit audit
                WHERE audit.raw_text = spots.bill_code
            )
            WHERE spots.import_batch_id = ?
            AND EXISTS (
                SELECT 1 FROM v_customer_normalization_audit audit
                WHERE audit.raw_text = spots.bill_code
                    AND (spots.customer_id != audit.customer_id OR spots.customer_id IS NULL)
                    AND audit.customer_id IS NOT NULL
            )
        """,
            (batch_id,),
        )

        corrections_made = cursor.rowcount
        if corrections_made > 0:
            tqdm.write(f"Auto-corrected {corrections_made} customer_id mismatches")

        return corrections_made

    def _create_import_batch(
        self, batch_id: str, import_mode: str, source_file: str, months: List[str]
    ) -> int:
        """Create import batch record."""
        try:
            with self.safe_transaction() as conn:
                cursor = conn.execute(
                    """
                    INSERT INTO import_batches 
                    (batch_id, import_mode, source_file, import_date, status, broadcast_months_affected)
                    VALUES (?, ?, ?, CURRENT_TIMESTAMP, 'RUNNING', ?)
                """,
                    (batch_id, import_mode, source_file, str(months)),
                )

                batch_record_id = cursor.lastrowid
                return batch_record_id

        except Exception as e:
            error_msg = f"Failed to create import batch: {str(e)}"
            tqdm.write(f"{error_msg}")
            raise BroadcastMonthImportError(error_msg)

    def _delete_broadcast_month_data_with_progress(
        self, months: List[str], conn
    ) -> int:
        """Delete existing data for specified broadcast months with progress tracking."""
        if not months:
            return 0

        total_deleted = 0

        # First, get counts for progress bar
        month_counts = {}
        for month in months:
            cursor = conn.execute(
                "SELECT COUNT(*) FROM spots WHERE broadcast_month = ?", (month,)
            )
            count = cursor.fetchone()[0]
            month_counts[month] = count
            total_deleted += count

        if total_deleted == 0:
            tqdm.write("No existing records to delete")
            return 0

        # Delete with progress tracking
        deleted_so_far = 0
        with tqdm(
            total=total_deleted, desc="Deleting existing records", unit=" records"
        ) as pbar:
            for month in months:
                count = month_counts[month]
                if count > 0:
                    cursor = conn.execute(
                        "DELETE FROM spots WHERE broadcast_month = ?", (month,)
                    )
                    actual_deleted = cursor.rowcount
                    deleted_so_far += actual_deleted
                    pbar.update(actual_deleted)
                    pbar.set_description(
                        f"Deleted {deleted_so_far:,}/{total_deleted:,}"
                    )

        tqdm.write(f"Deleted {total_deleted:,} existing records")
        return total_deleted

    def _import_excel_data_with_progress(
        self, excel_file: str, batch_id: str, conn, allowed_months: List[str]
    ) -> int:
        """
        Import Excel data using fixed column positions with comprehensive progress tracking.
        ENHANCED: Now supports sheet source tracking using source_file field.
        """
        from openpyxl import load_workbook
        from datetime import datetime
        import re

        # Verify batch exists
        cursor = conn.execute(
            "SELECT COUNT(*) FROM import_batches WHERE batch_id = ?", (batch_id,)
        )
        if cursor.fetchone()[0] == 0:
            raise BroadcastMonthImportError(
                f"batch_id {batch_id} not found in import_batches table"
            )

        # Extract filename for source tracking
        filename = SourceFileFormatter.extract_filename_from_path(excel_file)

        try:
            with suppress_verbose_logging(), suppress_stdout_stderr():
                worksheet, sheet_name, workbook = get_excel_worksheet_flexible(excel_file)
                tqdm.write(f"Using sheet: {sheet_name}")

            total_records = worksheet.max_row - 1
            imported_count = 0
            skipped_count = 0
            filtered_count = 0
            unmatched_customers = set()
            unmatched_agencies = set()
            sheet_source_stats = {}  # Track sheet source statistics

            # 🚀 PERFORMANCE BOOST: Pre-populate entity cache
            tqdm.write("Phase 1: Building entity cache for high-performance lookups...")
            self.batch_resolver.build_entity_cache_from_excel(excel_file)
            cache_stats = self.batch_resolver.get_performance_stats()
            tqdm.write(
                f"Cache ready: {cache_stats['cache_size']} entities pre-resolved"
            )

            with tqdm(
                total=total_records, desc="Processing Excel rows", unit=" rows"
            ) as pbar:
                for row_num, row in enumerate(
                    worksheet.iter_rows(min_row=2, values_only=True), start=2
                ):
                    pbar.update(1)

                    try:
                        if not any(row):
                            continue

                        # Use fixed position for broadcast_month
                        month_col_index = [
                            k
                            for k, v in EXCEL_COLUMN_POSITIONS.items()
                            if v == "broadcast_month"
                        ]
                        if not month_col_index:
                            skipped_count += 1
                            continue
                        month_value = row[month_col_index[0]]

                        if not month_value:
                            skipped_count += 1
                            continue

                        # Format broadcast_month
                        try:
                            if hasattr(month_value, "date"):
                                bm_date = month_value.date()
                            elif isinstance(month_value, str):
                                for fmt in [
                                    "%Y-%m-%d",
                                    "%m/%d/%Y",
                                    "%d/%m/%Y",
                                    "%Y-%m-%d %H:%M:%S",
                                ]:
                                    try:
                                        bm_date = datetime.strptime(
                                            month_value.strip(), fmt
                                        ).date()
                                        break
                                    except:
                                        continue
                                else:
                                    bm_date = datetime.strptime(
                                        str(month_value), "%Y-%m-%d"
                                    ).date()
                            else:
                                bm_date = month_value
                            broadcast_month_display = bm_date.strftime("%b-%y")
                        except:
                            skipped_count += 1
                            continue

                        if broadcast_month_display not in allowed_months:
                            filtered_count += 1
                            continue

                        spot_data = {
                            "import_batch_id": batch_id,
                            "broadcast_month": broadcast_month_display,
                        }

                        # NEW: Handle sheet source tracking
                        sheet_source = None

                        for col_idx, field_name in EXCEL_COLUMN_POSITIONS.items():
                            if field_name and col_idx < len(row):
                                val = row[col_idx]
                                if val is None or val == "":
                                    continue

                                # NEW: Capture sheet_source but don't store in database
                                if field_name == "sheet_source":
                                    sheet_source = str(val).strip() if val else None
                                    # Track statistics for logging
                                    if sheet_source:
                                        sheet_source_stats[sheet_source] = (
                                            sheet_source_stats.get(sheet_source, 0) + 1
                                        )
                                    continue  # Don't store this column in database

                                if field_name == "bill_code":
                                    spot_data[field_name] = str(val).strip()

                                elif field_name == "air_date":
                                    try:
                                        spot_data[field_name] = (
                                            val.date().isoformat()
                                            if hasattr(val, "date")
                                            else str(val).strip()
                                        )
                                    except:
                                        spot_data[field_name] = str(val).strip()

                                elif field_name in [
                                    "gross_rate",
                                    "station_net",
                                    "spot_value",
                                    "broker_fees",
                                ]:
                                    try:
                                        spot_data[field_name] = float(val)
                                    except:
                                        spot_data[field_name] = None

                                elif field_name == "day_of_week":
                                    try:
                                        spot_data[field_name] = normalize_broadcast_day(
                                            str(val).strip()
                                        )
                                    except:
                                        spot_data[field_name] = str(val).strip()

                                elif (
                                    field_name == "revenue_type"
                                ):  # NEW: normalize A&O → Internal Ad Sales
                                    spot_data[field_name] = self.normalize_revenue_type(
                                        str(val).strip()
                                    )

                                elif (
                                    field_name == "spot_type"
                                ):  # NEW: enforce CHECK constraint
                                    spot_data[field_name] = self.normalize_spot_type(
                                        str(val).strip()
                                    )

                                elif (
                                    field_name != "broadcast_month"
                                ):  # already handled above
                                    spot_data[field_name] = str(val).strip()

                        # NEW: Set source_file with sheet tracking
                        spot_data["source_file"] = (
                            SourceFileFormatter.format_source_file(
                                filename, sheet_source
                            )
                        )

                        if "market_name" in spot_data:
                            market_id = self._lookup_market_id(
                                spot_data["market_name"], conn
                            )
                            if market_id:
                                spot_data["market_id"] = market_id

                        # NEW CODE - High-performance cached lookups:
                        if "bill_code" in spot_data:
                            entity_result = self.batch_resolver.lookup_entities_cached(
                                spot_data["bill_code"], conn
                            )
                            if entity_result.customer_id:
                                spot_data["customer_id"] = entity_result.customer_id
                            else:
                                unmatched_customers.add(spot_data["bill_code"])
                            if entity_result.agency_id:
                                spot_data["agency_id"] = entity_result.agency_id
                            else:
                                if ":" in spot_data["bill_code"]:
                                    unmatched_agencies.add(
                                        spot_data["bill_code"].split(":", 1)[0].strip()
                                    )

                        if "language_id" not in spot_data:
                            spot_data["language_id"] = 1

                        if not spot_data.get("bill_code") or not spot_data.get(
                            "air_date"
                        ):
                            skipped_count += 1
                            continue

                        with suppress_verbose_logging():
                            fields = list(spot_data.keys())
                            placeholders = ", ".join(["?"] * len(fields))
                            field_names = ", ".join(fields)
                            values = [spot_data[field] for field in fields]

                            query = f"INSERT INTO spots ({field_names}) VALUES ({placeholders})"
                            conn.execute(query, values)
                            imported_count += 1

                        if imported_count % 1000 == 0:
                            pbar.set_description(f"Imported {imported_count:,} records")

                    except Exception as row_error:
                        skipped_count += 1
                        if skipped_count <= 5:
                            tqdm.write(
                                f"Row {row_num} error: {str(row_error)[:100]}..."
                            )
                        continue

            workbook.close()

            # Enhanced completion logging with performance statistics
            final_stats = self.batch_resolver.get_performance_stats()
            tqdm.write(f"Import complete: {imported_count:,} records imported")
            tqdm.write(f"Entity resolution performance:")
            tqdm.write(f"   Cache hit rate: {final_stats['cache_hit_rate_percent']}%")
            tqdm.write(f"   Total lookups: {final_stats['total_lookups']:,}")
            tqdm.write(f"   Cache hits: {final_stats['cache_hits']:,}")

            if sheet_source_stats:
                tqdm.write(f"Sheet breakdown:")
                for sheet, count in sorted(sheet_source_stats.items()):
                    tqdm.write(f"   {sheet}: {count:,} records")

            if skipped_count:
                tqdm.write(f"Skipped: {skipped_count:,} records (errors)")
            if filtered_count:
                tqdm.write(f"Filtered: {filtered_count:,} records (closed months)")
            if unmatched_customers:
                tqdm.write(f"Unmatched customers: {len(unmatched_customers)}")
            if unmatched_agencies:
                tqdm.write(f"Unmatched agencies: {len(unmatched_agencies)}")

            return imported_count

        except Exception as e:
            raise BroadcastMonthImportError(f"Excel import failed: {e}")

    def _lookup_entities_with_aliases(self, bill_code: str, conn) -> dict:
        """
        ENHANCED: Use dashboard-compatible customer resolution
        This prevents future customer_id mismatches
        """
        result = {"customer_id": None, "agency_id": None}

        # PRIORITY 1: Use normalization system (what dashboard expects)
        cursor = conn.execute(
            """
            SELECT customer_id FROM v_customer_normalization_audit 
            WHERE raw_text = ?
        """,
            (bill_code,),
        )
        audit_result = cursor.fetchone()

        if audit_result:
            result["customer_id"] = audit_result[0]
            return result

        # PRIORITY 2: Create using normalization-compatible format
        if ":" in bill_code:
            # Create as agency:customer (dashboard-friendly)
            normalized_name = bill_code.strip()
        else:
            # Create as customer-only but log for review
            normalized_name = bill_code.strip()
            logger.warning(f"Creating customer-only record: {bill_code}")

        cursor = conn.execute(
            """
            INSERT INTO customers (normalized_name, is_active)
            VALUES (?, 1)
        """,
            (normalized_name,),
        )

        result["customer_id"] = cursor.lastrowid
        return result

    def _lookup_customer_id(self, customer_name: str, conn) -> Optional[int]:
        """Direct customer lookup (no alias)."""
        try:
            cursor = conn.execute(
                """
                SELECT customer_id FROM customers 
                WHERE normalized_name = ? AND is_active = 1
            """,
                (customer_name,),
            )

            result = cursor.fetchone()
            return result[0] if result else None
        except:
            return None

    def _lookup_agency_id(self, agency_name: str, conn) -> Optional[int]:
        """Direct agency lookup (no alias)."""
        try:
            cursor = conn.execute(
                """
                SELECT agency_id FROM agencies 
                WHERE agency_name = ? AND is_active = 1
            """,
                (agency_name,),
            )

            result = cursor.fetchone()
            return result[0] if result else None
        except:
            return None

    def _complete_import_batch(self, batch_id: str, result, conn):
        """Mark import batch as completed with statistics."""
        try:
            conn.execute(
                """
                UPDATE import_batches 
                SET status = 'COMPLETED',
                    completed_at = CURRENT_TIMESTAMP,
                    records_imported = ?,
                    records_deleted = ?
                WHERE batch_id = ?
            """,
                (
                    result.records_imported
                    if hasattr(result, "records_imported")
                    else 0,
                    result.records_deleted if hasattr(result, "records_deleted") else 0,
                    batch_id,
                ),
            )
        except Exception as e:
            logger.error(f"Failed to update batch completion: {e}")

    def _fail_import_batch(self, batch_id: str, error_message: str):
        """Mark import batch as failed."""
        try:
            with self.safe_transaction() as conn:
                conn.execute(
                    """
                    UPDATE import_batches 
                    SET status = 'FAILED',
                        completed_at = CURRENT_TIMESTAMP,
                        error_summary = ?
                    WHERE batch_id = ?
                """,
                    (error_message, batch_id),
                )
        except Exception as e:
            logger.error(f"Failed to mark batch as failed: {e}")

    def get_import_history(self, limit: int = 10) -> List[Dict[str, Any]]:
        """Get recent import history."""
        try:
            with self.safe_connection() as conn:
                cursor = conn.execute(
                    """
                    SELECT batch_id, import_mode, source_file, import_date, status,
                        records_imported, records_deleted, broadcast_months_affected
                    FROM import_batches 
                    ORDER BY import_date DESC 
                    LIMIT ?
                """,
                    (limit,),
                )

                history = []
                for row in cursor.fetchall():
                    history.append(
                        {
                            "batch_id": row[0],
                            "import_mode": row[1],
                            "source_file": row[2],
                            "import_date": row[3],
                            "status": row[4],
                            "records_imported": row[5] or 0,
                            "records_deleted": row[6] or 0,
                            "months_affected": eval(row[7]) if row[7] else [],
                        }
                    )

                return history

        except Exception as e:
            logger.error(f"Failed to get import history: {e}")
            return []

    def cleanup_failed_imports(self) -> int:
        """Clean up any failed or stuck import batches."""
        try:
            with self.safe_transaction() as conn:
                cursor = conn.execute("""
                    UPDATE import_batches 
                    SET status = 'FAILED',
                        error_summary = 'Import timed out or was interrupted',
                        completed_at = CURRENT_TIMESTAMP
                    WHERE status = 'RUNNING' 
                      AND import_date < datetime('now', '-1 hour')
                """)

                cleaned_count = cursor.rowcount

                if cleaned_count > 0:
                    tqdm.write(f"Cleaned up {cleaned_count} failed import batches")

                return cleaned_count

        except Exception as e:
            logger.error(f"Failed to cleanup imports: {e}")
            return 0


# Convenience functions for simple usage
def execute_import(
    excel_file: str,
    import_mode: str,
    db_path: str,
    closed_by: str = None,
    dry_run: bool = False,
) -> ImportResult:
    """Simple function to execute an import."""
    db_connection = DatabaseConnection(db_path)
    service = BroadcastMonthImportService(db_connection)

    try:
        return service.execute_month_replacement(
            excel_file, import_mode, closed_by, dry_run
        )
    finally:
        db_connection.close()


# Test and example usage
if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Test broadcast month import service")
    parser.add_argument("excel_file", help="Excel file to import")
    parser.add_argument(
        "--db-path", default="data/database/production.db", help="Database path"
    )
    parser.add_argument(
        "--mode",
        choices=["WEEKLY_UPDATE", "HISTORICAL", "MANUAL"],
        default="WEEKLY_UPDATE",
        help="Import mode",
    )
    parser.add_argument("--closed-by", help="Required for HISTORICAL mode")
    parser.add_argument(
        "--dry-run", action="store_true", help="Validate and preview without importing"
    )
    parser.add_argument(
        "--validate-only", action="store_true", help="Only validate the import"
    )
    parser.add_argument("--history", action="store_true", help="Show import history")
    parser.add_argument("--cleanup", action="store_true", help="Cleanup failed imports")
    parser.add_argument("--verbose", action="store_true", help="Enable verbose logging")

    args = parser.parse_args()

    # Setup logging - default to WARNING to reduce noise
    level = logging.DEBUG if args.verbose else logging.WARNING
    logging.basicConfig(level=level, format="%(levelname)s: %(message)s")

    if not Path(args.db_path).exists():
        print(f"Database not found: {args.db_path}")
        sys.exit(1)

    # Create service
    db_connection = DatabaseConnection(args.db_path)
    service = BroadcastMonthImportService(db_connection)

    try:
        if args.history:
            history = service.get_import_history()
            if history:
                print(f"Recent import history:")
                for record in history:
                    print(
                        f"  {record['batch_id']}: {record['import_mode']} - {record['status']} - {record['records_imported']} imported"
                    )
            else:
                print("No import history found")

        elif args.cleanup:
            cleaned = service.cleanup_failed_imports()
            print(f"Cleaned up {cleaned} failed import batches")

        elif args.validate_only:
            validation = service.validate_import(args.excel_file, args.mode)
            if validation.is_valid:
                print(f"Validation passed for {args.mode} mode")
                if validation.closed_months_found:
                    print(f"Includes closed months: {validation.closed_months_found}")
            else:
                print(f"Validation failed: {validation.error_message}")
                print(f"Solution: {validation.suggested_action}")

        else:
            # Execute import with clean output
            result = service.execute_month_replacement(
                args.excel_file, args.mode, args.closed_by, args.dry_run
            )

            print(f"\nImport Results:")
            print(f"  Status: {'Success' if result.success else 'Failed'}")
            print(f"  Batch ID: {result.batch_id}")
            print(f"  Mode: {result.import_mode}")
            print(f"  Duration: {result.duration_seconds:.1f} seconds")
            print(f"  Months: {', '.join(result.broadcast_months_affected)}")
            print(f"  Records deleted: {result.records_deleted:,}")
            print(f"  Records imported: {result.records_imported:,}")

            if result.closed_months:
                print(f"  Months closed: {', '.join(result.closed_months)}")

            if result.error_messages:
                print(f"  Errors:")
                for error in result.error_messages:
                    print(f"    - {error}")

            if not result.success:
                sys.exit(1)

    finally:
        db_connection.close()
