#!/usr/bin/env python3
"""
Smart Filtered Month Import - Import only new/open months from complete workbook.
Respects already-closed months and only imports what's needed.
"""

import sys
import argparse
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Set, Dict, Any
from dataclasses import dataclass
from tqdm import tqdm

# Add src to path
sys.path.insert(0, str(Path(__file__).parent / "src"))

from src.database.connection import DatabaseConnection
from src.services.month_closure_service import MonthClosureService, MonthClosureError
from src.services.broadcast_month_import_service import BroadcastMonthImportService
from src.utils.broadcast_month_utils import extract_broadcast_months_from_excel
from src.services.base_service import BaseService

logger = logging.getLogger(__name__)


@dataclass
class FilteredImportResult:
    """Result of filtered import operation."""
    success: bool
    total_months_in_excel: int
    months_skipped_closed: List[str]
    months_imported: List[str]
    months_closed: List[str]
    records_imported: int
    duration_seconds: float
    error_messages: List[str]


class SmartFilteredImporter(BaseService):
    """Smart importer that only imports new/open months from complete workbook."""
    
    def __init__(self, db_connection: DatabaseConnection):
        super().__init__(db_connection)
        self.month_service = MonthClosureService(db_connection)
        self.import_service = BroadcastMonthImportService(db_connection)
    
    def analyze_workbook(self, excel_file: str, target_year: int = None) -> Dict[str, Any]:
        """
        Analyze workbook to determine what needs to be imported.
        
        Args:
            excel_file: Path to Excel workbook
            target_year: Optional year filter (e.g., 2025)
            
        Returns:
            Analysis results
        """
        print(f"Analyzing workbook: {excel_file}")
        
        # Extract all months from Excel
        all_months = list(extract_broadcast_months_from_excel(excel_file))
        
        # Filter by year if specified
        if target_year:
            filtered_months = []
            for month in all_months:
                try:
                    from src.utils.broadcast_month_utils import BroadcastMonthParser
                    parser = BroadcastMonthParser()
                    month_year = parser.extract_year_from_broadcast_month(month)
                    if month_year == target_year:
                        filtered_months.append(month)
                except:
                    continue
            all_months = filtered_months
        
        print(f"Found {len(all_months)} months in workbook: {sorted(all_months)}")
        
        # Check which months are already closed
        closed_months = self.month_service.get_closed_months(all_months)
        open_months = [month for month in all_months if month not in closed_months]
        
        print(f"Already closed: {len(closed_months)} months")
        for month in sorted(closed_months):
            print(f"   {month} - SKIP (already closed)")
        
        print(f"Open/new months: {len(open_months)} months")
        for month in sorted(open_months):
            print(f"   {month} - IMPORT")
        
        return {
            'total_months': len(all_months),
            'closed_months': closed_months,
            'open_months': open_months,
            'all_months': all_months
        }
    
    def _analyze_excel_structure(self, source_excel: str, target_months: List[str]) -> Dict[str, Any]:
        """
        Phase 1: Analyze Excel structure and gather statistics.
        
        Args:
            source_excel: Original Excel file
            target_months: Months to include in analysis
            
        Returns:
            Analysis results dict
        """
        print(f"Phase 1: Analyzing Excel structure...")
        
        from openpyxl import load_workbook
        from src.utils.broadcast_month_utils import BroadcastMonthParser
        
        parser = BroadcastMonthParser()
        
        # Load source workbook for analysis
        source_wb = load_workbook(source_excel, read_only=True, data_only=True)
        source_ws = source_wb.active
        
        # Find Month column
        header_row = list(source_ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        month_col_index = None
        
        for i, header in enumerate(header_row):
            if header and str(header).strip() == 'Month':
                month_col_index = i
                break
        
        if month_col_index is None:
            source_wb.close()
            raise ValueError("Month column not found in Excel file")
        
        # Get total rows for progress tracking
        total_rows = source_ws.max_row - 1  # Subtract header row
        
        # Analyze with progress bar
        month_stats = {}
        total_target_rows = 0
        
        with tqdm(total=total_rows, desc="Analyzing structure", unit=" rows") as pbar:
            for row_num, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True), start=2):
                pbar.update(1)
                
                if not any(cell for cell in row):
                    continue
                
                # Analyze month distribution
                if month_col_index < len(row):
                    month_value = row[month_col_index]
                    
                    if month_value is not None:
                        try:
                            display_month = parser.parse_excel_date_to_broadcast_month(month_value)
                            
                            if display_month not in month_stats:
                                month_stats[display_month] = 0
                            month_stats[display_month] += 1
                            
                            if display_month in target_months:
                                total_target_rows += 1
                                
                        except Exception:
                            continue
                
                # Update description periodically
                if row_num % 10000 == 0:
                    pbar.set_description(f"Analyzed {len(month_stats)} months")
        
        source_wb.close()
        
        analysis_result = {
            'header_row': header_row,
            'month_col_index': month_col_index,
            'total_rows': total_rows,
            'month_stats': month_stats,
            'target_row_estimate': total_target_rows,
            'target_months': target_months
        }
        
        print(f"Analysis complete:")
        print(f"   Total rows: {total_rows:,}")
        print(f"   Months found: {len(month_stats)} {sorted(month_stats.keys())}")
        print(f"   Target months: {len(target_months)} {target_months}")
        print(f"   Estimated target rows: {total_target_rows:,}")
        
        return analysis_result
    
    def _create_filtered_excel(self, source_excel: str, analysis: Dict[str, Any], temp_file: str) -> str:
        """
        Phase 2: Create filtered Excel file using analysis results.
        
        Args:
            source_excel: Original Excel file
            analysis: Results from analysis phase
            temp_file: Path for temporary filtered file
            
        Returns:
            Path to filtered Excel file
        """
        print(f"Phase 2: Creating filtered Excel...")
        
        from openpyxl import load_workbook, Workbook
        from src.utils.broadcast_month_utils import BroadcastMonthParser
        
        parser = BroadcastMonthParser()
        
        # Load source workbook for filtering
        source_wb = load_workbook(source_excel, read_only=True, data_only=True)
        source_ws = source_wb.active
        
        # Create new workbook
        target_wb = Workbook()
        target_ws = target_wb.active
        
        # Get analysis data
        header_row = analysis['header_row']
        month_col_index = analysis['month_col_index']
        total_rows = analysis['total_rows']
        target_months = analysis['target_months']
        
        # Copy header row
        target_ws.append(header_row)
        
        rows_copied = 0
        rows_skipped = 0
        
        # Filter with progress bar
        with tqdm(total=total_rows, desc="Filtering rows", unit=" rows") as pbar:
            for row_num, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True), start=2):
                pbar.update(1)
                
                if not any(cell for cell in row):
                    continue
                
                # Check if this row's month should be included
                if month_col_index < len(row):
                    month_value = row[month_col_index]
                    
                    if month_value is not None:
                        try:
                            display_month = parser.parse_excel_date_to_broadcast_month(month_value)
                            
                            if display_month in target_months:
                                target_ws.append(row)
                                rows_copied += 1
                            else:
                                rows_skipped += 1
                                
                        except Exception:
                            rows_skipped += 1
                            continue
                
                # Update progress description periodically
                if row_num % 5000 == 0:
                    pbar.set_description(f"Copied: {rows_copied:,} | Skipped: {rows_skipped:,}")
        
        # Save filtered workbook
        target_wb.save(temp_file)
        source_wb.close()
        target_wb.close()
        
        print(f"Filtered Excel created: {rows_copied:,} rows copied, {rows_skipped:,} rows skipped")
        return temp_file
    
    def create_filtered_excel(self, source_excel: str, target_months: List[str], temp_file: str) -> str:
        """
        Create a temporary Excel file containing only specified months using two-phase approach.
        
        Args:
            source_excel: Original Excel file
            target_months: Months to include
            temp_file: Path for temporary filtered file
            
        Returns:
            Path to filtered Excel file
        """
        try:
            # Phase 1: Analyze Excel structure
            analysis = self._analyze_excel_structure(source_excel, target_months)
            
            # Phase 2: Create filtered Excel
            return self._create_filtered_excel(source_excel, analysis, temp_file)
            
        except Exception as e:
            print(f"Error during Excel filtering: {str(e)}")
            raise
    
    def execute_filtered_import(self, 
                              excel_file: str, 
                              target_year: int,
                              closed_by: str,
                              dry_run: bool = False,
                              close_imported_months: bool = True) -> FilteredImportResult:
        """
        Execute smart filtered import.
        
        Args:
            excel_file: Path to complete Excel workbook
            target_year: Year to focus on (e.g., 2025)
            closed_by: Who is performing the import
            dry_run: If True, show what would happen without importing
            close_imported_months: If True, close imported months after import
            
        Returns:
            FilteredImportResult with details
        """
        start_time = datetime.now()
        
        result = FilteredImportResult(
            success=False,
            total_months_in_excel=0,
            months_skipped_closed=[],
            months_imported=[],
            months_closed=[],
            records_imported=0,
            duration_seconds=0.0,
            error_messages=[]
        )
        
        try:
            # Step 1: Analyze workbook
            analysis = self.analyze_workbook(excel_file, target_year)
            
            result.total_months_in_excel = analysis['total_months']
            result.months_skipped_closed = analysis['closed_months']
            
            if not analysis['open_months']:
                print("All months are already closed - no import needed!")
                result.success = True
                return result
            
            print(f"\nImport Plan:")
            print(f"  • Skip {len(analysis['closed_months'])} already-closed months")
            print(f"  • Import {len(analysis['open_months'])} open months: {analysis['open_months']}")
            
            if dry_run:
                print(f"\nDRY RUN - No changes would be made")
                result.success = True
                return result
            
            # Step 2: Create filtered Excel with only open months
            temp_excel = f"temp_filtered_{target_year}_{int(start_time.timestamp())}.xlsx"
            temp_path = Path(temp_excel)
            
            try:
                print(f"\nStarting two-phase filtering process...")
                filtered_file = self.create_filtered_excel(
                    excel_file, 
                    analysis['open_months'], 
                    temp_excel
                )
                
                # Step 3: Import filtered data
                print(f"\nStep 3: Importing filtered data...")
                import_result = self.import_service.execute_month_replacement(
                    filtered_file,
                    'HISTORICAL' if close_imported_months else 'MANUAL',
                    closed_by,
                    dry_run=False
                )
                
                if import_result.success:
                    result.months_imported = analysis['open_months']
                    result.records_imported = import_result.records_imported
                    result.months_closed = import_result.closed_months if close_imported_months else []
                    
                    print(f"Filtered import successful!")
                    print(f"   Records imported: {result.records_imported:,}")
                    print(f"   Months imported: {len(result.months_imported)} {result.months_imported}")
                    if result.months_closed:
                        print(f"   Months closed: {len(result.months_closed)} {result.months_closed}")
                    
                else:
                    result.error_messages.extend(import_result.error_messages)
                    
            finally:
                # Clean up temporary file
                if temp_path.exists():
                    temp_path.unlink()
                    print(f"Cleaned up temporary file")
            
            result.success = import_result.success if 'import_result' in locals() else False
            
        except Exception as e:
            error_msg = f"Filtered import failed: {str(e)}"
            result.error_messages.append(error_msg)
            logger.error(error_msg, exc_info=True)
        
        # Calculate duration
        end_time = datetime.now()
        result.duration_seconds = (end_time - start_time).total_seconds()
        
        return result


def main():
    """CLI interface for smart filtered import."""
    parser = argparse.ArgumentParser(
        description="Smart Filtered Month Import - Import only new months from complete workbook",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Import only new 2025 months from complete workbook
  python smart_import.py data/raw/2025_complete.xlsx --year 2025 --closed-by "Kurt"
  
  # Dry run to see what would be imported
  python smart_import.py data/raw/2025_complete.xlsx --year 2025 --closed-by "Kurt" --dry-run
  
  # Import without closing months (for testing)
  python smart_import.py data/raw/2025_complete.xlsx --year 2025 --closed-by "Kurt" --no-close
        """
    )
    
    parser.add_argument("excel_file", help="Path to complete Excel workbook")
    parser.add_argument("--year", type=int, required=True, help="Target year (e.g., 2025)")
    parser.add_argument("--closed-by", required=True, help="Name of person performing import")
    parser.add_argument("--db-path", default="data/database/production.db", help="Database path")
    parser.add_argument("--dry-run", action="store_true", help="Show what would be imported without doing it")
    parser.add_argument("--no-close", action="store_true", help="Don't close months after import")
    parser.add_argument("--verbose", action="store_true", help="Enable verbose logging")
    
    args = parser.parse_args()
    
    # Setup logging
    level = logging.DEBUG if args.verbose else logging.INFO
    logging.basicConfig(level=level, format='%(levelname)s: %(message)s')
    
    # Validate inputs
    if not Path(args.excel_file).exists():
        print(f"Excel file not found: {args.excel_file}")
        sys.exit(1)
    
    if not Path(args.db_path).exists():
        print(f"Database not found: {args.db_path}")
        sys.exit(1)
    
    print(f"Smart Filtered Import Tool")
    print(f"Excel file: {args.excel_file}")
    print(f"Target year: {args.year}")
    print(f"Closed by: {args.closed_by}")
    if args.dry_run:
        print(f"Mode: DRY RUN")
    print("=" * 60)
    
    # Execute filtered import
    db_connection = DatabaseConnection(args.db_path)
    importer = SmartFilteredImporter(db_connection)
    
    try:
        result = importer.execute_filtered_import(
            args.excel_file,
            args.year,
            args.closed_by,
            args.dry_run,
            close_imported_months=not args.no_close
        )
        
        # Display results
        print(f"\n" + "=" * 60)
        print(f"SMART FILTERED IMPORT {'PREVIEW' if args.dry_run else 'COMPLETED'}")
        print(f"=" * 60)
        
        print(f"Results:")
        print(f"  Success: {'✓' if result.success else '✗'}")
        print(f"  Duration: {result.duration_seconds:.2f} seconds")
        print(f"  Total months in Excel: {result.total_months_in_excel}")
        print(f"  Months skipped (closed): {len(result.months_skipped_closed)}")
        print(f"  Months imported: {len(result.months_imported)}")
        
        if not args.dry_run:
            print(f"  Records imported: {result.records_imported:,}")
            if result.months_closed:
                print(f"  Months closed: {len(result.months_closed)}")
        
        if result.months_skipped_closed:
            print(f"\nSkipped (already closed): {result.months_skipped_closed}")
        
        if result.months_imported:
            print(f"\nImported: {result.months_imported}")
        
        if result.months_closed:
            print(f"\nClosed: {result.months_closed}")
        
        if result.error_messages:
            print(f"\nErrors:")
            for error in result.error_messages:
                print(f"  • {error}")
        
        if result.success and not args.dry_run:
            print(f"\nSmart import completed successfully!")
            print(f"Only new/open months were imported - closed months were protected")
        
        sys.exit(0 if result.success else 1)
        
    except KeyboardInterrupt:
        print(f"\nImport cancelled by user")
        sys.exit(1)
    except Exception as e:
        print(f"Unexpected error: {e}")
        if args.verbose:
            import traceback
            traceback.print_exc()
        sys.exit(1)
    
    finally:
        db_connection.close()


if __name__ == "__main__":
    main()