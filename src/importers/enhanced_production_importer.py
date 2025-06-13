#!/usr/bin/env python3
"""
FIXED Enhanced production-grade Excel importer with batch tracking support.
Addresses the stalling issue in bulk import operations.
"""

import sys
import sqlite3
import logging
import time
import gc
import psutil
import os
from pathlib import Path
from datetime import datetime, date
from typing import Optional, List, Dict, Any, Tuple
from dataclasses import dataclass

# Add src to path
sys.path.insert(0, str(Path(__file__).parent.parent))

try:
    from openpyxl import load_workbook
    print("✓ openpyxl available")
except ImportError:
    print("❌ openpyxl not available. Run: uv add openpyxl")
    sys.exit(1)

# Progress tracking
class ProgressTracker:
    """Tracks and displays import progress without spamming CLI."""
    
    def __init__(self, total_records: int, update_interval: int = 1000):
        self.total_records = total_records
        self.update_interval = update_interval
        self.processed = 0
        self.imported = 0
        self.skipped = 0
        self.last_update = 0
        self.start_time = time.time()
        
        # Memory monitoring
        try:
            self.process = psutil.Process(os.getpid())
            self.initial_memory = self.process.memory_info().rss / 1024 / 1024
            self.peak_memory = self.initial_memory
            self.memory_monitoring = True
        except:
            self.memory_monitoring = False
    
    def update(self, imported: bool = True):
        """Update progress counters."""
        self.processed += 1
        if imported:
            self.imported += 1
        else:
            self.skipped += 1
        
        # Track peak memory if available
        if self.memory_monitoring:
            try:
                current_memory = self.process.memory_info().rss / 1024 / 1024
                if current_memory > self.peak_memory:
                    self.peak_memory = current_memory
                
                # Force garbage collection for large files
                if current_memory - self.initial_memory > 500 and self.processed % 10000 == 0:
                    gc.collect()
            except:
                pass
        
        # Update display at intervals
        if self.processed - self.last_update >= self.update_interval or self.processed == self.total_records:
            self._display_progress()
            self.last_update = self.processed
    
    def _display_progress(self):
        """Display progress without spamming."""
        elapsed = time.time() - self.start_time
        rate = self.processed / elapsed if elapsed > 0 else 0
        percent = (self.processed / self.total_records) * 100 if self.total_records > 0 else 0
        
        progress_text = f"\rProgress: {self.processed:,}/{self.total_records:,} ({percent:.1f}%) " \
                       f"| Imported: {self.imported:,} | Skipped: {self.skipped:,} " \
                       f"| Rate: {rate:.0f} records/sec"
        
        # Add memory info if available
        if self.memory_monitoring:
            try:
                current_memory = self.process.memory_info().rss / 1024 / 1024
                memory_used = current_memory - self.initial_memory
                progress_text += f" | Memory: {current_memory:.0f}MB (+{memory_used:.0f}MB)"
            except:
                pass
        
        print(progress_text, end='', flush=True)
        
        if self.processed == self.total_records:
            print()  # New line at completion


@dataclass
class ImportResults:
    """Comprehensive import results."""
    success: bool
    total_records: int
    records_processed: int
    records_imported: int
    records_skipped: int
    new_agencies_created: int
    new_customers_created: int
    customers_normalized: int
    duration_seconds: float
    records_per_second: float
    errors: List[str]
    error_summary: Dict[str, int]
    batch_id: Optional[str] = None
    broadcast_months_processed: List[str] = None
    
    def __post_init__(self):
        if self.broadcast_months_processed is None:
            self.broadcast_months_processed = []


class EnhancedProductionExcelImporter:
    """FIXED Production-grade Excel importer with comprehensive column handling."""
    
    # Complete flexible column mapping for both 2024 and 2025 Excel formats
    COLUMN_MAPPING = {
        # ===================================================================
        # CORE FIELDS - Same in both 2024 and 2025 formats
        # ===================================================================
        'Bill Code': 'bill_code',
        'Start Date': 'air_date',
        'End Date': 'end_date',
        'Time In': 'time_in',
        'Time out': 'time_out',
        'Length': 'length_seconds',
        'Comments': 'program',
        'Language': 'language_code',
        'Line': 'line_number',
        'Type': 'spot_type',
        'Make Good': 'make_good',
        'Spot Value': 'spot_value',
        'Month': 'broadcast_month',
        'Broker Fees': 'broker_fees',
        'Revenue Type': 'revenue_type',
        'Billing Type': 'billing_type',
        'Market': 'market_name',
        
        # ===================================================================
        # 2024 FORMAT COLUMNS
        # ===================================================================
        'Day(s)': 'day_of_week',
        'Media/Name/Program': 'media',
        'Format': 'format',
        'Units-Spot count': 'sequence_number',
        'Agency/Episode# or cut number': 'estimate',
        'Unit rate Gross': 'gross_rate',
        'Sales/rep com: revenue sharing': 'priority',
        'Station Net': 'station_net',
        'Sales Person': 'sales_person',
        'Agency?': 'agency_flag',
        'Affidavit?': 'affidavit_flag',
        'Notarize?': 'contract',
        
        # ===================================================================
        # 2025 FORMAT COLUMNS (Alternative mappings)
        # ===================================================================
        'Day': 'day_of_week',              # Alternative to 'Day(s)'
        'Show Name': 'media',              # Alternative to 'Media/Name/Program'
        'Show': 'format',                  # Alternative to 'Format'
        'Spots': 'sequence_number',        # Alternative to 'Units-Spot count'
        'Estimate': 'estimate',            # Alternative to 'Agency/Episode# or cut number'
        'Gross': 'gross_rate',             # Alternative to 'Unit rate Gross'
        ' Gross ': 'gross_rate',           # Handle potential spacing variations
        'Priority': 'priority',            # Alternative to 'Sales/rep com: revenue sharing'
        'Net': 'station_net',              # Alternative to 'Station Net'
        ' Net ': 'station_net',            # Handle potential spacing variations
        'AE': 'sales_person',              # Alternative to 'Sales Person'
        'Agency': 'agency_flag',           # Alternative to 'Agency?'
        'Affidavit': 'affidavit_flag',     # Alternative to 'Affidavit?'
        'Notarize': 'contract',            # Alternative to 'Notarize?'
        
        # ===================================================================
        # ADDITIONAL VARIATIONS (Common alternatives)
        # ===================================================================
        'Account Executive': 'sales_person',   # Another common variation
        'Salesperson': 'sales_person',         # Another common variation
        'Gross Rate': 'gross_rate',            # Another common variation
        'Net Rate': 'station_net',             # Another common variation
        'Air Date': 'air_date',                # Alternative to 'Start Date'
        'Date': 'air_date',                    # Simple date column
        'Program': 'media',                    # Alternative to show/media fields
        'Episode': 'estimate',                 # Alternative to estimate field
    }
    
    def __init__(self, database_path: str):
        """Initialize the FIXED enhanced production importer."""
        self.database_path = database_path
        self.column_indexes = {}
        self.stats = {
            'agencies_created': 0,
            'customers_created': 0,
            'customers_normalized': 0,
            'errors': [],
            'error_types': {},
            'broadcast_months_found': set()
        }
        
        # CRITICAL FIX: Pre-cache database lookups to avoid repeated queries
        self._agency_cache = {}
        self._customer_cache = {}
        self._market_cache = {}
        
        # CRITICAL FIX: Use WAL mode for better concurrency
        self._setup_database_for_bulk_import()

    def _setup_database_for_bulk_import(self):
        """CRITICAL FIX: Configure database for bulk import operations."""
        try:
            conn = sqlite3.connect(self.database_path)
            # Enable WAL mode for better concurrency
            conn.execute("PRAGMA journal_mode=WAL")
            # Increase cache size for better performance
            conn.execute("PRAGMA cache_size=10000")
            # Disable synchronous writes during bulk import
            conn.execute("PRAGMA synchronous=OFF")
            # Increase temp store size
            conn.execute("PRAGMA temp_store=MEMORY")
            conn.close()
            print("✓ Database configured for bulk import")
        except Exception as e:
            print(f"⚠️  Could not optimize database settings: {e}")

    def validate_and_confirm_column_mapping(self, worksheet, interactive=True):
        """
        Validate Excel columns against expected mappings and get user confirmation.
        
        Args:
            worksheet: Excel worksheet object
            interactive: If True, prompt user for confirmation/corrections
            
        Returns:
            bool: True if mapping is acceptable, False to abort
        """
        print("🔍 Analyzing Excel Column Structure...")
        
        # Get actual headers from Excel
        header_row = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        actual_headers = [str(header).strip() if header else '' for header in header_row]
        actual_headers = [h for h in actual_headers if h]  # Remove empty headers
        
        print(f"📋 Found {len(actual_headers)} columns in Excel file")
        
        # Check which columns will be mapped
        mapped_columns = []
        unmapped_columns = []
        missing_critical_columns = []
        
        for header in actual_headers:
            if header in self.COLUMN_MAPPING:
                mapped_columns.append(header)
            else:
                unmapped_columns.append(header)
        
        # Check for critical missing columns
        critical_columns = ['Bill Code', 'Start Date', 'Month']  # Core required columns
        alternative_mappings = {
            'Start Date': ['Air Date', 'Date'],
            'Month': ['Broadcast Month', 'Period'],
            'Sales Person': ['AE', 'Account Executive', 'Salesperson'],
            'Unit rate Gross': ['Gross', 'Gross Rate', 'Rate'],
            'Station Net': ['Net', 'Net Rate']
        }
        
        for critical in critical_columns:
            if critical not in mapped_columns:
                # Check for alternatives
                found_alternative = False
                if critical in alternative_mappings:
                    for alt in alternative_mappings[critical]:
                        if alt in actual_headers:
                            found_alternative = True
                            break
                
                if not found_alternative:
                    missing_critical_columns.append(critical)
        
        # Display analysis
        print(f"\n📊 Column Mapping Analysis:")
        print(f"  ✅ Mapped columns: {len(mapped_columns)}")
        print(f"  ❓ Unmapped columns: {len(unmapped_columns)}")
        print(f"  ❌ Missing critical: {len(missing_critical_columns)}")
        
        if mapped_columns:
            print(f"\n✅ Successfully Mapped Columns:")
            for col in sorted(mapped_columns):
                db_field = self.COLUMN_MAPPING[col]
                print(f"  '{col}' → {db_field}")
        
        if unmapped_columns:
            print(f"\n❓ Unmapped Columns (will be ignored):")
            for col in sorted(unmapped_columns):
                print(f"  '{col}'")
        
        if missing_critical_columns:
            print(f"\n❌ Missing Critical Columns:")
            for col in missing_critical_columns:
                print(f"  '{col}' - Required for import")
                if col in alternative_mappings:
                    print(f"    Alternatives: {', '.join(alternative_mappings[col])}")
        
        # Interactive confirmation
        if interactive:
            print(f"\n🚨 COLUMN MAPPING CONFIRMATION")
            print(f"Excel file structure analysis complete.")
            
            if missing_critical_columns:
                print(f"❌ Cannot proceed: Missing critical columns {missing_critical_columns}")
                print(f"Please ensure your Excel file contains the required columns.")
                return False
            
            if unmapped_columns:
                print(f"⚠️  Warning: {len(unmapped_columns)} columns will be ignored during import.")
                print(f"This might indicate column naming differences between files.")
            
            print(f"\nProceed with import using current column mapping?")
            print(f"  - {len(mapped_columns)} columns will be imported")
            print(f"  - {len(unmapped_columns)} columns will be ignored")
            
            while True:
                response = input(f"\nContinue with import? (yes/no): ").strip().lower()
                if response in ['yes', 'y']:
                    return True
                elif response in ['no', 'n']:
                    print(f"❌ Import cancelled by user")
                    return False
                else:
                    print("Please enter 'yes' or 'no'")
        
        return len(missing_critical_columns) == 0
    
    def import_with_batch_id(self, 
                           excel_file_path: str, 
                           batch_id: Optional[str] = None,
                           limit: Optional[int] = None) -> ImportResults:
        """
        FIXED Enhanced import method with proper error handling and connection management.
        
        Args:
            excel_file_path: Path to Excel file
            batch_id: Optional batch ID for tracking (integrates with import service)
            limit: Optional limit on records to process
            
        Returns:
            ImportResults with batch and month information
        """
        print(f"🚀 Starting FIXED production import: {Path(excel_file_path).name}")
        if batch_id:
            print(f"📋 Batch ID: {batch_id}")
        
        start_time = time.time()
        
        # Validate files
        if not Path(excel_file_path).exists():
            return self._create_error_result(f"Excel file not found: {excel_file_path}")
        
        if not Path(self.database_path).exists():
            return self._create_error_result(f"Database not found: {self.database_path}")
        
        try:
            # Step 1: Load and analyze Excel file
            print("📊 Loading Excel file...")
            
            # Always use data_only=True for better performance
            workbook = load_workbook(excel_file_path, data_only=True)
            worksheet = workbook.active
            
            # Parse headers
            self._parse_headers(worksheet)
            
            # Count total rows
            total_rows = worksheet.max_row - 1
            if limit:
                total_rows = min(total_rows, limit)
            
            print(f"📈 Found {total_rows:,} records to process")
            print(f"📋 Mapped {len(self.column_indexes)} columns")
            
            # CRITICAL FIX: Pre-load caches before processing
            self._preload_database_caches()
            
            # Step 2: Process in single transaction for better performance
            progress = ProgressTracker(total_rows)
            
            # CRITICAL FIX: Use single transaction instead of chunked processing
            records_imported = self._process_sequential_fixed(worksheet, total_rows, batch_id, progress, limit)
            
            workbook.close()
            
            # Step 3: Generate results
            end_time = time.time()
            duration = end_time - start_time
            
            results = ImportResults(
                success=True,
                total_records=total_rows,
                records_processed=progress.processed,
                records_imported=progress.imported,
                records_skipped=progress.skipped,
                new_agencies_created=self.stats['agencies_created'],
                new_customers_created=self.stats['customers_created'],
                customers_normalized=self.stats['customers_normalized'],
                duration_seconds=duration,
                records_per_second=progress.processed / duration if duration > 0 else 0,
                errors=self.stats['errors'][:10],  # First 10 errors
                error_summary=self.stats['error_types'],
                batch_id=batch_id,
                broadcast_months_processed=sorted(list(self.stats['broadcast_months_found']))
            )
            
            self._print_final_results(results)
            return results
            
        except Exception as e:
            return self._create_error_result(f"Import failed: {str(e)}")
    
    def _preload_database_caches(self):
        """CRITICAL FIX: Pre-load database caches to avoid repeated queries."""
        try:
            conn = sqlite3.connect(self.database_path)
            conn.row_factory = sqlite3.Row
            
            # Pre-load agencies
            cursor = conn.execute("SELECT agency_id, agency_name FROM agencies")
            for row in cursor.fetchall():
                self._agency_cache[row['agency_name']] = row['agency_id']
            
            # Pre-load customers
            cursor = conn.execute("SELECT customer_id, normalized_name FROM customers")
            for row in cursor.fetchall():
                self._customer_cache[row['normalized_name']] = row['customer_id']
            
            # Pre-load markets
            cursor = conn.execute("SELECT market_id, market_name FROM markets")
            for row in cursor.fetchall():
                self._market_cache[row['market_name'].lower()] = row['market_id']
            
            conn.close()
            
            print(f"✓ Pre-loaded caches: {len(self._agency_cache)} agencies, {len(self._customer_cache)} customers, {len(self._market_cache)} markets")
            
        except Exception as e:
            print(f"⚠️  Could not pre-load caches: {e}")
    
    def _parse_headers(self, worksheet, interactive=False):
        """Parse and validate Excel headers."""
        
        # First run the validation
        if not self.validate_and_confirm_column_mapping(worksheet, interactive):
            raise ValueError("Column mapping validation failed or cancelled by user")
        
        # Original header parsing logic (now with validation passed)
        header_row = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        
        self.column_indexes = {}
        missing_columns = []
        
        for col_idx, header in enumerate(header_row):
            if header and str(header).strip():
                clean_header = str(header).strip()
                if clean_header in self.COLUMN_MAPPING:
                    field_name = self.COLUMN_MAPPING[clean_header]
                    self.column_indexes[field_name] = col_idx
        
        # Check for required columns after mapping
        required_fields = ['bill_code', 'air_date']
        for field in required_fields:
            if field not in self.column_indexes:
                original_name = next((k for k, v in self.COLUMN_MAPPING.items() if v == field), field)
                missing_columns.append(original_name)
        
        if missing_columns:
            raise ValueError(f"Missing required columns after mapping: {', '.join(missing_columns)}")
        
        print(f"✅ Column mapping validated and confirmed")
        print(f"📋 Mapped {len(self.column_indexes)} columns for import")
    
    def _process_sequential_fixed(self, worksheet, total_rows: int, batch_id: Optional[str], progress: ProgressTracker, limit: Optional[int]) -> int:
        """CRITICAL FIX: Process rows in single transaction with proper error handling."""
        
        # CRITICAL FIX: Use single database connection for entire import
        db_conn = sqlite3.connect(self.database_path)
        db_conn.execute("PRAGMA foreign_keys = ON")
        db_conn.execute("PRAGMA journal_mode=WAL")  # Better for concurrent access
        
        try:
            # CRITICAL FIX: Single large transaction
            db_conn.execute("BEGIN IMMEDIATE")
            
            # CRITICAL FIX: Prepare INSERT statement once
            insert_query = self._prepare_insert_statement()
            
            row_count = 0
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                if limit and row_count >= limit:
                    break
                
                if not any(cell for cell in row):
                    continue
                
                # CRITICAL FIX: Use prepared statement with better error handling
                success = self._process_row_fixed(row, row_num, db_conn, insert_query, batch_id)
                progress.update(imported=success)
                row_count += 1
                
                # CRITICAL FIX: Commit periodically for large imports
                if row_count % 10000 == 0:
                    db_conn.commit()
                    db_conn.execute("BEGIN IMMEDIATE")
                    print(f"\n✓ Committed {row_count:,} records...")
            
            # Final commit
            db_conn.commit()
            print(f"\n✅ Final commit: {progress.imported:,} records imported successfully")
            
        except Exception as e:
            db_conn.rollback()
            print(f"\n❌ Transaction rolled back: {e}")
            raise
        finally:
            db_conn.close()
        
        return progress.imported
    
    def _prepare_insert_statement(self):
        """CRITICAL FIX: Prepare INSERT statement with known columns."""
        # Define the exact columns we want to insert
        columns = [
            'bill_code', 'air_date', 'end_date', 'day_of_week', 'time_in', 'time_out',
            'length_seconds', 'media', 'program', 'language_code', 'format', 
            'sequence_number', 'line_number', 'spot_type', 'estimate',
            'gross_rate', 'make_good', 'spot_value', 'broadcast_month', 'broker_fees',
            'priority', 'station_net', 'sales_person', 'revenue_type', 'billing_type',
            'agency_flag', 'affidavit_flag', 'contract', 'market_name',
            'customer_id', 'agency_id', 'market_id', 'language_id',
            'source_file', 'load_date', 'is_historical', 'import_batch_id'
        ]
        
        placeholders = ', '.join(['?' for _ in columns])
        field_names = ', '.join(columns)
        
        return f"INSERT INTO spots ({field_names}) VALUES ({placeholders})", columns
    
    def _process_row_fixed(self, row: tuple, row_num: int, db_conn: sqlite3.Connection, insert_query: tuple, batch_id: Optional[str] = None) -> bool:
        """
        CRITICAL FIX: Process a single row with proper error handling and prepared statements.
        
        Args:
            row: Excel row data
            row_num: Row number for error reporting
            db_conn: Database connection
            insert_query: Prepared INSERT statement tuple (query, columns)
            batch_id: Optional batch ID for tracking
            
        Returns:
            True if row was successfully imported
        """
        try:
            query, columns = insert_query
            
            # Extract all fields using column mapping
            spot_data = {}
            
            for field_name, col_idx in self.column_indexes.items():
                if col_idx < len(row):
                    raw_value = row[col_idx]
                    converted_value = self._convert_value(field_name, raw_value)
                    spot_data[field_name] = converted_value
            
            # Validate required fields
            if not spot_data.get('bill_code') or not spot_data.get('air_date'):
                return False
            
            # Track broadcast months found
            if spot_data.get('broadcast_month'):
                try:
                    from utils.broadcast_month_utils import BroadcastMonthParser
                    parser = BroadcastMonthParser()
                    display_month = parser.parse_excel_date_to_broadcast_month(spot_data['broadcast_month'])
                    self.stats['broadcast_months_found'].add(display_month)
                except:
                    pass  # Don't fail import on month parsing
            
            # Parse bill code for agency/customer
            agency_name, customer_name = self._parse_bill_code(spot_data['bill_code'])
            
            # CRITICAL FIX: Use cached lookups instead of database queries
            if agency_name:
                spot_data['agency_id'] = self._get_or_create_agency_cached(agency_name)
            
            spot_data['customer_id'] = self._get_or_create_customer_cached(customer_name)
            
            # Handle market mapping with cache
            if spot_data.get('market_name'):
                spot_data['market_id'] = self._get_market_id_cached(spot_data['market_name'])
            
            # Add metadata
            spot_data['source_file'] = Path(self.database_path).name
            spot_data['load_date'] = datetime.now()
            spot_data['is_historical'] = 0
            spot_data['import_batch_id'] = batch_id
            
            # CRITICAL FIX: Build values array in exact column order
            values = []
            for col in columns:
                values.append(spot_data.get(col))
            
            # CRITICAL FIX: Execute with proper error handling
            try:
                db_conn.execute(query, values)
                return True
            except sqlite3.IntegrityError as e:
                self._record_error(f"Row {row_num}", f"Database integrity error: {e}")
                return False
            except sqlite3.Error as e:
                self._record_error(f"Row {row_num}", f"Database error: {e}")
                return False
            
        except Exception as e:
            self._record_error(f"Row {row_num}", f"Processing error: {e}")
            return False
    
    def _get_or_create_agency_cached(self, agency_name: str) -> int:
        """CRITICAL FIX: Get or create agency using cache."""
        if agency_name in self._agency_cache:
            return self._agency_cache[agency_name]
        
        # Create new agency
        try:
            conn = sqlite3.connect(self.database_path)
            cursor = conn.execute("INSERT INTO agencies (agency_name) VALUES (?)", (agency_name,))
            agency_id = cursor.lastrowid
            conn.commit()
            conn.close()
            
            self._agency_cache[agency_name] = agency_id
            self.stats['agencies_created'] += 1
            return agency_id
        except sqlite3.IntegrityError:
            # Another process may have created it
            conn = sqlite3.connect(self.database_path)
            cursor = conn.execute("SELECT agency_id FROM agencies WHERE agency_name = ?", (agency_name,))
            row = cursor.fetchone()
            conn.close()
            if row:
                agency_id = row[0]
                self._agency_cache[agency_name] = agency_id
                return agency_id
            raise
    
    def _get_or_create_customer_cached(self, customer_name: str) -> int:
        """CRITICAL FIX: Get or create customer using cache."""
        if customer_name in self._customer_cache:
            return self._customer_cache[customer_name]
        
        # Create new customer
        try:
            conn = sqlite3.connect(self.database_path)
            cursor = conn.execute("INSERT INTO customers (normalized_name) VALUES (?)", (customer_name,))
            customer_id = cursor.lastrowid
            conn.commit()
            conn.close()
            
            self._customer_cache[customer_name] = customer_id
            self.stats['customers_created'] += 1
            return customer_id
        except sqlite3.IntegrityError:
            # Another process may have created it
            conn = sqlite3.connect(self.database_path)
            cursor = conn.execute("SELECT customer_id FROM customers WHERE normalized_name = ?", (customer_name,))
            row = cursor.fetchone()
            conn.close()
            if row:
                customer_id = row[0]
                self._customer_cache[customer_name] = customer_id
                return customer_id
            raise
    
    def _get_market_id_cached(self, market_name: str) -> Optional[int]:
        """CRITICAL FIX: Get market ID using cache."""
        market_key = market_name.lower()
        return self._market_cache.get(market_key)
    
    # [Keep all the other methods the same: _convert_value, _parse_bill_code, etc.]
    
    def _convert_value(self, field_name: str, raw_value: Any) -> Any:
        """Convert values with proper type handling for SQLite."""
        if raw_value is None or (isinstance(raw_value, str) and not raw_value.strip()):
            return None
        
        try:
            # Date fields
            if field_name in ['air_date', 'end_date']:
                return self._convert_date(raw_value)
            
            # Financial fields (use float for SQLite, allow negatives)
            elif field_name in ['gross_rate', 'station_net', 'spot_value', 'broker_fees']:
                return self._convert_float(raw_value)
            
            # Integer fields
            elif field_name in ['sequence_number', 'line_number', 'priority']:
                return self._convert_integer(raw_value)
            
            # String fields
            else:
                return self._convert_string(raw_value)
                
        except Exception:
            return None
    
    def _convert_date(self, value: Any) -> Optional[date]:
        """Convert date values."""
        if isinstance(value, date):
            return value
        if isinstance(value, datetime):
            return value.date()
        
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None
            
            for fmt in ['%m/%d/%y', '%m/%d/%Y', '%Y-%m-%d']:
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue
        
        return None
    
    def _convert_float(self, value: Any) -> Optional[float]:
        """Convert to float, allowing negative values."""
        if isinstance(value, (int, float)):
            return float(value)
        
        if isinstance(value, str):
            value = value.strip().replace('$', '').replace(',', '')
            if value.startswith('(') and value.endswith(')'):
                value = '-' + value[1:-1]
            
            try:
                return float(value)
            except ValueError:
                return None
        
        return None
    
    def _convert_integer(self, value: Any) -> Optional[int]:
        """Convert to integer."""
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            return int(value)
        if isinstance(value, str) and value.strip():
            try:
                return int(float(value.strip()))
            except ValueError:
                return None
        return None
    
    def _convert_string(self, value: Any) -> Optional[str]:
        """Convert to clean string."""
        if isinstance(value, str):
            cleaned = value.strip()
            return cleaned if cleaned else None
        return str(value) if value is not None else None
    
    def _parse_bill_code(self, bill_code: str) -> Tuple[Optional[str], str]:
        """Parse bill code into agency and customer."""
        if ':' in bill_code:
            parts = bill_code.split(':', 1)
            agency = parts[0].strip()
            customer = parts[1].strip()
            
            # Remove PRODUCTION suffixes
            for suffix in [' PRODUCTION', ' Production', ' PROD']:
                if customer.endswith(suffix):
                    customer = customer[:-len(suffix)].strip()
                    self.stats['customers_normalized'] += 1
                    break
            
            return (agency, customer)
        else:
            customer = bill_code.strip()
            for suffix in [' PRODUCTION', ' Production', ' PROD']:
                if customer.endswith(suffix):
                    customer = customer[:-len(suffix)].strip()
                    self.stats['customers_normalized'] += 1
                    break
            
            return (None, customer)
    
    def _record_error(self, context: str, error: str):
        """Record error with categorization."""
        full_error = f"{context}: {error}"
        self.stats['errors'].append(full_error)
        
        # Categorize error type
        error_type = 'ProcessingError'
        if 'integrity' in error.lower():
            error_type = 'IntegrityError'
        elif 'foreign key' in error.lower():
            error_type = 'ForeignKeyError'
        elif 'database' in error.lower():
            error_type = 'DatabaseError'
        
        self.stats['error_types'][error_type] = self.stats['error_types'].get(error_type, 0) + 1
    
    def _create_error_result(self, error_message: str) -> ImportResults:
        """Create error result."""
        return ImportResults(
            success=False,
            total_records=0,
            records_processed=0,
            records_imported=0,
            records_skipped=0,
            new_agencies_created=0,
            new_customers_created=0,
            customers_normalized=0,
            duration_seconds=0,
            records_per_second=0,
            errors=[error_message],
            error_summary={},
            batch_id=None,
            broadcast_months_processed=[]
        )
    
    def _print_final_results(self, results: ImportResults):
        """Print comprehensive final results."""
        print(f"\n{'='*60}")
        print(f"🎉 FIXED PRODUCTION IMPORT COMPLETED")
        print(f"{'='*60}")
        print(f"📊 Performance Metrics:")
        print(f"  Duration: {results.duration_seconds:.2f} seconds")
        print(f"  Processing Rate: {results.records_per_second:,.0f} records/second")
        
        if results.batch_id:
            print(f"  Batch ID: {results.batch_id}")
        
        print(f"")
        print(f"📈 Import Results:")
        print(f"  Total Records: {results.total_records:,}")
        print(f"  Successfully Imported: {results.records_imported:,}")
        print(f"  Skipped: {results.records_skipped:,}")
        print(f"  Success Rate: {(results.records_imported/results.total_records)*100:.1f}%")
        print(f"")
        print(f"🏢 Data Creation:")
        print(f"  New Agencies: {results.new_agencies_created}")
        print(f"  New Customers: {results.new_customers_created}")
        print(f"  Customer Names Normalized: {results.customers_normalized}")
        
        # Show broadcast months processed
        if results.broadcast_months_processed:
            print(f"")
            print(f"📅 Broadcast Months Processed:")
            print(f"  Months: {', '.join(results.broadcast_months_processed)}")
            print(f"  Total: {len(results.broadcast_months_processed)} months")
        
        if results.records_skipped > 0:
            print(f"\n⚠️  Error Analysis:")
            print(f"  Total Errors: {results.records_skipped}")
            
            if results.error_summary:
                print(f"  Error Types:")
                for error_type, count in sorted(results.error_summary.items(), key=lambda x: x[1], reverse=True):
                    percentage = (count / results.records_skipped) * 100
                    print(f"    {error_type}: {count} ({percentage:.1f}%)")
            
            if results.errors:
                print(f"\n  Sample Errors (first 5):")
                for error in results.errors[:5]:
                    print(f"    • {error}")


# Integration function for import service - FIXED
def import_excel_with_batch(excel_file_path: str, database_path: str, batch_id: str) -> int:
    """
    FIXED Integration function for import service.
    
    Args:
        excel_file_path: Path to Excel file
        database_path: Path to database
        batch_id: Batch ID for tracking
        
    Returns:
        Number of records imported
        
    Raises:
        Exception: If import fails
    """
    importer = EnhancedProductionExcelImporter(database_path)
    result = importer.import_with_batch_id(excel_file_path, batch_id)
    
    if not result.success:
        error_msg = "; ".join(result.errors) if result.errors else "Import failed"
        raise Exception(f"Excel import failed: {error_msg}")
    
    return result.records_imported


def main():
    """CLI interface for FIXED enhanced production importer."""
    import argparse
    
    parser = argparse.ArgumentParser(description="FIXED Enhanced Production Excel Importer")
    parser.add_argument("excel_file", help="Path to Excel file")
    parser.add_argument("--database", default="data/database/production.db", help="Database path")
    parser.add_argument("--batch-id", help="Optional batch ID for tracking")
    parser.add_argument("--limit", type=int, help="Limit number of records (for testing)")
    
    args = parser.parse_args()
    
    # Check system memory
    try:
        import psutil
        available_memory_gb = psutil.virtual_memory().available / (1024**3)
        print(f"💾 Available system memory: {available_memory_gb:.1f} GB")
    except:
        print("💾 Memory monitoring not available (psutil not installed)")
    
    importer = EnhancedProductionExcelImporter(args.database)
    results = importer.import_with_batch_id(args.excel_file, args.batch_id, args.limit)
    
    if results.success:
        print(f"\n✅ FIXED Import completed successfully!")
        sys.exit(0)
    else:
        print(f"\n❌ FIXED Import failed!")
        for error in results.errors:
            print(f"  {error}")
        sys.exit(1)


if __name__ == "__main__":
    main()