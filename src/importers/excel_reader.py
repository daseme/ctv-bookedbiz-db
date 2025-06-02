"""
Excel file reader for booked business reports.
Handles parsing Excel files into Spot objects with proper data type conversion.
"""

import logging
from pathlib import Path
from typing import List, Iterator, Optional, Dict, Any
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    raise ImportError("openpyxl is required. Install with: uv add openpyxl")

# Handle imports for both package use and direct execution
try:
    from ..models.entities import Spot
except ImportError:
    # Fall back to absolute import for direct execution
    import sys
    from pathlib import Path
    sys.path.insert(0, str(Path(__file__).parent.parent))
    from models.entities import Spot

logger = logging.getLogger(__name__)


class ExcelReadError(Exception):
    """Raised when there's an error reading the Excel file."""
    pass


class ExcelReader:
    """Reads booked business Excel files and converts to Spot objects."""
    
    # Expected column mappings from Excel to our database fields
    COLUMN_MAPPING = {
        # Your actual Excel column names → our field names
        'Bill Code': 'bill_code',
        'Start Date': 'air_date',  # Your column is "Start Date" not "Air Date"
        'End Date': 'end_date',
        'Day(s)': 'day_of_week',
        'Time In': 'time_in',
        'Time out': 'time_out',  # Note: lowercase 'o' in "out"
        'Length': 'length_seconds',
        'Media/Name/Program': 'media',  # Combined field
        'Comments': 'program',  # Using Comments for program field
        'Language': 'language_code',  # Your column is "Language" not "Lang."
        'Format': 'format',
        'Units-Spot count': 'sequence_number',  # Using this for sequence number
        'Line': 'line_number',
        'Type': 'spot_type',
        'Agency/Episode# or cut number': 'estimate',  # Using this field for estimate
        'Unit rate Gross': 'gross_rate',  # Your column is "Unit rate" not "Gross Rate"
        'Make Good': 'make_good',
        'Spot Value': 'spot_value',  # If you have both Gross and Spot Value
        'Month': 'broadcast_month',
        'Broker Fees': 'broker_fees',
        'Sales/rep com: revenue sharing': 'priority',  # Using this field for priority
        'Station Net': 'station_net',
        'Sales Person': 'sales_person',
        'Revenue Type': 'revenue_type',
        'Billing Type': 'billing_type',
        'Agency?': 'agency_flag',
        'Affidavit?': 'affidavit_flag',
        'Notarize?': 'contract',  # Using Notarize for contract field
        'Market': 'market_name'
    }
    
    def __init__(self, file_path: str):
        """Initialize Excel reader with file path."""
        self.file_path = Path(file_path)
        self.workbook = None
        self.worksheet = None
        self.headers = {}
        self.total_rows = 0
        
    def __enter__(self):
        """Context manager entry - load the workbook."""
        self.load_workbook()
        return self
        
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit - close the workbook."""
        if self.workbook:
            self.workbook.close()
    
    def load_workbook(self):
        """Load and validate the Excel workbook."""
        if not self.file_path.exists():
            raise ExcelReadError(f"File not found: {self.file_path}")
        
        if not self.file_path.suffix.lower() in ['.xlsx', '.xlsm', '.xls']:
            raise ExcelReadError(f"Invalid file type: {self.file_path.suffix}")
        
        try:
            logger.info(f"Loading Excel file: {self.file_path}")
            self.workbook = load_workbook(self.file_path, read_only=True, data_only=True)
            
            # Use the first worksheet
            self.worksheet = self.workbook.active
            
            # Parse headers
            self._parse_headers()
            
            # Count total rows (excluding header)
            self.total_rows = self.worksheet.max_row - 1
            
            logger.info(f"Loaded workbook with {self.total_rows} data rows")
            
        except Exception as e:
            raise ExcelReadError(f"Failed to load Excel file: {str(e)}")
    
    def _parse_headers(self):
        """Parse and validate column headers."""
        if not self.worksheet:
            raise ExcelReadError("Workbook not loaded")
        
        # Get headers from first row
        header_row = list(self.worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        
        # Create mapping of column index to field name
        self.headers = {}
        missing_columns = []
        
        for col_idx, header in enumerate(header_row):
            if header and str(header).strip():
                clean_header = str(header).strip()
                if clean_header in self.COLUMN_MAPPING:
                    field_name = self.COLUMN_MAPPING[clean_header]
                    self.headers[col_idx] = field_name
                    logger.debug(f"Mapped column {col_idx}: '{clean_header}' → {field_name}")
        
        # Check for required columns
        required_fields = ['bill_code', 'air_date']
        mapped_fields = set(self.headers.values())
        
        for field in required_fields:
            if field not in mapped_fields:
                # Find the original column name
                original_name = next((k for k, v in self.COLUMN_MAPPING.items() if v == field), field)
                missing_columns.append(original_name)
        
        if missing_columns:
            raise ExcelReadError(f"Missing required columns: {', '.join(missing_columns)}")
        
        logger.info(f"Successfully mapped {len(self.headers)} columns")
    
    def read_spots(self) -> Iterator[Spot]:
        """Read all spots from the Excel file as an iterator."""
        if not self.worksheet:
            raise ExcelReadError("Workbook not loaded")
        
        logger.info(f"Reading {self.total_rows} spots from Excel file")
        
        # Skip header row (start from row 2)
        for row_num, row in enumerate(self.worksheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Skip empty rows
                if not any(cell for cell in row):
                    continue
                
                # Convert row to Spot object
                spot = self._row_to_spot(row, row_num)
                if spot:
                    yield spot
                    
            except Exception as e:
                logger.warning(f"Error processing row {row_num}: {str(e)}")
                # Continue with next row instead of failing completely
                continue
    
    def read_all_spots(self) -> List[Spot]:
        """Read all spots from the Excel file as a list."""
        return list(self.read_spots())
    
    def _row_to_spot(self, row: tuple, row_num: int) -> Optional[Spot]:
        """Convert a row tuple to a Spot object."""
        try:
            # Extract values based on header mapping
            values = {}
            
            for col_idx, field_name in self.headers.items():
                if col_idx < len(row):
                    raw_value = row[col_idx]
                    converted_value = self._convert_value(field_name, raw_value)
                    values[field_name] = converted_value
            
            # Check if we have required fields
            if not values.get('bill_code') or not values.get('air_date'):
                logger.debug(f"Row {row_num}: Missing required fields, skipping")
                return None
            
            # Create Spot object with all available fields
            spot = Spot(
                bill_code=values.get('bill_code', ''),
                air_date=values.get('air_date', date.today()),
                end_date=values.get('end_date'),
                day_of_week=values.get('day_of_week'),
                time_in=values.get('time_in'),
                time_out=values.get('time_out'),
                length_seconds=values.get('length_seconds'),
                media=values.get('media'),
                program=values.get('program'),
                language_code=values.get('language_code'),
                format=values.get('format'),
                sequence_number=values.get('sequence_number'),
                line_number=values.get('line_number'),
                spot_type=values.get('spot_type'),
                estimate=values.get('estimate'),
                gross_rate=values.get('gross_rate'),
                make_good=values.get('make_good'),
                spot_value=values.get('spot_value'),
                broadcast_month=values.get('broadcast_month'),
                broker_fees=values.get('broker_fees'),
                priority=values.get('priority'),
                station_net=values.get('station_net'),
                sales_person=values.get('sales_person'),
                revenue_type=values.get('revenue_type'),
                billing_type=values.get('billing_type'),
                agency_flag=values.get('agency_flag'),
                affidavit_flag=values.get('affidavit_flag'),
                contract=values.get('contract'),
                market_name=values.get('market_name'),
                source_file=str(self.file_path.name)
            )
            
            return spot
            
        except Exception as e:
            logger.error(f"Failed to convert row {row_num} to Spot: {str(e)}")
            raise
    
    def _convert_value(self, field_name: str, raw_value: Any) -> Any:
        """Convert raw Excel value to appropriate Python type."""
        if raw_value is None:
            return None
        
        # Handle empty strings
        if isinstance(raw_value, str) and not raw_value.strip():
            return None
        
        try:
            # Date fields
            if field_name in ['air_date', 'end_date', 'effective_date']:
                return self._convert_date(raw_value)
            
            # Decimal/currency fields
            elif field_name in ['gross_rate', 'station_net', 'spot_value', 'broker_fees']:
                return self._convert_decimal(raw_value)
            
            # Integer fields
            elif field_name in ['sequence_number', 'line_number', 'priority']:
                return self._convert_integer(raw_value)
            
            # String fields - just clean up
            else:
                return self._convert_string(raw_value)
                
        except Exception as e:
            logger.warning(f"Failed to convert {field_name} value '{raw_value}': {str(e)}")
            return None
    
    def _convert_date(self, value: Any) -> Optional[date]:
        """Convert various date formats to Python date."""
        if value is None:
            return None
        
        # If it's already a date or datetime
        if isinstance(value, date):
            return value
        if isinstance(value, datetime):
            return value.date()
        
        # If it's a string, try to parse it
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None
            
            # Try common date formats
            for fmt in ['%m/%d/%y', '%m/%d/%Y', '%Y-%m-%d', '%m-%d-%y', '%m-%d-%Y']:
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue
        
        # If it's a number (Excel date serial)
        if isinstance(value, (int, float)):
            try:
                # Excel date serial (days since 1900-01-01, with some quirks)
                from datetime import timedelta
                excel_epoch = date(1900, 1, 1)
                # Excel incorrectly treats 1900 as a leap year, so subtract 2 days
                return excel_epoch + timedelta(days=int(value) - 2)
            except (ValueError, OverflowError):
                pass
        
        logger.warning(f"Could not convert date value: {value} (type: {type(value)})")
        return None
    
    def _convert_decimal(self, value: Any) -> Optional[Decimal]:
        """Convert value to Decimal for currency fields."""
        if value is None:
            return None

        if isinstance(value, (int, float)):
            return Decimal(str(value))

        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None

            cleaned = value.replace('$', '').replace(',', '').strip()
            if cleaned.startswith('(') and cleaned.endswith(')'):
                cleaned = '-' + cleaned[1:-1]

            try:
                return Decimal(cleaned)
            except InvalidOperation:
                logger.warning(f"Could not convert decimal value after cleaning: '{cleaned}' (original: '{value}')")
                return None

        logger.warning(f"Could not convert decimal value: {value} (type: {type(value)})")
        return None

    
    def _convert_integer(self, value: Any) -> Optional[int]:
        """Convert value to integer."""
        if value is None:
            return None
        
        if isinstance(value, int):
            return value
        
        if isinstance(value, float):
            return int(value)
        
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None
            
            try:
                return int(float(value))  # Handle "1.0" strings
            except ValueError:
                pass
        
        logger.warning(f"Could not convert integer value: {value}")
        return None
    
    def _convert_string(self, value: Any) -> Optional[str]:
        """Convert value to clean string."""
        if value is None:
            return None
        
        if isinstance(value, str):
            cleaned = value.strip()
            return cleaned if cleaned else None
        
        # Convert other types to string
        return str(value)
    
    def get_file_info(self) -> Dict[str, Any]:
        """Get information about the loaded file."""
        if not self.worksheet:
            return {}
        
        return {
            'file_path': str(self.file_path),
            'file_name': self.file_path.name,
            'total_rows': self.total_rows,
            'columns_mapped': len(self.headers),
            'worksheet_name': self.worksheet.title
        }


# Convenience function for simple usage
def read_excel_file(file_path: str) -> List[Spot]:
    """Simple function to read an Excel file and return list of Spots."""
    with ExcelReader(file_path) as reader:
        return reader.read_all_spots()


# Example usage
if __name__ == "__main__":
    import sys
    from pathlib import Path

    sys.path.insert(0, str(Path(__file__).parent.parent))

    from models.entities import Spot

    if len(sys.argv) != 2:
        print("Usage: python excel_reader.py <excel_file_path>")
        sys.exit(1)

    file_path = sys.argv[1]

    try:
        with ExcelReader(file_path) as reader:
            print(f"File info: {reader.get_file_info()}")

            spots = reader.read_all_spots()
            print(f"Successfully read {len(spots)} spots")

            for i, spot in enumerate(spots[:3]):
                print(f"\nSpot {i+1}:")
                print(f"  Bill Code: {spot.bill_code}")
                print(f"  Air Date: {spot.air_date}")
                print(f"  Gross Rate: {spot.gross_rate}")
                print(f"  Sales Person: {spot.sales_person}")

    except ExcelReadError as e:
        print(f"Excel reading error: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"Unexpected error: {e}")
        sys.exit(1)

    
    def _convert_integer(self, value: Any) -> Optional[int]:
        """Convert value to integer."""
        if value is None:
            return None
        
        if isinstance(value, int):
            return value
        
        if isinstance(value, float):
            return int(value)
        
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None
            
            try:
                return int(float(value))  # Handle "1.0" strings
            except ValueError:
                pass
        
        logger.warning(f"Could not convert integer value: {value}")
        return None
    
    def _convert_string(self, value: Any) -> Optional[str]:
        """Convert value to clean string."""
        if value is None:
            return None
        
        if isinstance(value, str):
            cleaned = value.strip()
            return cleaned if cleaned else None
        
        # Convert other types to string
        return str(value)
    
    def get_file_info(self) -> Dict[str, Any]:
        """Get information about the loaded file."""
        if not self.worksheet:
            return {}
        
        return {
            'file_path': str(self.file_path),
            'file_name': self.file_path.name,
            'total_rows': self.total_rows,
            'columns_mapped': len(self.headers),
            'worksheet_name': self.worksheet.title
        }


# Convenience function for simple usage
def read_excel_file(file_path: str) -> List[Spot]:
    """Simple function to read an Excel file and return list of Spots."""
    with ExcelReader(file_path) as reader:
        return reader.read_all_spots()


# Example usage
if __name__ == "__main__":
    import sys
    from pathlib import Path
    
    # Add src to path for direct execution
    sys.path.insert(0, str(Path(__file__).parent.parent))
    
    # Now import with absolute paths
    from models.entities import Spot
    
    if len(sys.argv) != 2:
        print("Usage: python excel_reader.py <excel_file_path>")
        sys.exit(1)
    
    file_path = sys.argv[1]
    
    try:
        with ExcelReader(file_path) as reader:
            print(f"File info: {reader.get_file_info()}")
            
            spots = reader.read_all_spots()
            print(f"Successfully read {len(spots)} spots")
            
            # Show first few spots
            for i, spot in enumerate(spots[:3]):
                print(f"\nSpot {i+1}:")
                print(f"  Bill Code: {spot.bill_code}")
                print(f"  Air Date: {spot.air_date}")
                print(f"  Gross Rate: {spot.gross_rate}")
                print(f"  Sales Person: {spot.sales_person}")
                
    except ExcelReadError as e:
        print(f"Excel reading error: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"Unexpected error: {e}")
        sys.exit(1)