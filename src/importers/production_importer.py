#!/usr/bin/env python3
"""
Production-grade Excel importer combining the best of both approaches.
Handles all 29 columns with robust error handling, progress tracking, and timing.
"""

import sys
import sqlite3
import logging
import time
from pathlib import Path
from datetime import datetime, date
from typing import Optional, List, Dict, Any, Tuple
from dataclasses import dataclass

# Add src to path
sys.path.insert(0, str(Path(__file__).parent / "src"))

try:
    from openpyxl import load_workbook
    print("✓ openpyxl available")
except ImportError:
    print("❌ openpyxl not available. Run: uv add openpyxl")
    sys.exit(1)

# Progress tracking
class ProgressTracker:
    """Tracks and displays import progress without spamming CLI."""
    
    def __init__(self, total_records: int, update_interval: int = 1000):
        self.total_records = total_records
        self.update_interval = update_interval
        self.processed = 0
        self.imported = 0
        self.skipped = 0
        self.last_update = 0
        self.start_time = time.time()
    
    def update(self, imported: bool = True):
        """Update progress counters."""
        self.processed += 1
        if imported:
            self.imported += 1
        else:
            self.skipped += 1
        
        # Update display at intervals
        if self.processed - self.last_update >= self.update_interval or self.processed == self.total_records:
            self._display_progress()
            self.last_update = self.processed
    
    def _display_progress(self):
        """Display progress without spamming."""
        elapsed = time.time() - self.start_time
        rate = self.processed / elapsed if elapsed > 0 else 0
        percent = (self.processed / self.total_records) * 100 if self.total_records > 0 else 0
        
        print(f"\rProgress: {self.processed:,}/{self.total_records:,} ({percent:.1f}%) "
              f"| Imported: {self.imported:,} | Skipped: {self.skipped:,} "
              f"| Rate: {rate:.0f} records/sec", end='', flush=True)
        
        if self.processed == self.total_records:
            print()  # New line at completion


@dataclass
class ImportResults:
    """Comprehensive import results."""
    success: bool
    total_records: int
    records_processed: int
    records_imported: int
    records_skipped: int
    new_agencies_created: int
    new_customers_created: int
    customers_normalized: int
    duration_seconds: float
    records_per_second: float
    errors: List[str]
    error_summary: Dict[str, int]


class ProductionExcelImporter:
    """Production-grade Excel importer with comprehensive column handling."""
    
    # Complete column mapping for all 29 Excel columns
    COLUMN_MAPPING = {
        'Bill Code': 'bill_code',
        'Start Date': 'air_date',
        'End Date': 'end_date',
        'Day(s)': 'day_of_week',
        'Time In': 'time_in',
        'Time out': 'time_out',
        'Length': 'length_seconds',
        'Media/Name/Program': 'media',
        'Comments': 'program',
        'Language': 'language_code',
        'Format': 'format',
        'Units-Spot count': 'sequence_number',
        'Line': 'line_number',
        'Type': 'spot_type',
        'Agency/Episode# or cut number': 'estimate',
        'Unit rate Gross': 'gross_rate',
        'Make Good': 'make_good',
        'Spot Value': 'spot_value',
        'Month': 'broadcast_month',
        'Broker Fees': 'broker_fees',
        'Sales/rep com: revenue sharing': 'priority',
        'Station Net': 'station_net',
        'Sales Person': 'sales_person',
        'Revenue Type': 'revenue_type',
        'Billing Type': 'billing_type',
        'Agency?': 'agency_flag',
        'Affidavit?': 'affidavit_flag',
        'Notarize?': 'contract',
        'Market': 'market_name'
    }
    
    def __init__(self, database_path: str):
        """Initialize the production importer."""
        self.database_path = database_path
        self.column_indexes = {}
        self.stats = {
            'agencies_created': 0,
            'customers_created': 0,
            'customers_normalized': 0,
            'errors': [],
            'error_types': {}
        }
    
    def import_excel_file(self, excel_file_path: str, limit: Optional[int] = None) -> ImportResults:
        """Import Excel file with full production features."""
        print(f"🚀 Starting production import: {Path(excel_file_path).name}")
        start_time = time.time()
        
        # Validate files
        if not Path(excel_file_path).exists():
            return self._create_error_result(f"Excel file not found: {excel_file_path}")
        
        if not Path(self.database_path).exists():
            return self._create_error_result(f"Database not found: {self.database_path}")
        
        try:
            # Step 1: Load and analyze Excel file
            print("📊 Loading Excel file...")
            workbook = load_workbook(excel_file_path, read_only=True, data_only=True)
            worksheet = workbook.active
            
            # Parse headers
            self._parse_headers(worksheet)
            
            # Count total rows
            total_rows = worksheet.max_row - 1
            if limit:
                total_rows = min(total_rows, limit)
            
            print(f"📈 Found {total_rows:,} records to process")
            print(f"📋 Mapped {len(self.column_indexes)} columns")
            
            # Step 2: Process with progress tracking
            progress = ProgressTracker(total_rows)
            
            db_conn = sqlite3.connect(self.database_path)
            db_conn.execute("PRAGMA foreign_keys = ON")
            
            try:
                db_conn.execute("BEGIN")
                
                row_count = 0
                for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                    if limit and row_count >= limit:
                        break
                    
                    if not any(cell for cell in row):
                        continue
                    
                    success = self._process_row(row, row_num, db_conn)
                    progress.update(imported=success)
                    row_count += 1
                
                db_conn.commit()
                print("✅ Transaction committed successfully")
                
            except Exception as e:
                db_conn.rollback()
                print(f"\n❌ Transaction rolled back: {e}")
                raise
            finally:
                db_conn.close()
                workbook.close()
            
            # Step 3: Generate results
            end_time = time.time()
            duration = end_time - start_time
            
            results = ImportResults(
                success=True,
                total_records=total_rows,
                records_processed=progress.processed,
                records_imported=progress.imported,
                records_skipped=progress.skipped,
                new_agencies_created=self.stats['agencies_created'],
                new_customers_created=self.stats['customers_created'],
                customers_normalized=self.stats['customers_normalized'],
                duration_seconds=duration,
                records_per_second=progress.processed / duration if duration > 0 else 0,
                errors=self.stats['errors'][:10],  # First 10 errors
                error_summary=self.stats['error_types']
            )
            
            self._print_final_results(results)
            return results
            
        except Exception as e:
            return self._create_error_result(f"Import failed: {str(e)}")
    
    def _parse_headers(self, worksheet):
        """Parse and validate Excel headers."""
        header_row = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        
        self.column_indexes = {}
        missing_columns = []
        
        for col_idx, header in enumerate(header_row):
            if header and str(header).strip():
                clean_header = str(header).strip()
                if clean_header in self.COLUMN_MAPPING:
                    field_name = self.COLUMN_MAPPING[clean_header]
                    self.column_indexes[field_name] = col_idx
        
        # Check for required columns
        required_fields = ['bill_code', 'air_date']
        for field in required_fields:
            if field not in self.column_indexes:
                original_name = next((k for k, v in self.COLUMN_MAPPING.items() if v == field), field)
                missing_columns.append(original_name)
        
        if missing_columns:
            raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")
    
    def _process_row(self, row: tuple, row_num: int, db_conn: sqlite3.Connection) -> bool:
        """Process a single row with comprehensive error handling."""
        try:
            # Extract all fields using column mapping
            spot_data = {}
            
            for field_name, col_idx in self.column_indexes.items():
                if col_idx < len(row):
                    raw_value = row[col_idx]
                    converted_value = self._convert_value(field_name, raw_value)
                    spot_data[field_name] = converted_value
            
            # Validate required fields
            if not spot_data.get('bill_code') or not spot_data.get('air_date'):
                return False
            
            # Parse bill code for agency/customer
            agency_name, customer_name = self._parse_bill_code(spot_data['bill_code'])
            
            # Handle agency
            if agency_name:
                spot_data['agency_id'] = self._get_or_create_agency(db_conn, agency_name)
            
            # Handle customer
            spot_data['customer_id'] = self._get_or_create_customer(db_conn, customer_name)
            
            # Handle market mapping
            if spot_data.get('market_name'):
                spot_data['market_id'] = self._get_market_id(db_conn, spot_data['market_name'])
            
            # Insert into database
            self._insert_spot(db_conn, spot_data, Path(self.database_path).name)
            
            return True
            
        except Exception as e:
            self._record_error(f"Row {row_num}", str(e))
            return False
    
    def _convert_value(self, field_name: str, raw_value: Any) -> Any:
        """Convert values with proper type handling for SQLite."""
        if raw_value is None or (isinstance(raw_value, str) and not raw_value.strip()):
            return None
        
        try:
            # Date fields
            if field_name in ['air_date', 'end_date']:
                return self._convert_date(raw_value)
            
            # Financial fields (use float for SQLite, allow negatives)
            elif field_name in ['gross_rate', 'station_net', 'spot_value', 'broker_fees']:
                return self._convert_float(raw_value)
            
            # Integer fields
            elif field_name in ['sequence_number', 'line_number', 'priority']:
                return self._convert_integer(raw_value)
            
            # String fields
            else:
                return self._convert_string(raw_value)
                
        except Exception:
            return None
    
    def _convert_date(self, value: Any) -> Optional[date]:
        """Convert date values."""
        if isinstance(value, date):
            return value
        if isinstance(value, datetime):
            return value.date()
        
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None
            
            for fmt in ['%m/%d/%y', '%m/%d/%Y', '%Y-%m-%d']:
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue
        
        return None
    
    def _convert_float(self, value: Any) -> Optional[float]:
        """Convert to float, allowing negative values."""
        if isinstance(value, (int, float)):
            return float(value)
        
        if isinstance(value, str):
            value = value.strip().replace('$', '').replace(',', '')
            if value.startswith('(') and value.endswith(')'):
                value = '-' + value[1:-1]
            
            try:
                return float(value)
            except ValueError:
                return None
        
        return None
    
    def _convert_integer(self, value: Any) -> Optional[int]:
        """Convert to integer."""
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            return int(value)
        if isinstance(value, str) and value.strip():
            try:
                return int(float(value.strip()))
            except ValueError:
                return None
        return None
    
    def _convert_string(self, value: Any) -> Optional[str]:
        """Convert to clean string."""
        if isinstance(value, str):
            cleaned = value.strip()
            return cleaned if cleaned else None
        return str(value) if value is not None else None
    
    def _parse_bill_code(self, bill_code: str) -> Tuple[Optional[str], str]:
        """Parse bill code into agency and customer."""
        if ':' in bill_code:
            parts = bill_code.split(':', 1)
            agency = parts[0].strip()
            customer = parts[1].strip()
            
            # Remove PRODUCTION suffixes
            for suffix in [' PRODUCTION', ' Production', ' PROD']:
                if customer.endswith(suffix):
                    customer = customer[:-len(suffix)].strip()
                    self.stats['customers_normalized'] += 1
                    break
            
            return (agency, customer)
        else:
            customer = bill_code.strip()
            for suffix in [' PRODUCTION', ' Production', ' PROD']:
                if customer.endswith(suffix):
                    customer = customer[:-len(suffix)].strip()
                    self.stats['customers_normalized'] += 1
                    break
            
            return (None, customer)
    
    def _get_or_create_agency(self, conn: sqlite3.Connection, agency_name: str) -> int:
        """Get or create agency."""
        cursor = conn.execute("SELECT agency_id FROM agencies WHERE agency_name = ?", (agency_name,))
        row = cursor.fetchone()
        
        if row:
            return row[0]
        
        cursor = conn.execute("INSERT INTO agencies (agency_name) VALUES (?)", (agency_name,))
        self.stats['agencies_created'] += 1
        return cursor.lastrowid
    
    def _get_or_create_customer(self, conn: sqlite3.Connection, customer_name: str) -> int:
        """Get or create customer."""
        cursor = conn.execute("SELECT customer_id FROM customers WHERE normalized_name = ?", (customer_name,))
        row = cursor.fetchone()
        
        if row:
            return row[0]
        
        cursor = conn.execute("INSERT INTO customers (normalized_name) VALUES (?)", (customer_name,))
        self.stats['customers_created'] += 1
        return cursor.lastrowid
    
    def _get_market_id(self, conn: sqlite3.Connection, market_name: str) -> Optional[int]:
        """Get market ID."""
        cursor = conn.execute("SELECT market_id FROM markets WHERE LOWER(market_name) = LOWER(?)", (market_name,))
        row = cursor.fetchone()
        return row[0] if row else None
    
    def _insert_spot(self, conn: sqlite3.Connection, spot_data: Dict[str, Any], source_file: str):
        """Insert spot with all 29 columns."""
        # Add metadata
        spot_data['source_file'] = source_file
        spot_data['load_date'] = datetime.now()
        
        # Build dynamic insert query for all available fields
        fields = list(spot_data.keys())
        placeholders = ', '.join(['?' for _ in fields])
        field_names = ', '.join(fields)
        values = [spot_data[field] for field in fields]
        
        query = f"INSERT INTO spots ({field_names}) VALUES ({placeholders})"
        conn.execute(query, values)
    
    def _record_error(self, context: str, error: str):
        """Record error with categorization."""
        full_error = f"{context}: {error}"
        self.stats['errors'].append(full_error)
        
        # Categorize error type
        error_type = type(error).__name__ if hasattr(error, '__name__') else 'ProcessingError'
        self.stats['error_types'][error_type] = self.stats['error_types'].get(error_type, 0) + 1
    
    def _create_error_result(self, error_message: str) -> ImportResults:
        """Create error result."""
        return ImportResults(
            success=False,
            total_records=0,
            records_processed=0,
            records_imported=0,
            records_skipped=0,
            new_agencies_created=0,
            new_customers_created=0,
            customers_normalized=0,
            duration_seconds=0,
            records_per_second=0,
            errors=[error_message],
            error_summary={}
        )
    
    def _print_final_results(self, results: ImportResults):
        """Print comprehensive final results."""
        print(f"\n{'='*60}")
        print(f"🎉 PRODUCTION IMPORT COMPLETED")
        print(f"{'='*60}")
        print(f"📊 Performance Metrics:")
        print(f"  Duration: {results.duration_seconds:.2f} seconds")
        print(f"  Processing Rate: {results.records_per_second:,.0f} records/second")
        print(f"")
        print(f"📈 Import Results:")
        print(f"  Total Records: {results.total_records:,}")
        print(f"  Successfully Imported: {results.records_imported:,}")
        print(f"  Skipped: {results.records_skipped:,}")
        print(f"  Success Rate: {(results.records_imported/results.total_records)*100:.1f}%")
        print(f"")
        print(f"🏢 Data Creation:")
        print(f"  New Agencies: {results.new_agencies_created}")
        print(f"  New Customers: {results.new_customers_created}")
        print(f"  Customer Names Normalized: {results.customers_normalized}")
        
        if results.records_skipped > 0:
            print(f"\n⚠️  Error Analysis:")
            print(f"  Total Errors: {results.records_skipped}")
            
            if results.error_summary:
                print(f"  Error Types:")
                for error_type, count in sorted(results.error_summary.items(), key=lambda x: x[1], reverse=True):
                    percentage = (count / results.records_skipped) * 100
                    print(f"    {error_type}: {count} ({percentage:.1f}%)")
            
            if results.errors:
                print(f"\n  Sample Errors (first 5):")
                for error in results.errors[:5]:
                    print(f"    • {error}")


def main():
    """CLI interface for production importer."""
    import argparse
    
    parser = argparse.ArgumentParser(description="Production Excel Importer")
    parser.add_argument("excel_file", help="Path to Excel file")
    parser.add_argument("--database", default="data/database/production.db", help="Database path")
    parser.add_argument("--limit", type=int, help="Limit number of records (for testing)")
    
    args = parser.parse_args()
    
    importer = ProductionExcelImporter(args.database)
    results = importer.import_excel_file(args.excel_file, args.limit)
    
    if results.success:
        print(f"\n✅ Import completed successfully!")
        sys.exit(0)
    else:
        print(f"\n❌ Import failed!")
        for error in results.errors:
            print(f"  {error}")
        sys.exit(1)


if __name__ == "__main__":
    main()