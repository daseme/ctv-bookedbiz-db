#!/usr/bin/env python3
"""
ENHANCED bulk import historical data command with automatic market setup.
Handles large Excel files and automatically creates missing markets and schedule assignments.
"""

import sys
import argparse
import gc
import psutil
import os
from pathlib import Path
from datetime import datetime
from typing import Optional, Generator, Set, Dict, List, Tuple
import sqlite3

# Add src to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from services.broadcast_month_import_service import BroadcastMonthImportService
from services.import_integration_utilities import get_excel_import_summary
from utils.broadcast_month_utils import BroadcastMonthParser
from database.connection import DatabaseConnection


class MarketSetupManager:
    """Manages automatic market creation and schedule assignment setup."""
    
    def __init__(self, db_connection: DatabaseConnection):
        self.db = db_connection
        self.parser = BroadcastMonthParser()
    
    def scan_excel_for_markets(self, excel_file: str) -> Dict[str, Dict]:
        """
        Scan Excel file to extract all market codes and their earliest dates.
        
        Returns:
            Dict mapping market_code -> {'earliest_date': date, 'spot_count': int}
        """
        print(f"🔍 Scanning Excel file for market codes: {excel_file}")
        
        try:
            from openpyxl import load_workbook
            
            workbook = load_workbook(excel_file, read_only=True, data_only=True)
            worksheet = workbook.active
            
            # Find required columns
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
            market_col_index = None
            air_date_col_index = None
            
            for i, header in enumerate(header_row):
                if header:
                    header_str = str(header).strip().lower()
                    if header_str in ['market_name', 'market', 'market_code']:
                        market_col_index = i
                    elif header_str in ['air_date', 'date', 'airdate']:
                        air_date_col_index = i
            
            if market_col_index is None:
                raise ValueError("Market column not found in Excel file")
            if air_date_col_index is None:
                raise ValueError("Air date column not found in Excel file")
            
            print(f"📍 Found Market column at index {market_col_index}")
            print(f"📍 Found Air Date column at index {air_date_col_index}")
            
            markets_data = {}
            row_count = 0
            
            # Process rows to extract market data
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(cell for cell in row):
                    continue
                
                if market_col_index < len(row) and air_date_col_index < len(row):
                    market_value = row[market_col_index]
                    air_date_value = row[air_date_col_index]
                    
                    if market_value and air_date_value:
                        market_code = str(market_value).strip()
                        
                        # Parse air date
                        try:
                            if isinstance(air_date_value, datetime):
                                air_date = air_date_value.date()
                            else:
                                # Handle string dates
                                air_date = datetime.strptime(str(air_date_value), '%Y-%m-%d').date()
                        except:
                            continue
                        
                        # Track market data
                        if market_code not in markets_data:
                            markets_data[market_code] = {
                                'earliest_date': air_date,
                                'latest_date': air_date,
                                'spot_count': 0
                            }
                        else:
                            if air_date < markets_data[market_code]['earliest_date']:
                                markets_data[market_code]['earliest_date'] = air_date
                            if air_date > markets_data[market_code]['latest_date']:
                                markets_data[market_code]['latest_date'] = air_date
                        
                        markets_data[market_code]['spot_count'] += 1
                        row_count += 1
                
                # Progress reporting
                if row_count > 0 and row_count % 10000 == 0:
                    print(f"📊 Processed {row_count:,} rows, found {len(markets_data)} markets")
            
            workbook.close()
            
            print(f"✅ Excel scan complete:")
            print(f"   📊 {row_count:,} total spots analyzed")
            print(f"   🎯 {len(markets_data)} unique markets found")
            
            # Display market summary
            for market_code, data in sorted(markets_data.items()):
                print(f"   📋 {market_code}: {data['spot_count']:,} spots ({data['earliest_date']} to {data['latest_date']})")
            
            return markets_data
            
        except Exception as e:
            raise RuntimeError(f"Failed to scan Excel for markets: {str(e)}")
    
    def get_existing_markets(self) -> Dict[str, int]:
        """Get existing markets from database. Returns market_code -> market_id mapping."""
        with self.db.transaction() as conn:
            cursor = conn.execute("SELECT market_code, market_id FROM markets")
            return {row[0]: row[1] for row in cursor.fetchall()}
    
    def get_markets_with_schedules(self) -> Set[int]:
        """Get market IDs that already have schedule assignments."""
        with self.db.transaction() as conn:
            cursor = conn.execute("SELECT DISTINCT market_id FROM schedule_market_assignments")
            return {row[0] for row in cursor.fetchall()}
    
    def create_missing_markets(self, excel_markets: Dict[str, Dict]) -> Dict[str, int]:
        """
        Create missing markets in database.
        
        Returns:
            Dict mapping market_code -> market_id for all markets (existing + new)
        """
        existing_markets = self.get_existing_markets()
        missing_markets = set(excel_markets.keys()) - set(existing_markets.keys())
        
        if not missing_markets:
            print("✅ All markets already exist in database")
            return existing_markets
        
        print(f"🏗️  Creating {len(missing_markets)} missing markets...")
        
        with self.db.transaction() as conn:
            for market_code in sorted(missing_markets):
                # Create market name from code
                market_name = self._generate_market_name(market_code)
                
                cursor = conn.execute("""
                    INSERT INTO markets (market_code, market_name) 
                    VALUES (?, ?)
                """, (market_code, market_name))
                
                market_id = cursor.lastrowid
                existing_markets[market_code] = market_id
                
                print(f"   ✅ Created market: {market_code} ({market_name}) - ID: {market_id}")
        
        print(f"✅ Market creation complete")
        return existing_markets
    
    def _generate_market_name(self, market_code: str) -> str:
        """Generate a proper market name from market code."""
        # Known mappings
        name_mappings = {
            'NYC': 'NEW YORK',
            'LAX': 'LOS ANGELES', 
            'SFO': 'SAN FRANCISCO',
            'SEA': 'SEATTLE',
            'CHI': 'CHICAGO',
            'MSP': 'MINNEAPOLIS',
            'DAL': 'DALLAS',
            'HOU': 'HOUSTON',
            'WDC': 'WASHINGTON DC',
            'CVC': 'CENTRAL VALLEY',
            'CMP': 'CHI MSP',
            'MMT': 'MAMMOTH',
            'ADMIN': 'ADMINISTRATIVE'
        }
        
        if market_code in name_mappings:
            return name_mappings[market_code]
        else:
            # Generate name from code
            return market_code.upper().replace('_', ' ')
    
    def setup_schedule_assignments(self, excel_markets: Dict[str, Dict], market_mapping: Dict[str, int]) -> int:
        """
        Create schedule assignments for markets that don't have them.
        Uses earliest spot date as effective_start_date.
        
        Returns:
            Number of schedule assignments created
        """
        markets_with_schedules = self.get_markets_with_schedules()
        assignments_created = 0
        
        print(f"🗓️  Setting up schedule assignments...")
        
        with self.db.transaction() as conn:
            for market_code, market_data in excel_markets.items():
                market_id = market_mapping[market_code]
                
                if market_id in markets_with_schedules:
                    print(f"   • {market_code}: Already has schedule assignment")
                    continue
                
                # Create schedule assignment with earliest date
                earliest_date = market_data['earliest_date']
                
                cursor = conn.execute("""
                    INSERT INTO schedule_market_assignments 
                    (schedule_id, market_id, effective_start_date, assignment_priority)
                    VALUES (1, ?, ?, 1)
                """, (market_id, earliest_date))
                
                assignments_created += 1
                print(f"   ✅ {market_code}: Created schedule assignment (effective from {earliest_date})")
        
        print(f"✅ Schedule assignment setup complete: {assignments_created} assignments created")
        return assignments_created
    
    def update_existing_schedule_dates(self, excel_markets: Dict[str, Dict], market_mapping: Dict[str, int]) -> int:
        """
        Update existing schedule assignments if Excel data has earlier dates.
        
        Returns:
            Number of schedule assignments updated
        """
        print(f"📅 Checking existing schedule dates...")
        updates_made = 0
        
        with self.db.transaction() as conn:
            for market_code, market_data in excel_markets.items():
                market_id = market_mapping[market_code]
                earliest_excel_date = market_data['earliest_date']
                
                # Check current schedule start date
                cursor = conn.execute("""
                    SELECT effective_start_date 
                    FROM schedule_market_assignments 
                    WHERE market_id = ?
                """, (market_id,))
                
                result = cursor.fetchone()
                if not result:
                    continue
                
                current_start_date = datetime.strptime(result[0], '%Y-%m-%d').date()
                
                if earliest_excel_date < current_start_date:
                    # Update to earlier date
                    cursor = conn.execute("""
                        UPDATE schedule_market_assignments 
                        SET effective_start_date = ?
                        WHERE market_id = ?
                    """, (earliest_excel_date, market_id))
                    
                    updates_made += 1
                    days_earlier = (current_start_date - earliest_excel_date).days
                    print(f"   ✅ {market_code}: Updated start date from {current_start_date} to {earliest_excel_date} ({days_earlier} days earlier)")
                else:
                    print(f"   • {market_code}: Current date {current_start_date} already covers Excel data")
        
        if updates_made > 0:
            print(f"✅ Schedule date updates complete: {updates_made} dates updated")
        else:
            print(f"✅ All existing schedule dates already cover the data")
        
        return updates_made
    
    def execute_market_setup(self, excel_file: str) -> Dict:
        """
        Execute complete market setup process.
        
        Returns:
            Summary of actions taken
        """
        print(f"🚀 Starting automatic market setup...")
        start_time = datetime.now()
        
        # Step 1: Scan Excel for markets
        excel_markets = self.scan_excel_for_markets(excel_file)
        
        # Step 2: Create missing markets
        market_mapping = self.create_missing_markets(excel_markets)
        
        # Step 3: Setup schedule assignments
        assignments_created = self.setup_schedule_assignments(excel_markets, market_mapping)
        
        # Step 4: Update existing schedule dates if needed
        dates_updated = self.update_existing_schedule_dates(excel_markets, market_mapping)
        
        duration = (datetime.now() - start_time).total_seconds()
        
        summary = {
            'markets_found': len(excel_markets),
            'markets_created': len(excel_markets) - len(self.get_existing_markets()) + len([m for m in excel_markets.keys() if m not in self.get_existing_markets()]),
            'assignments_created': assignments_created,
            'dates_updated': dates_updated,
            'duration_seconds': duration,
            'excel_markets': excel_markets
        }
        
        print(f"✅ Market setup complete in {duration:.2f} seconds!")
        return summary


class EnhancedHistoricalImporter:
    """Enhanced historical importer with automatic market setup."""
    
    def __init__(self, db_connection: DatabaseConnection):
        self.db = db_connection
        self.market_manager = MarketSetupManager(db_connection)
        self.import_service = BroadcastMonthImportService(db_connection)
        self.process = psutil.Process(os.getpid())
    
    def execute_enhanced_import(self, 
                              excel_file: str,
                              expected_year: int,
                              closed_by: str,
                              auto_setup_markets: bool = True,
                              dry_run: bool = False) -> Dict:
        """
        Execute enhanced historical import with optional market setup.
        
        Returns:
            Comprehensive results dictionary
        """
        start_time = datetime.now()
        batch_id = f"enhanced_historical_{int(start_time.timestamp())}"
        
        print(f"🚀 Enhanced Historical Import Starting...")
        print(f"📋 Excel file: {excel_file}")
        print(f"📅 Expected year: {expected_year}")
        print(f"👤 Closed by: {closed_by}")
        print(f"🏗️  Auto-setup markets: {auto_setup_markets}")
        print(f"🔍 Dry run: {dry_run}")
        print(f"📋 Batch ID: {batch_id}")
        print("=" * 70)
        
        results = {
            'success': False,
            'batch_id': batch_id,
            'market_setup': None,
            'import_result': None,
            'duration_seconds': 0,
            'error_messages': []
        }
        
        try:
            # Step 1: Market setup (if enabled)
            if auto_setup_markets and not dry_run:
                print(f"🏗️  STEP 1: Automatic Market Setup")
                market_setup_result = self.market_manager.execute_market_setup(excel_file)
                results['market_setup'] = market_setup_result
                
                print(f"📊 Market Setup Results:")
                print(f"   🎯 Markets found in Excel: {market_setup_result['markets_found']}")
                print(f"   🏗️  New markets created: {market_setup_result.get('markets_created', 0)}")
                print(f"   🗓️  Schedule assignments created: {market_setup_result['assignments_created']}")
                print(f"   📅 Schedule dates updated: {market_setup_result['dates_updated']}")
                print()
            
            elif auto_setup_markets and dry_run:
                print(f"🔍 DRY RUN: Would perform market setup analysis")
                excel_markets = self.market_manager.scan_excel_for_markets(excel_file)
                results['market_setup'] = {'would_analyze': len(excel_markets)}
            
            # Step 2: Historical data import
            print(f"📦 STEP 2: Historical Data Import")
            
            if dry_run:
                print(f"🔍 DRY RUN: Would import historical data")
                # Perform analysis without importing
                summary = get_excel_import_summary(excel_file, self.db.db_path)
                results['import_result'] = {
                    'would_import': True,
                    'months_found': len(summary['months_in_excel']),
                    'total_spots': summary['total_existing_spots_affected']
                }
            else:
                # Execute actual import
                import_result = self.import_service.execute_month_replacement(
                    excel_file,
                    'HISTORICAL',
                    closed_by,
                    dry_run=False
                )
                results['import_result'] = import_result
            
            results['success'] = True
            
        except Exception as e:
            error_msg = f"Enhanced import failed: {str(e)}"
            results['error_messages'].append(error_msg)
            print(f"❌ {error_msg}")
        
        # Calculate total duration
        results['duration_seconds'] = (datetime.now() - start_time).total_seconds()
        
        return results


def display_enhanced_preview(excel_file: str, expected_year: int, db_path: str, auto_setup: bool):
    """Display comprehensive preview of what the enhanced import would do."""
    print(f"📋 Enhanced Historical Import Preview")
    print(f"=" * 70)
    print(f"Excel file: {excel_file}")
    print(f"Expected year: {expected_year}")
    print(f"Auto-setup markets: {auto_setup}")
    print()
    
    try:
        db_connection = DatabaseConnection(db_path)
        
        if auto_setup:
            # Show market setup preview
            market_manager = MarketSetupManager(db_connection)
            excel_markets = market_manager.scan_excel_for_markets(excel_file)
            existing_markets = market_manager.get_existing_markets()
            
            missing_markets = set(excel_markets.keys()) - set(existing_markets.keys())
            
            print(f"🏗️  Market Setup Preview:")
            print(f"   📊 Markets in Excel: {len(excel_markets)}")
            print(f"   ✅ Already exist: {len(existing_markets)}")
            print(f"   🏗️  Will create: {len(missing_markets)}")
            
            if missing_markets:
                print(f"   📋 New markets to create: {sorted(missing_markets)}")
            print()
        
        # Show import preview
        summary = get_excel_import_summary(excel_file, db_path)
        
        print(f"📦 Import Preview:")
        print(f"   📅 Months found: {len(summary['months_in_excel'])}")
        print(f"   📊 Total spots to process: {summary['total_existing_spots_affected']:,}")
        print(f"   🔒 Closed months: {len(summary['closed_months'])}")
        print(f"   📂 Open months: {len(summary['open_months'])}")
        
        db_connection.close()
        
    except Exception as e:
        print(f"❌ Error generating preview: {e}")


def main():
    parser = argparse.ArgumentParser(
        description="ENHANCED bulk import historical data with automatic market setup",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Enhanced import with automatic market setup
  python enhanced_bulk_import_historical.py data/2023_complete.xlsx --year 2023 --closed-by "Kurt" --auto-setup

  # Preview what would happen
  python enhanced_bulk_import_historical.py data/2023_complete.xlsx --year 2023 --closed-by "Kurt" --auto-setup --dry-run

  # Traditional import without market setup
  python enhanced_bulk_import_historical.py data/2023_complete.xlsx --year 2023 --closed-by "Kurt"

Enhanced Features:
  • Automatic market detection from Excel
  • Missing market creation with proper naming  
  • Schedule assignment setup with correct dates
  • Existing schedule date updates when needed
  • Comprehensive preview and validation
        """
    )
    
    parser.add_argument("excel_file", help="Path to Excel file to import")
    parser.add_argument("--year", type=int, required=True, 
                       help="Expected year for validation (e.g., 2023)")
    parser.add_argument("--closed-by", required=True,
                       help="Name/ID of person performing the import")
    parser.add_argument("--auto-setup", action="store_true",
                       help="Automatically create missing markets and schedule assignments")
    parser.add_argument("--db-path", default="data/database/production.db",
                       help="Database path (default: production.db)")
    parser.add_argument("--dry-run", action="store_true",
                       help="Preview import without making changes")
    parser.add_argument("--force", action="store_true",
                       help="Skip confirmation prompts (use with caution)")
    parser.add_argument("--verbose", action="store_true",
                       help="Enable verbose logging")
    
    args = parser.parse_args()
    
    # Setup logging
    import logging
    level = logging.DEBUG if args.verbose else logging.INFO
    logging.basicConfig(level=level, format='%(levelname)s: %(message)s')
    
    # Validate inputs
    if not Path(args.excel_file).exists():
        print(f"❌ Excel file not found: {args.excel_file}")
        sys.exit(1)
    
    if not Path(args.db_path).exists():
        print(f"❌ Database not found: {args.db_path}")
        print(f"Run: uv run python scripts/setup_database.py --db-path {args.db_path}")
        sys.exit(1)
    
    if args.year < 2000 or args.year > 2030:
        print(f"❌ Invalid year: {args.year}. Expected range: 2000-2030")
        sys.exit(1)
    
    try:
        print(f"🚀 Enhanced Historical Import Tool")
        print(f"=" * 50)
        
        # Display enhanced preview
        display_enhanced_preview(args.excel_file, args.year, args.db_path, args.auto_setup)
        
        # Get confirmation unless forced or dry run
        if not args.dry_run and not args.force:
            print(f"\n🚨 CONFIRMATION REQUIRED")
            action_list = ["Import historical data and close months"]
            if args.auto_setup:
                action_list.insert(0, "Create missing markets and schedule assignments")
            
            print(f"This will:")
            for i, action in enumerate(action_list, 1):
                print(f"  {i}. {action}")
            
            response = input(f"\nProceed with enhanced import? (type 'yes' to confirm): ").strip().lower()
            if response != 'yes':
                print(f"❌ Import cancelled by user")
                sys.exit(0)
        
        # Execute enhanced import
        db_connection = DatabaseConnection(args.db_path)
        importer = EnhancedHistoricalImporter(db_connection)
        
        try:
            results = importer.execute_enhanced_import(
                args.excel_file,
                args.year,
                args.closed_by,
                args.auto_setup,
                args.dry_run
            )
            
            # Display results
            print(f"\n" + "=" * 70)
            print(f"🎉 ENHANCED IMPORT {'PREVIEW' if args.dry_run else 'COMPLETED'}")
            print(f"=" * 70)
            
            print(f"📊 Overall Results:")
            print(f"  Success: {'✅' if results['success'] else '❌'}")
            print(f"  Duration: {results['duration_seconds']:.2f} seconds")
            print(f"  Batch ID: {results['batch_id']}")
            
            if results['market_setup']:
                setup = results['market_setup']
                print(f"\n🏗️  Market Setup Results:")
                if not args.dry_run:
                    print(f"  Markets found: {setup['markets_found']}")
                    print(f"  New markets created: {setup.get('markets_created', 0)}")
                    print(f"  Schedule assignments created: {setup['assignments_created']}")
                    print(f"  Schedule dates updated: {setup['dates_updated']}")
                else:
                    print(f"  Would analyze: {setup.get('would_analyze', 0)} markets")
            
            if results['import_result']:
                import_res = results['import_result']
                print(f"\n📦 Import Results:")
                if not args.dry_run and hasattr(import_res, 'success'):
                    print(f"  Records deleted: {import_res.records_deleted:,}")
                    print(f"  Records imported: {import_res.records_imported:,}")
                    print(f"  Months affected: {len(import_res.broadcast_months_affected)}")
                    if hasattr(import_res, 'closed_months') and import_res.closed_months:
                        print(f"  Months closed: {import_res.closed_months}")
                else:
                    print(f"  Would import: {import_res.get('total_spots', 0):,} spots")
                    print(f"  Months found: {import_res.get('months_found', 0)}")
            
            if results['error_messages']:
                print(f"\n❌ Errors:")
                for error in results['error_messages']:
                    print(f"  • {error}")
                sys.exit(1)
            
            if results['success'] and not args.dry_run:
                print(f"\n✅ Enhanced historical import completed successfully!")
                print(f"💡 Next steps:")
                print(f"  • All historical data is now imported and protected")
                print(f"  • Markets and schedules are properly configured")
                print(f"  • Database is ready for current operations")
            
        finally:
            db_connection.close()
    
    except Exception as e:
        print(f"❌ Enhanced import error: {e}")
        if args.verbose:
            import traceback
            traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()