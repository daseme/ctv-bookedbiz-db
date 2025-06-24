#!/usr/bin/env python3
"""
Enhanced weekly update command with automatic market setup.
Replaces open month data while protecting closed historical months.
Automatically creates missing markets and schedule assignments as needed.
"""

import sys
import argparse
from pathlib import Path
from datetime import datetime
from typing import Dict, Set

# Add src to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from services.broadcast_month_import_service import BroadcastMonthImportService, BroadcastMonthImportError
from services.import_integration_utilities import get_excel_import_summary, validate_excel_for_import
from database.connection import DatabaseConnection


class WeeklyMarketSetupManager:
    """Lightweight market setup manager optimized for weekly updates."""
    
    def __init__(self, db_connection: DatabaseConnection):
        self.db = db_connection
    
    def scan_excel_for_new_markets(self, excel_file: str) -> Dict[str, Dict]:
        """
        Quick scan of Excel file to detect any new market codes.
        Optimized for smaller weekly files.
        
        Returns:
            Dict mapping new market_code -> {'earliest_date': date, 'spot_count': int}
        """
        print(f"🔍 Scanning weekly Excel for new markets...")
        
        try:
            from openpyxl import load_workbook
            
            workbook = load_workbook(excel_file, read_only=True, data_only=True)
            worksheet = workbook.active
            
            # Find required columns
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
            market_col_index = None
            air_date_col_index = None
            
            for i, header in enumerate(header_row):
                if header:
                    header_str = str(header).strip().lower()
                    if header_str in ['market_name', 'market', 'market_code']:
                        market_col_index = i
                    elif header_str in ['air_date', 'date', 'airdate']:
                        air_date_col_index = i
            
            if market_col_index is None:
                print("   ℹ️  No market column found - skipping market detection")
                return {}
            
            # Get existing markets to identify new ones
            existing_markets = self.get_existing_markets()
            
            markets_data = {}
            new_markets_found = set()
            
            # Process rows to find new markets
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(cell for cell in row):
                    continue
                
                if market_col_index < len(row):
                    market_value = row[market_col_index]
                    
                    if market_value:
                        market_code = str(market_value).strip()
                        
                        # Only track if it's a NEW market
                        if market_code not in existing_markets:
                            new_markets_found.add(market_code)
                            
                            # Get air date if available
                            air_date = None
                            if air_date_col_index and air_date_col_index < len(row):
                                air_date_value = row[air_date_col_index]
                                if air_date_value:
                                    try:
                                        if isinstance(air_date_value, datetime):
                                            air_date = air_date_value.date()
                                        else:
                                            air_date = datetime.strptime(str(air_date_value), '%Y-%m-%d').date()
                                    except:
                                        pass
                            
                            # Track new market data
                            if market_code not in markets_data:
                                markets_data[market_code] = {
                                    'earliest_date': air_date,
                                    'latest_date': air_date,
                                    'spot_count': 0
                                }
                            else:
                                if air_date:
                                    if not markets_data[market_code]['earliest_date'] or air_date < markets_data[market_code]['earliest_date']:
                                        markets_data[market_code]['earliest_date'] = air_date
                                    if not markets_data[market_code]['latest_date'] or air_date > markets_data[market_code]['latest_date']:
                                        markets_data[market_code]['latest_date'] = air_date
                            
                            markets_data[market_code]['spot_count'] += 1
            
            workbook.close()
            
            if markets_data:
                print(f"   🆕 Found {len(markets_data)} new markets:")
                for market_code, data in sorted(markets_data.items()):
                    date_range = ""
                    if data['earliest_date']:
                        if data['latest_date'] != data['earliest_date']:
                            date_range = f" ({data['earliest_date']} to {data['latest_date']})"
                        else:
                            date_range = f" ({data['earliest_date']})"
                    print(f"      📋 {market_code}: {data['spot_count']:,} spots{date_range}")
            else:
                print(f"   ✅ No new markets found - all markets already exist")
            
            return markets_data
            
        except Exception as e:
            print(f"   ⚠️  Warning: Could not scan for new markets: {str(e)}")
            return {}
    
    def get_existing_markets(self) -> Dict[str, int]:
        """Get existing markets from database."""
        with self.db.transaction() as conn:
            cursor = conn.execute("SELECT market_code, market_id FROM markets")
            return {row[0]: row[1] for row in cursor.fetchall()}
    
    def create_new_markets(self, new_markets: Dict[str, Dict]) -> Dict[str, int]:
        """Create new markets found in weekly data."""
        if not new_markets:
            return {}
        
        print(f"🏗️  Creating {len(new_markets)} new markets for weekly update...")
        
        created_markets = {}
        
        with self.db.transaction() as conn:
            for market_code, market_data in sorted(new_markets.items()):
                # Generate market name
                market_name = self._generate_market_name(market_code)
                
                cursor = conn.execute("""
                    INSERT INTO markets (market_code, market_name) 
                    VALUES (?, ?)
                """, (market_code, market_name))
                
                market_id = cursor.lastrowid
                created_markets[market_code] = market_id
                
                print(f"   ✅ Created: {market_code} ({market_name}) - ID: {market_id}")
        
        return created_markets
    
    def setup_schedules_for_new_markets(self, new_markets: Dict[str, Dict], market_mapping: Dict[str, int]) -> int:
        """Setup schedule assignments for newly created markets."""
        if not new_markets:
            return 0
        
        print(f"🗓️  Setting up schedules for new markets...")
        assignments_created = 0
        
        with self.db.transaction() as conn:
            for market_code, market_data in new_markets.items():
                market_id = market_mapping[market_code]
                
                # Use earliest date if available, otherwise use current date
                effective_date = market_data['earliest_date'] if market_data['earliest_date'] else datetime.now().date()
                
                cursor = conn.execute("""
                    INSERT INTO schedule_market_assignments 
                    (schedule_id, market_id, effective_start_date, assignment_priority)
                    VALUES (1, ?, ?, 1)
                """, (market_id, effective_date))
                
                assignments_created += 1
                print(f"   ✅ {market_code}: Schedule assignment created (effective from {effective_date})")
        
        return assignments_created
    
    def _generate_market_name(self, market_code: str) -> str:
        """Generate proper market name from code."""
        name_mappings = {
            'NYC': 'NEW YORK',
            'LAX': 'LOS ANGELES', 
            'SFO': 'SAN FRANCISCO',
            'SEA': 'SEATTLE',
            'CHI': 'CHICAGO',
            'MSP': 'MINNEAPOLIS',
            'DAL': 'DALLAS',
            'HOU': 'HOUSTON',
            'WDC': 'WASHINGTON DC',
            'CVC': 'CENTRAL VALLEY',
            'CMP': 'CHI MSP',
            'MMT': 'MAMMOTH',
            'ADMIN': 'ADMINISTRATIVE'
        }
        
        return name_mappings.get(market_code, market_code.upper().replace('_', ' '))
    
    def execute_weekly_market_setup(self, excel_file: str) -> Dict:
        """Execute lightweight market setup for weekly data."""
        start_time = datetime.now()
        
        # Step 1: Scan for new markets
        new_markets = self.scan_excel_for_new_markets(excel_file)
        
        if not new_markets:
            return {
                'new_markets_found': 0,
                'markets_created': 0,
                'schedules_created': 0,
                'duration_seconds': (datetime.now() - start_time).total_seconds()
            }
        
        # Step 2: Create new markets
        created_markets = self.create_new_markets(new_markets)
        
        # Step 3: Setup schedules for new markets
        schedules_created = self.setup_schedules_for_new_markets(new_markets, created_markets)
        
        duration = (datetime.now() - start_time).total_seconds()
        
        return {
            'new_markets_found': len(new_markets),
            'markets_created': len(created_markets),
            'schedules_created': schedules_created,
            'duration_seconds': duration,
            'new_markets': new_markets
        }


def display_enhanced_weekly_preview(excel_file: str, db_path: str, auto_setup: bool):
    """Display what the enhanced weekly update would do."""
    print(f"📋 Enhanced Weekly Update Preview")
    print(f"=" * 60)
    print(f"Excel file: {excel_file}")
    print(f"Auto-setup markets: {auto_setup}")
    print()
    
    try:
        db_connection = DatabaseConnection(db_path)
        
        # Market setup preview
        if auto_setup:
            market_manager = WeeklyMarketSetupManager(db_connection)
            new_markets = market_manager.scan_excel_for_new_markets(excel_file)
            
            if new_markets:
                print(f"🏗️  Market Setup Preview:")
                print(f"   🆕 New markets detected: {len(new_markets)}")
                print(f"   📋 Will create: {sorted(new_markets.keys())}")
                print()
        
        # Standard import preview
        can_proceed = display_weekly_update_preview(excel_file, db_path)
        
        db_connection.close()
        return can_proceed
        
    except Exception as e:
        print(f"❌ Error generating enhanced preview: {e}")
        return False


def display_weekly_update_preview(excel_file: str, db_path: str):
    """Display what the weekly update would do (original functionality)."""
    print(f"📦 Weekly Update Preview:")
    
    try:
        # Get Excel summary
        summary = get_excel_import_summary(excel_file, db_path)
        
        print(f"   📅 Months found: {len(summary['months_in_excel'])}")
        print(f"   📊 Total existing spots: {summary['total_existing_spots_affected']:,}")
        
        # Validation check
        validation = validate_excel_for_import(excel_file, 'WEEKLY_UPDATE', db_path)
        
        if validation.is_valid:
            print(f"   ✅ Weekly update allowed")
        else:
            print(f"   ❌ Weekly update BLOCKED")
            print(f"      Reason: {validation.error_message}")
            print(f"      Solution: {validation.suggested_action}")
            return False
        
        # Show month details
        print(f"\n📅 Month Details:")
        for month_info in summary['month_details']:
            status = "CLOSED" if month_info['is_closed'] else "OPEN"
            spots = month_info['existing_spots']
            revenue = month_info['existing_revenue']
            status_icon = "🔒" if month_info['is_closed'] else "📂"
            print(f"   {status_icon} {month_info['month']}: {status} - {spots:,} spots (${revenue:,.2f})")
        
        # Import impact
        open_months = summary['open_months']
        print(f"\n🎯 Import Impact:")
        print(f"   • {len(open_months)} open months will be updated")
        print(f"   • {summary['total_existing_spots_affected']:,} existing spots will be replaced")
        print(f"   • Closed months will be protected (no changes)")
        
        if summary['closed_months']:
            print(f"   • {len(summary['closed_months'])} closed months: {summary['closed_months']}")
        
        return True
        
    except Exception as e:
        print(f"❌ Error generating preview: {e}")
        return False


def get_user_confirmation(total_spots: int, open_months: int, new_markets: int = 0, force: bool = False) -> bool:
    """Get user confirmation for the enhanced update."""
    if force:
        return True
    
    print(f"\n🚨 CONFIRMATION REQUIRED")
    actions = [
        f"REPLACE {total_spots:,} existing spots",
        f"Update {open_months} open months",
        "Preserve all closed/historical months"
    ]
    
    if new_markets > 0:
        actions.insert(0, f"Create {new_markets} new markets")
    
    print(f"This will:")
    for action in actions:
        print(f"  • {action}")
    
    while True:
        response = input(f"\nProceed with enhanced weekly update? (yes/no): ").strip().lower()
        if response in ['yes', 'y']:
            return True
        elif response in ['no', 'n', '']:
            return False
        else:
            print("Please enter 'yes' or 'no'")


def main():
    parser = argparse.ArgumentParser(
        description="Enhanced weekly update with automatic market setup",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Standard weekly update (original functionality)
  python enhanced_weekly_update.py data/booked_business_current.xlsx

  # Enhanced weekly update with automatic market setup
  python enhanced_weekly_update.py data/booked_business_current.xlsx --auto-setup

  # Preview what would happen
  python enhanced_weekly_update.py data/weekly_data.xlsx --auto-setup --dry-run

Enhanced Features:
  • Automatic detection of new markets in weekly data
  • Missing market creation with proper naming
  • Schedule assignment setup for new markets
  • Backward compatible with existing workflows
        """
    )
    
    parser.add_argument("excel_file", help="Path to Excel file to import")
    parser.add_argument("--auto-setup", action="store_true",
                       help="Automatically create missing markets and schedule assignments")
    parser.add_argument("--db-path", default="data/database/production.db",
                       help="Database path (default: production.db)")
    parser.add_argument("--dry-run", action="store_true",
                       help="Preview import without making changes")
    parser.add_argument("--force", action="store_true",
                       help="Skip confirmation prompts")
    parser.add_argument("--verbose", action="store_true",
                       help="Enable verbose logging")
    
    args = parser.parse_args()
    
    # Setup logging
    import logging
    level = logging.DEBUG if args.verbose else logging.INFO
    logging.basicConfig(level=level, format='%(levelname)s: %(message)s')
    
    # Validate inputs
    if not Path(args.excel_file).exists():
        print(f"❌ Excel file not found: {args.excel_file}")
        sys.exit(1)
    
    if not Path(args.db_path).exists():
        print(f"❌ Database not found: {args.db_path}")
        print(f"Run: uv run python scripts/setup_database.py --db-path {args.db_path}")
        sys.exit(1)
    
    try:
        print(f"🔄 Enhanced Weekly Update Tool")
        print(f"Excel file: {args.excel_file}")
        print(f"Database: {args.db_path}")
        print(f"Auto-setup: {args.auto_setup}")
        if args.dry_run:
            print(f"Mode: DRY RUN (no changes will be made)")
        print()
        
        # Market setup results
        market_setup_results = None
        
        # Execute market setup if enabled
        if args.auto_setup and not args.dry_run:
            db_connection = DatabaseConnection(args.db_path)
            market_manager = WeeklyMarketSetupManager(db_connection)
            market_setup_results = market_manager.execute_weekly_market_setup(args.excel_file)
            db_connection.close()
            
            if market_setup_results['markets_created'] > 0:
                print(f"✅ Market setup completed:")
                print(f"   🆕 New markets created: {market_setup_results['markets_created']}")
                print(f"   🗓️  Schedule assignments created: {market_setup_results['schedules_created']}")
                print()
        
        # Display enhanced preview and validate
        can_proceed = display_enhanced_weekly_preview(args.excel_file, args.db_path, args.auto_setup)
        
        if not can_proceed:
            print(f"\n❌ Weekly update cannot proceed due to validation errors")
            print(f"💡 Common solutions:")
            print(f"  • Remove closed month data from Excel file")
            print(f"  • Use historical import mode for closed months")
            print(f"  • Check if months need to be manually closed first")
            sys.exit(1)
        
        # Get confirmation unless forced or dry run
        if not args.dry_run and not args.force:
            summary = get_excel_import_summary(args.excel_file, args.db_path)
            new_market_count = 0
            
            if args.auto_setup and market_setup_results:
                new_market_count = market_setup_results['markets_created']
            elif args.auto_setup:
                # Dry run - estimate new markets
                db_connection = DatabaseConnection(args.db_path)
                market_manager = WeeklyMarketSetupManager(db_connection)
                new_markets = market_manager.scan_excel_for_new_markets(args.excel_file)
                new_market_count = len(new_markets)
                db_connection.close()
            
            confirmed = get_user_confirmation(
                summary['total_existing_spots_affected'],
                len(summary['open_months']),
                new_market_count,
                args.force
            )
            
            if not confirmed:
                print(f"❌ Weekly update cancelled by user")
                sys.exit(0)
        
        # Execute the standard update
        db_connection = DatabaseConnection(args.db_path)
        service = BroadcastMonthImportService(db_connection)
        
        try:
            result = service.execute_month_replacement(
                args.excel_file,
                'WEEKLY_UPDATE',
                closed_by=None,  # Not needed for weekly updates
                dry_run=args.dry_run
            )
            
            # Display results
            print(f"\n{'='*60}")
            if args.dry_run:
                print(f"🔍 ENHANCED WEEKLY UPDATE DRY RUN COMPLETED")
            else:
                print(f"🎉 ENHANCED WEEKLY UPDATE COMPLETED")
            print(f"{'='*60}")
            
            # Market setup results
            if market_setup_results and not args.dry_run:
                print(f"🏗️  Market Setup Results:")
                print(f"   New markets found: {market_setup_results['new_markets_found']}")
                print(f"   Markets created: {market_setup_results['markets_created']}")
                print(f"   Schedule assignments created: {market_setup_results['schedules_created']}")
                print(f"   Setup duration: {market_setup_results['duration_seconds']:.2f} seconds")
                print()
            
            # Import results
            print(f"📦 Import Results:")
            print(f"   Success: {'✅' if result.success else '❌'}")
            print(f"   Batch ID: {result.batch_id}")
            print(f"   Duration: {result.duration_seconds:.2f} seconds")
            print(f"   Months updated: {len(result.broadcast_months_affected)}")
            
            if not args.dry_run:
                print(f"   Records deleted: {result.records_deleted:,}")
                print(f"   Records imported: {result.records_imported:,}")
                print(f"   Net change: {result.records_imported - result.records_deleted:+,}")
            
            if result.error_messages:
                print(f"\n❌ Errors:")
                for error in result.error_messages:
                    print(f"  • {error}")
                sys.exit(1)
            
            if not args.dry_run and result.success:
                print(f"\n✅ Enhanced weekly update completed successfully!")
                print(f"📋 Summary:")
                print(f"   • Updated {len(result.broadcast_months_affected)} open months")
                if market_setup_results and market_setup_results['markets_created'] > 0:
                    print(f"   • Created {market_setup_results['markets_created']} new markets")
                print(f"   • Historical/closed months preserved")
                print(f"   • Database ready for reporting")
                
                print(f"\n💡 Next steps:")
                print(f"   • Generate reports with updated data")
                print(f"   • Close current month when ready: uv run python src/cli/close_month.py")
            
        finally:
            db_connection.close()
    
    except BroadcastMonthImportError as e:
        print(f"❌ Import error: {e}")
        sys.exit(1)
    except KeyboardInterrupt:
        print(f"\n❌ Weekly update cancelled by user")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Unexpected error: {e}")
        if args.verbose:
            import traceback
            traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()