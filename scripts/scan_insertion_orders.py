#!/usr/bin/env python3
"""Scan K drive insertion order spreadsheets and write a summary JSON.

Standalone CLI script with no web app dependency.
Designed to run on a schedule via systemd timer.

Usage:
    python3 scripts/scan_insertion_orders.py
    python3 scripts/scan_insertion_orders.py --io-path /custom/path
    python3 scripts/scan_insertion_orders.py --output data/pending_orders.json
"""

import argparse
import json
import logging
import os
import sys
import tempfile
from datetime import datetime

import openpyxl

DEFAULT_IO_PATH = "/mnt/k-drive/Insertion Orders"
DEFAULT_OUTPUT = "data/pending_orders.json"

# Run Sheet column positions (0-indexed)
COL_BILL_CODE = 0
COL_AIR_DATE = 1
COL_END_DATE = 2
COL_GROSS_RATE = 15
COL_MONTH = 18
COL_STATION_NET = 21
COL_SALES_PERSON = 22
COL_REVENUE_TYPE = 23
COL_CONTRACT = 27
COL_MARKET = 28

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger(__name__)


def parse_date(value):
    """Convert a cell value to an ISO date string, or None."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def parse_float(value):
    """Convert a cell value to float, defaulting to 0.0."""
    if value is None:
        return 0.0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0


def cell_str(value):
    """Convert a cell value to a stripped string."""
    if value is None:
        return ""
    return str(value).strip()


def scan_workbook(filepath, folder_name):
    """Read a single workbook's Run Sheet and return an order dict.

    Returns None if there is no Run Sheet or no data rows.
    """
    wb = openpyxl.load_workbook(
        filepath, read_only=True, data_only=True
    )
    try:
        if "Run Sheet" not in wb.sheetnames:
            return None

        ws = wb["Run Sheet"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        if not rows:
            return None

        filename = os.path.basename(filepath)
        spot_count = 0
        total_gross = 0.0
        total_net = 0.0
        date_starts = []
        date_ends = []
        bill_code = ""
        sales_person = ""
        contract = ""
        market = ""
        revenue_type = ""

        for row in rows:
            if not row or len(row) <= COL_MARKET:
                continue

            # Skip completely empty rows
            if all(cell is None for cell in row):
                continue

            spot_count += 1
            total_gross += parse_float(row[COL_GROSS_RATE])
            total_net += parse_float(row[COL_STATION_NET])

            air_date = parse_date(row[COL_AIR_DATE])
            if air_date:
                date_starts.append(air_date)

            end_date = parse_date(row[COL_END_DATE])
            if end_date:
                date_ends.append(end_date)

            if not bill_code:
                bill_code = cell_str(row[COL_BILL_CODE])
            if not sales_person:
                sales_person = cell_str(row[COL_SALES_PERSON])
            if not contract:
                contract = cell_str(row[COL_CONTRACT])
            if not market:
                market = cell_str(row[COL_MARKET])
            if not revenue_type:
                revenue_type = cell_str(row[COL_REVENUE_TYPE])

        if spot_count == 0:
            return None

        # Derive customer from bill_code ("Agency:Customer" pattern)
        customer = ""
        if ":" in bill_code:
            customer = bill_code.split(":", 1)[1].strip()
        else:
            customer = bill_code

        date_range_start = min(date_starts) if date_starts else None
        date_range_end = (
            max(date_ends) if date_ends
            else (max(date_starts) if date_starts else None)
        )

        return {
            "folder": folder_name,
            "file": filename,
            "customer": customer,
            "bill_code": bill_code,
            "sales_person": sales_person,
            "contract": contract,
            "market": market,
            "revenue_type": revenue_type,
            "spot_count": spot_count,
            "total_gross": round(total_gross, 2),
            "total_net": round(total_net, 2),
            "date_range_start": date_range_start,
            "date_range_end": date_range_end,
        }
    finally:
        wb.close()


def parse_folder_name(folder_name):
    """Extract contract number and customer from folder name.

    Folder format: "NNNN Customer Name" where leading digits are the
    contract number and the rest is the customer name.
    """
    parts = folder_name.split(None, 1)
    if len(parts) == 2 and parts[0].isdigit():
        return parts[0], parts[1]
    return "", folder_name


def order_from_folder(folder_name):
    """Create a stub order entry from just the folder name.

    Used when a folder has no .xlsx Run Sheet files.
    """
    contract, customer = parse_folder_name(folder_name)
    return {
        "folder": folder_name,
        "file": None,
        "customer": customer,
        "bill_code": "",
        "sales_person": "",
        "contract": contract,
        "market": "",
        "revenue_type": "",
        "spot_count": 0,
        "total_gross": 0.0,
        "total_net": 0.0,
        "date_range_start": None,
        "date_range_end": None,
    }


def scan_all(io_path):
    """Walk the insertion orders directory and return orders + errors."""
    orders = []
    errors = []

    if not os.path.isdir(io_path):
        logger.error(
            "Insertion orders path not found: %s "
            "(is the K drive mounted?)",
            io_path,
        )
        sys.exit(1)

    for entry in sorted(os.listdir(io_path)):
        folder_path = os.path.join(io_path, entry)
        if not os.path.isdir(folder_path):
            continue

        folder_has_order = False
        for filename in sorted(os.listdir(folder_path)):
            if not filename.endswith(".xlsx"):
                continue
            if filename.startswith("~$"):
                continue

            filepath = os.path.join(folder_path, filename)
            try:
                order = scan_workbook(filepath, entry)
                if order:
                    orders.append(order)
                    folder_has_order = True
            except Exception as exc:
                msg = f"{entry}/{filename}: {exc}"
                logger.warning("Error scanning %s", msg)
                errors.append(msg)

        if not folder_has_order:
            orders.append(order_from_folder(entry))

    return orders, errors


def write_json(output_path, orders, errors):
    """Write the result JSON atomically (write tmp then rename)."""
    result = {
        "scanned_at": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
        "order_count": len(orders),
        "error_count": len(errors),
        "orders": orders,
        "errors": errors,
    }

    output_dir = os.path.dirname(output_path) or "."
    os.makedirs(output_dir, exist_ok=True)

    fd, tmp_path = tempfile.mkstemp(
        dir=output_dir, suffix=".tmp", prefix=".pending_orders_"
    )
    try:
        with os.fdopen(fd, "w") as f:
            json.dump(result, f, indent=2)
        os.replace(tmp_path, output_path)
    except Exception:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
        raise

    logger.info(
        "Wrote %s: %d orders, %d errors",
        output_path,
        len(orders),
        len(errors),
    )


def main():
    parser = argparse.ArgumentParser(
        description="Scan insertion order spreadsheets from K drive"
    )
    parser.add_argument(
        "--io-path",
        default=DEFAULT_IO_PATH,
        help=f"Path to Insertion Orders directory (default: {DEFAULT_IO_PATH})",
    )
    parser.add_argument(
        "--output",
        default=DEFAULT_OUTPUT,
        help=f"Output JSON path (default: {DEFAULT_OUTPUT})",
    )
    args = parser.parse_args()

    logger.info("Scanning insertion orders from: %s", args.io_path)
    orders, errors = scan_all(args.io_path)
    write_json(args.output, orders, errors)

    if errors:
        logger.warning("%d files had errors", len(errors))
        sys.exit(0)

    logger.info("Scan complete: %d orders found", len(orders))


if __name__ == "__main__":
    main()
