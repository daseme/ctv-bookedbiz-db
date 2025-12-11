#!/usr/bin/env python3
"""
Database-Driven AE Budget Sheet Generator - Clean Architecture Implementation
Pulls data directly from CTV BookedBiz SQLite database instead of Excel files
Compliant with CTV BookedBiz DB Coding Guidelines
"""

from dataclasses import dataclass, replace
from typing import Dict, List, Optional, Tuple, Generator
from pathlib import Path
from datetime import datetime, date
from decimal import Decimal
import pandas as pd
import sqlite3
import logging
import sys
import subprocess

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)

# Domain Entities (Immutable)
@dataclass(frozen=True)
class Customer:
    """Immutable customer entity"""
    customer_id: int
    name: str
    sector: str
    ae_name: str
    revenue_2025: Decimal
    spot_count: int

@dataclass(frozen=True)
class AEPerformance:
    """Immutable AE performance metrics"""
    ae_name: str
    total_revenue: Decimal
    customer_count: int
    spot_count: int
    customers: List[Customer]
    
    def top_customers(self, limit: int = 10) -> List[Customer]:
        """Get top customers by revenue (pure function)"""
        return sorted(self.customers, key=lambda c: c.revenue_2025, reverse=True)[:limit]

@dataclass(frozen=True)
class SectorSummary:
    """Immutable sector performance summary"""
    sector_name: str
    revenue_by_year: Dict[int, Decimal]
    total_revenue: Decimal
    customer_count: int

@dataclass(frozen=True)
class MonthlyRevenue:
    """Monthly revenue breakdown"""
    broadcast_month: str
    revenue: Decimal
    spot_count: int

@dataclass(frozen=True)
class LanguageRevenue:
    """Language revenue breakdown"""
    language_code: str
    language_name: str
    revenue: Decimal
    spot_count: int
    customer_count: int

# Repository Pattern (Data Access Layer)
class RevenueRepository:
    """Repository for revenue data access with parameterized queries only"""
    
    def __init__(self, db_path: str):
        self._db_path = db_path
    
    def get_ae_revenue_2025(self, ae_name: str) -> List[Customer]:
        """Get customer revenue data for AE in 2025 using gross rates"""
        with sqlite3.connect(self._db_path) as conn:
            cursor = conn.execute("""
                SELECT 
                    COALESCE(s.customer_id, -1) as customer_id,
                    COALESCE(c.normalized_name, s.bill_code) as customer_name,
                    COALESCE(sec.sector_name, 'Unknown') as sector,
                    s.sales_person as ae_name,
                    SUM(COALESCE(s.gross_rate, 0)) as total_revenue,
                    COUNT(*) as spot_count
                FROM spots s
                LEFT JOIN customers c ON s.customer_id = c.customer_id
                LEFT JOIN sectors sec ON c.sector_id = sec.sector_id
                WHERE UPPER(TRIM(s.sales_person)) = UPPER(TRIM(?))
                    AND s.broadcast_month LIKE '%-25'
                    AND (s.revenue_type != 'Trade' OR s.revenue_type IS NULL)
                    AND COALESCE(s.gross_rate, 0) > 0
                GROUP BY 
                    s.customer_id, 
                    COALESCE(c.normalized_name, s.bill_code),
                    COALESCE(sec.sector_name, 'Unknown'),
                    s.sales_person
                ORDER BY total_revenue DESC
            """, (ae_name,))
            
            customers = []
            for row in cursor.fetchall():
                customers.append(Customer(
                    customer_id=row[0],
                    name=row[1],
                    sector=row[2],
                    ae_name=row[3],
                    revenue_2025=Decimal(str(row[4])),
                    spot_count=row[5]
                ))
            
            return customers
    
    def get_ae_new_vs_returning_customers(self, ae_name: str) -> Tuple[List[Customer], List[Customer]]:
        """Get new vs returning customers for 2025 (comparing to 2024)"""
        with sqlite3.connect(self._db_path) as conn:
            # First, let's get all 2025 data aggregated by TRUE normalized name
            cursor_2025 = conn.execute("""
                WITH normalized_customers_2025 AS (
                    SELECT 
                        COALESCE(c.normalized_name, s.bill_code) as normalized_name,
                        COALESCE(sec.sector_name, 'Unknown') as sector,
                        SUM(COALESCE(s.gross_rate, 0)) as total_revenue,
                        COUNT(*) as spot_count,
                        -- Get the first customer_id for this normalized name
                        MIN(COALESCE(s.customer_id, -1)) as customer_id,
                        -- Get the most common display name for this normalized name
                        MAX(COALESCE(c.normalized_name, s.bill_code)) as display_name
                    FROM spots s
                    LEFT JOIN customers c ON s.customer_id = c.customer_id
                    LEFT JOIN sectors sec ON c.sector_id = sec.sector_id
                    WHERE UPPER(TRIM(s.sales_person)) = UPPER(TRIM(?))
                        AND s.broadcast_month LIKE '%-25'
                        AND (s.revenue_type != 'Trade' OR s.revenue_type IS NULL)
                        AND COALESCE(s.gross_rate, 0) > 0
                    GROUP BY 
                        COALESCE(c.normalized_name, s.bill_code),
                        COALESCE(sec.sector_name, 'Unknown')
                )
                SELECT 
                    customer_id,
                    display_name,
                    normalized_name,
                    sector,
                    total_revenue,
                    spot_count
                FROM normalized_customers_2025
                ORDER BY total_revenue DESC
            """, (ae_name,))
            
            customers_2025 = {}
            for row in cursor_2025.fetchall():
                normalized_key = row[2].strip().upper() if row[2] else 'UNKNOWN'
                customers_2025[normalized_key] = Customer(
                    customer_id=row[0],
                    name=row[1],  # Use display name for reports
                    sector=row[3],
                    ae_name=ae_name,
                    revenue_2025=Decimal(str(row[4])),
                    spot_count=row[5]
                )
            
            # Get 2024 customers aggregated by normalized name
            cursor_2024 = conn.execute("""
                SELECT DISTINCT
                    COALESCE(c.normalized_name, s.bill_code) as normalized_name
                FROM spots s
                LEFT JOIN customers c ON s.customer_id = c.customer_id
                WHERE UPPER(TRIM(s.sales_person)) = UPPER(TRIM(?))
                    AND s.broadcast_month LIKE '%-24'
                    AND (s.revenue_type != 'Trade' OR s.revenue_type IS NULL)
                    AND COALESCE(s.gross_rate, 0) > 0
            """, (ae_name,))
            
            customers_2024 = set()
            for row in cursor_2024.fetchall():
                normalized_key = row[0].strip().upper() if row[0] else 'UNKNOWN'
                customers_2024.add(normalized_key)
            
            logger.info(f"   📊 Found {len(customers_2024)} unique normalized customers in 2024")
            logger.info(f"   📊 Found {len(customers_2025)} unique normalized customers in 2025")
            
            # Debug: Show sample customer names for verification
            if logger.level <= logging.DEBUG:
                sample_2024 = list(customers_2024)[:3] if customers_2024 else []
                sample_2025 = [(k, v.name) for k, v in list(customers_2025.items())[:3]]
                logger.debug(f"Sample 2024 normalized: {sample_2024}")
                logger.debug(f"Sample 2025 (norm_key, display): {sample_2025}")
            
            # Classify customers as new or returning based on normalized names
            new_customers = []
            returning_customers = []
            
            for normalized_key, customer in customers_2025.items():
                if normalized_key in customers_2024:
                    returning_customers.append(customer)
                    logger.debug(f"RETURNING: '{customer.name}' -> normalized: '{normalized_key}'")
                else:
                    new_customers.append(customer)
                    logger.debug(f"NEW: '{customer.name}' -> normalized: '{normalized_key}'")
            
            # Sort by revenue descending
            new_customers.sort(key=lambda c: c.revenue_2025, reverse=True)
            returning_customers.sort(key=lambda c: c.revenue_2025, reverse=True)
            
            logger.info(f"   👥 Customer analysis: {len(new_customers)} new, {len(returning_customers)} returning")
            
            # Additional debug: Show which customers are being classified where
            if new_customers:
                logger.info(f"   🆕 Top new customers: {[c.name for c in new_customers[:3]]}")
            if returning_customers:
                logger.info(f"   🔄 Top returning customers: {[c.name for c in returning_customers[:3]]}")
            
            return new_customers, returning_customers
    
    def get_ae_sector_trends(self, ae_name: str, years: List[int]) -> List[SectorSummary]:
        """Get sector performance trends for multiple years"""
        year_conditions = " OR ".join([f"s.broadcast_month LIKE '%-{str(year)[-2:]}'" for year in years])
        
        with sqlite3.connect(self._db_path) as conn:
            # Debug: Check what broadcast months exist for this AE
            debug_cursor = conn.execute("""
                SELECT DISTINCT s.broadcast_month, COUNT(*) as spots
                FROM spots s
                WHERE UPPER(TRIM(s.sales_person)) = UPPER(TRIM(?))
                    AND (s.revenue_type != 'Trade' OR s.revenue_type IS NULL)
                    AND COALESCE(s.gross_rate, 0) > 0
                GROUP BY s.broadcast_month
                ORDER BY s.broadcast_month
            """, (ae_name,))
            
            available_months = debug_cursor.fetchall()
            logger.debug(f"Available months for {ae_name}: {available_months}")
            
            cursor = conn.execute(f"""
                SELECT 
                    COALESCE(sec.sector_name, 'Unknown') as sector,
                    s.broadcast_month,
                    SUM(COALESCE(s.gross_rate, 0)) as revenue,
                    COUNT(DISTINCT COALESCE(s.customer_id, s.bill_code)) as customers
                FROM spots s
                LEFT JOIN customers c ON s.customer_id = c.customer_id
                LEFT JOIN sectors sec ON c.sector_id = sec.sector_id
                WHERE UPPER(TRIM(s.sales_person)) = UPPER(TRIM(?))
                    AND ({year_conditions})
                    AND (s.revenue_type != 'Trade' OR s.revenue_type IS NULL)
                    AND COALESCE(s.gross_rate, 0) > 0
                GROUP BY 
                    COALESCE(sec.sector_name, 'Unknown'),
                    s.broadcast_month
                ORDER BY sector, s.broadcast_month
            """, (ae_name,))
            
            # Group by sector and aggregate customer counts properly
            sector_data = {}
            years_found = set()
            
            for row in cursor.fetchall():
                sector = row[0]
                month = row[1]
                revenue = Decimal(str(row[2]))
                
                if sector not in sector_data:
                    sector_data[sector] = {
                        'revenue_by_year': {},
                        'total_revenue': Decimal('0'),
                        'customer_ids': set()
                    }
                
                # Extract year from broadcast_month (format: "Jan-25")
                if month and '-' in month:
                    year_suffix = month.split('-')[1]
                    # Handle both 2-digit and 4-digit years properly
                    if len(year_suffix) == 2:
                        year_int = int(year_suffix)
                        # Assume 00-49 = 2000-2049, 50-99 = 1950-1999
                        year = 2000 + year_int if year_int <= 49 else 1900 + year_int
                    else:
                        year = int(year_suffix)  # 4-digit year
                    
                    years_found.add(year)
                    
                    if year not in sector_data[sector]['revenue_by_year']:
                        sector_data[sector]['revenue_by_year'][year] = Decimal('0')
                    
                    sector_data[sector]['revenue_by_year'][year] += revenue
                    sector_data[sector]['total_revenue'] += revenue
            
            logger.info(f"   📅 Years found in data: {sorted(years_found)} (requested: {years})")
            
            # Now get the actual distinct customer count per sector for this AE
            cursor = conn.execute(f"""
                SELECT 
                    COALESCE(sec.sector_name, 'Unknown') as sector,
                    COUNT(DISTINCT COALESCE(s.customer_id, s.bill_code)) as unique_customers
                FROM spots s
                LEFT JOIN customers c ON s.customer_id = c.customer_id
                LEFT JOIN sectors sec ON c.sector_id = sec.sector_id
                WHERE UPPER(TRIM(s.sales_person)) = UPPER(TRIM(?))
                    AND ({year_conditions})
                    AND (s.revenue_type != 'Trade' OR s.revenue_type IS NULL)
                    AND COALESCE(s.gross_rate, 0) > 0
                GROUP BY COALESCE(sec.sector_name, 'Unknown')
            """, (ae_name,))
            
            # Update customer counts with corrected values
            for row in cursor.fetchall():
                sector = row[0]
                customer_count = row[1]
                if sector in sector_data:
                    sector_data[sector]['customer_count'] = customer_count
            
            # Convert to SectorSummary objects
            summaries = []
            for sector, data in sector_data.items():
                summaries.append(SectorSummary(
                    sector_name=sector,
                    revenue_by_year=data['revenue_by_year'],
                    total_revenue=data['total_revenue'],
                    customer_count=data.get('customer_count', 0)
                ))
            
            return sorted(summaries, key=lambda s: s.total_revenue, reverse=True)
    
    def get_ae_monthly_breakdown_2025(self, ae_name: str) -> List[MonthlyRevenue]:
        """Get monthly revenue breakdown for 2025 using gross rates"""
        with sqlite3.connect(self._db_path) as conn:
            cursor = conn.execute("""
                SELECT 
                    s.broadcast_month,
                    SUM(COALESCE(s.gross_rate, 0)) as revenue,
                    COUNT(*) as spot_count
                FROM spots s
                WHERE UPPER(TRIM(s.sales_person)) = UPPER(TRIM(?))
                    AND s.broadcast_month LIKE '%-25'
                    AND (s.revenue_type != 'Trade' OR s.revenue_type IS NULL)
                    AND COALESCE(s.gross_rate, 0) > 0
                GROUP BY s.broadcast_month
                ORDER BY 
                    CASE SUBSTR(s.broadcast_month, 1, 3)
                        WHEN 'Jan' THEN 1 WHEN 'Feb' THEN 2 WHEN 'Mar' THEN 3
                        WHEN 'Apr' THEN 4 WHEN 'May' THEN 5 WHEN 'Jun' THEN 6
                        WHEN 'Jul' THEN 7 WHEN 'Aug' THEN 8 WHEN 'Sep' THEN 9
                        WHEN 'Oct' THEN 10 WHEN 'Nov' THEN 11 WHEN 'Dec' THEN 12
                        ELSE 13
                    END
            """, (ae_name,))
            
            monthly = []
            for row in cursor.fetchall():
                monthly.append(MonthlyRevenue(
                    broadcast_month=row[0],
                    revenue=Decimal(str(row[1])),
                    spot_count=row[2]
                ))
            
            return monthly
    
    def get_ae_language_analysis(self, ae_name: str) -> List[LanguageRevenue]:
        """Get revenue breakdown by language for 2025 using gross rates"""
        with sqlite3.connect(self._db_path) as conn:
            cursor = conn.execute("""
                SELECT 
                    COALESCE(s.language_code, 'UNK') as language_code,
                    COALESCE(l.language_name, s.language_code, 'Unknown Language') as language_name,
                    SUM(COALESCE(s.gross_rate, 0)) as revenue,
                    COUNT(*) as spot_count,
                    COUNT(DISTINCT COALESCE(s.customer_id, s.bill_code)) as customer_count
                FROM spots s
                LEFT JOIN languages l ON s.language_code = l.language_code
                WHERE UPPER(TRIM(s.sales_person)) = UPPER(TRIM(?))
                    AND s.broadcast_month LIKE '%-25'
                    AND (s.revenue_type != 'Trade' OR s.revenue_type IS NULL)
                    AND COALESCE(s.gross_rate, 0) > 0
                GROUP BY 
                    COALESCE(s.language_code, 'UNK'),
                    COALESCE(l.language_name, s.language_code, 'Unknown Language')
                ORDER BY 
                    revenue DESC
            """, (ae_name,))
            
            results = []
            for row in cursor.fetchall():
                results.append(LanguageRevenue(
                    language_code=row[0],
                    language_name=row[1], 
                    revenue=Decimal(str(row[2])),
                    spot_count=row[3],
                    customer_count=row[4]
                ))
            
            logger.info(f"   📊 Language analysis found {len(results)} languages")
            return results

    def get_ae_raw_data_2025(self, ae_name: str) -> List[Dict]:
        """Get detailed raw spot data for AE in 2025 - matches RevenueDB format with CONSISTENT filtering"""
        with sqlite3.connect(self._db_path) as conn:
            cursor = conn.execute("""
                SELECT 
                    s.spot_id,
                    UPPER(TRIM(s.sales_person)) as ae_name,
                    COALESCE(sec.sector_name, 'Unknown') as sector,
                    COALESCE(c.normalized_name, s.bill_code) as customer,
                    s.broadcast_month,
                    COALESCE(s.gross_rate, 0) as gross_rate,
                    COALESCE(s.station_net, 0) as station_net,
                    s.air_date,
                    s.spot_type,
                    s.length_seconds,
                    s.time_in,
                    s.time_out,
                    COALESCE(s.language_code, '') as language_code,
                    COALESCE(l.language_name, s.language_code, '') as language_name,
                    COALESCE(m.market_code, '') as market_code,
                    COALESCE(a.agency_name, '') as agency_name,
                    s.revenue_type,
                    s.comments,
                    s.contract,
                    s.priority,
                    s.sequence_number,
                    s.line_number,
                    s.format,
                    s.media,
                    s.make_good,
                    s.broker_fees,
                    s.estimate,
                    s.affidavit_flag,
                    s.billing_type,
                    s.agency_flag
                FROM spots s
                LEFT JOIN customers c ON s.customer_id = c.customer_id
                LEFT JOIN sectors sec ON c.sector_id = sec.sector_id
                LEFT JOIN languages l ON s.language_code = l.language_code
                LEFT JOIN markets m ON s.market_id = m.market_id
                LEFT JOIN agencies a ON s.agency_id = a.agency_id
                WHERE UPPER(TRIM(s.sales_person)) = UPPER(TRIM(?))
                    AND s.broadcast_month LIKE '%-25'
                    AND (s.revenue_type != 'Trade' OR s.revenue_type IS NULL)
                    AND COALESCE(s.gross_rate, 0) > 0
                ORDER BY 
                    s.broadcast_month,
                    COALESCE(c.normalized_name, s.bill_code),
                    s.air_date,
                    s.spot_id
            """, (ae_name,))
            
            # Convert to list of dictionaries for easier DataFrame creation
            columns = [desc[0] for desc in cursor.description]
            raw_data = []
            
            for row in cursor.fetchall():
                row_dict = dict(zip(columns, row))
                raw_data.append(row_dict)
            
            return raw_data
    
    def get_available_aes(self) -> List[str]:
        """Get list of AEs with 2025 revenue data using gross rates"""
        with sqlite3.connect(self._db_path) as conn:
            cursor = conn.execute("""
                SELECT DISTINCT TRIM(s.sales_person) as ae_name
                FROM spots s
                WHERE s.broadcast_month LIKE '%-25'
                    AND (s.revenue_type != 'Trade' OR s.revenue_type IS NULL)
                    AND COALESCE(s.gross_rate, 0) > 0
                    AND s.sales_person IS NOT NULL 
                    AND TRIM(s.sales_person) != ''
                ORDER BY ae_name
            """)
            
            return [row[0] for row in cursor.fetchall()]
    
    def get_database_info(self) -> Dict[str, int]:
        """Get basic database statistics using gross rates"""
        with sqlite3.connect(self._db_path) as conn:
            cursor = conn.execute("""
                SELECT 
                    COUNT(*) as total_spots,
                    COUNT(DISTINCT sales_person) as total_aes,
                    COUNT(DISTINCT customer_id) as total_customers,
                    COUNT(*) FILTER (WHERE broadcast_month LIKE '%-25') as spots_2025
                FROM spots
                WHERE (revenue_type != 'Trade' OR revenue_type IS NULL)
                    AND gross_rate > 0
            """)
            
            row = cursor.fetchone()
            return {
                'total_spots': row[0],
                'total_aes': row[1], 
                'total_customers': row[2],
                'spots_2025': row[3]
            }

# Business Logic Layer
class RevenueService:
    """Service for revenue calculations and business logic"""
    
    def __init__(self, revenue_repo: RevenueRepository):
        self._revenue_repo = revenue_repo
    
    def calculate_ae_performance(self, ae_name: str) -> AEPerformance:
        """Calculate comprehensive AE performance metrics (pure business logic)"""
        customers = self._revenue_repo.get_ae_revenue_2025(ae_name)
        
        total_revenue = sum(customer.revenue_2025 for customer in customers)
        total_spots = sum(customer.spot_count for customer in customers)
        
        return AEPerformance(
            ae_name=ae_name,
            total_revenue=total_revenue,
            customer_count=len(customers),
            spot_count=total_spots,
            customers=customers
        )
    
    def get_sector_analysis(self, ae_name: str, years: List[int]) -> List[SectorSummary]:
        """Get sector trend analysis for specified years"""
        return self._revenue_repo.get_ae_sector_trends(ae_name, years)
    
    def get_monthly_performance(self, ae_name: str) -> List[MonthlyRevenue]:
        """Get monthly performance breakdown"""
        return self._revenue_repo.get_ae_monthly_breakdown_2025(ae_name)
    
    def get_language_analysis(self, ae_name: str) -> List[LanguageRevenue]:
        """Get language revenue analysis"""
        return self._revenue_repo.get_ae_language_analysis(ae_name)
    
    def get_raw_data_2025(self, ae_name: str) -> List[Dict]:
        """Get raw spot data for AE in 2025"""
        return self._revenue_repo.get_ae_raw_data_2025(ae_name)
    
    def get_new_vs_returning_customers(self, ae_name: str) -> Tuple[List[Customer], List[Customer]]:
        """Get new vs returning customers for 2025 (comparing to 2024)"""
        return self._revenue_repo.get_ae_new_vs_returning_customers(ae_name)

# File Transfer Service
class TailscaleFileTransfer:
    """Handle file transfers to desktop via Tailscale"""
    
    def __init__(self, desktop_name: str = "desktop-7402tkp"):
        self._desktop_name = desktop_name
    
    def copy_file_to_desktop(self, local_file_path: str) -> bool:
        """Copy file to desktop via Tailscale file cp"""
        try:
            # Use sudo tailscale file cp as shown in your example
            cmd = ["sudo", "tailscale", "file", "cp", local_file_path, f"{self._desktop_name}:"]
            
            logger.info(f"📤 Transferring {Path(local_file_path).name} to desktop...")
            
            result = subprocess.run(
                cmd, 
                capture_output=True, 
                text=True, 
                timeout=30
            )
            
            if result.returncode == 0:
                logger.info(f"   ✅ Successfully transferred to {self._desktop_name}")
                return True
            else:
                logger.warning(f"   ⚠️  Transfer failed: {result.stderr}")
                return False
                
        except subprocess.TimeoutExpired:
            logger.error(f"   ❌ Transfer timed out after 30 seconds")
            return False
        except Exception as e:
            logger.error(f"   ❌ Transfer error: {e}")
            return False
    
    def copy_directory_contents(self, directory_path: Path) -> Tuple[int, int]:
        """Copy all files in directory to desktop. Returns (success_count, total_count)"""
        files = list(directory_path.glob("*.xlsx"))
        
        if not files:
            logger.warning("No Excel files found to transfer")
            return 0, 0
        
        success_count = 0
        total_count = len(files)
        
        logger.info(f"📤 Transferring {total_count} files to desktop via Tailscale...")
        
        for file_path in files:
            if self.copy_file_to_desktop(str(file_path)):
                success_count += 1
        
        return success_count, total_count
    
    def test_connection(self) -> bool:
        """Test if desktop is reachable via Tailscale"""
        try:
            # Try to ping the desktop to verify connectivity
            cmd = ["tailscale", "ping", self._desktop_name]
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=10)
            
            if result.returncode == 0:
                logger.info(f"✅ Desktop {self._desktop_name} is reachable via Tailscale")
                return True
            else:
                logger.warning(f"⚠️  Desktop {self._desktop_name} not reachable: {result.stderr}")
                return False
                
        except subprocess.TimeoutExpired:
            logger.error("❌ Tailscale ping timed out")
            return False
        except Exception as e:
            logger.error(f"❌ Connection test error: {e}")
            return False

# Presentation Layer
class ExcelReportGenerator:
    """Generate Excel reports from domain entities"""
    
    def create_ae_worksheet(
        self, 
        ae_performance: AEPerformance, 
        sector_analysis: List[SectorSummary],
        monthly_breakdown: List[MonthlyRevenue],
        language_analysis: List[LanguageRevenue],
        new_customers: List[Customer],
        returning_customers: List[Customer],
        raw_data_2025: List[Dict],
        output_dir: Path
    ) -> str:
        """Create individual AE budget worksheet with professional formatting"""
        ae_name = ae_performance.ae_name
        output_file = output_dir / f"{ae_name}_Budget_Planning_2026.xlsx"
        
        logger.info(f"Creating worksheet for {ae_name}")
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Customer detail sheet (with numeric values and totals)
            self._create_customer_detail_sheet(writer, ae_performance)
            
            # Raw Data 2025 sheet (numeric values with totals)
            self._create_raw_data_sheet(writer, raw_data_2025, ae_name)
            
            # Monthly breakdown sheet (with numeric values and totals)
            self._create_monthly_breakdown_sheet(writer, monthly_breakdown)
            
            # Sector trends sheet (with numeric values and totals)
            self._create_sector_trends_sheet(writer, sector_analysis)
            
            # Language analysis sheet (with numeric values and totals)
            self._create_language_sheet(writer, language_analysis)
            
            # New vs Returning Customers sheet (with numeric values and totals)
            self._create_new_vs_returning_sheet(writer, new_customers, returning_customers)
            
            # Apply professional formatting to all sheets
            self._apply_professional_formatting(writer.book)
        
        logger.info(f"Created: {output_file}")
        return str(output_file)
    
    def _apply_professional_formatting(self, workbook) -> None:
        """Apply professional Excel formatting to all sheets"""
        from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Create currency style
        currency_style = NamedStyle(name="currency")
        currency_style.number_format = '$#,##0.00'
        
        # Create header style
        header_style = NamedStyle(name="header")
        header_style.font = Font(bold=True, size=11)
        header_style.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        header_style.alignment = Alignment(horizontal="center", vertical="center")
        header_style.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Create total row style
        total_style = NamedStyle(name="total")
        total_style.font = Font(bold=True, size=11)
        total_style.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        total_style.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='double'), bottom=Side(style='thin')
        )
        
        # Register styles with workbook
        try:
            workbook.add_named_style(currency_style)
            workbook.add_named_style(header_style)
            workbook.add_named_style(total_style)
        except ValueError:
            # Styles already exist
            pass
        
        # Apply formatting to each worksheet
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            
            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Format header row (row 1)
            if ws.max_row > 0:
                for cell in ws[1]:
                    if cell.value:
                        cell.style = header_style
            
            # Apply currency formatting to revenue columns
            self._format_currency_columns(ws)
            
            # Format total rows (look for rows containing "TOTAL")
            self._format_total_rows(ws, total_style)
    
    def _format_currency_columns(self, worksheet) -> None:
        """Apply currency formatting to columns containing dollar amounts"""
        from openpyxl.utils import get_column_letter
        
        currency_keywords = ['revenue', 'total', 'avg', 'rate', 'gross', 'net']
        
        # Check header row for currency columns
        if worksheet.max_row == 0:
            return
            
        header_row = list(worksheet[1])
        currency_columns = []
        
        for col_idx, cell in enumerate(header_row, 1):
            if cell.value and any(keyword in str(cell.value).lower() for keyword in currency_keywords):
                currency_columns.append(col_idx)
        
        # Apply currency formatting to identified columns
        for col_idx in currency_columns:
            column_letter = get_column_letter(col_idx)
            for row_idx in range(2, worksheet.max_row + 1):
                cell = worksheet[f"{column_letter}{row_idx}"]
                if cell.value and isinstance(cell.value, (int, float)):
                    cell.style = "currency"
    
    def _format_total_rows(self, worksheet, total_style) -> None:
        """Apply total row formatting to rows containing 'TOTAL'"""
        for row_idx in range(1, worksheet.max_row + 1):
            row = list(worksheet[row_idx])
            if any(cell.value and 'TOTAL' in str(cell.value).upper() for cell in row):
                for cell in row:
                    if cell.value is not None:
                        current_style = cell.style
                        cell.style = total_style
                        # Preserve currency formatting for total rows
                        if hasattr(current_style, 'number_format') and '$' in current_style.number_format:
                            cell.number_format = '$#,##0.00'
    
    def _create_customer_detail_sheet(self, writer: pd.ExcelWriter, ae_performance: AEPerformance) -> None:
        """Create customer detail sheet with numeric values and totals"""
        if not ae_performance.customers:
            return
        
        customer_data = []
        for customer in sorted(ae_performance.customers, key=lambda c: c.revenue_2025, reverse=True):
            customer_data.append({
                'Customer ID': customer.customer_id if customer.customer_id > 0 else 'Unmatched',
                'Customer': customer.name,
                'Sector': customer.sector,
                '2025 Revenue': float(customer.revenue_2025),  # Numeric for Excel formatting
                'Spots': customer.spot_count,
                'Avg per Spot': float(customer.revenue_2025 / max(customer.spot_count, 1))  # Numeric for Excel formatting
            })
        
        # Add totals row
        total_revenue = sum(c.revenue_2025 for c in ae_performance.customers)
        total_spots = sum(c.spot_count for c in ae_performance.customers)
        
        customer_data.append({
            'Customer ID': '',
            'Customer': 'TOTAL',
            'Sector': '',
            '2025 Revenue': float(total_revenue),
            'Spots': total_spots,
            'Avg per Spot': float(total_revenue / max(total_spots, 1))
        })
        
        customer_df = pd.DataFrame(customer_data)
        customer_df.to_excel(writer, sheet_name='Customer Detail', index=False)
    
    def _create_raw_data_sheet(self, writer: pd.ExcelWriter, raw_data: List[Dict], ae_name: str) -> None:
        """Create raw data sheet matching RevenueDB format with numeric values and totals"""
        if not raw_data:
            # Create empty sheet with message
            empty_df = pd.DataFrame({'Message': [f'No raw data available for {ae_name}']})
            empty_df.to_excel(writer, sheet_name='Raw Data 2025', index=False)
            return
        
        # Convert to DataFrame
        raw_df = pd.DataFrame(raw_data)
        
        # Create month columns similar to your RevenueDB format
        # Get all unique broadcast months and sort them
        months = sorted(raw_df['broadcast_month'].dropna().unique()) if 'broadcast_month' in raw_df.columns else []
        
        # Create pivot table to match your format: Customer rows x Month columns
        if months:
            # Create pivot table with customers as rows and months as revenue columns using gross rates
            pivot_df = raw_df.pivot_table(
                index=['ae_name', 'sector', 'customer'],
                columns='broadcast_month',
                values='gross_rate',  # Using gross_rate for budget planning
                aggfunc='sum',
                fill_value=0
            ).reset_index()
            
            # Rename columns to match your format
            pivot_df.columns.name = None
            pivot_df = pivot_df.rename(columns={
                'ae_name': 'AE',
                'sector': 'Sector', 
                'customer': 'Customer'
            })
            
            # Add total column
            month_columns = [col for col in pivot_df.columns if col not in ['AE', 'Sector', 'Customer']]
            pivot_df['Total'] = pivot_df[month_columns].sum(axis=1)
            
            # Sort by total revenue descending
            pivot_df = pivot_df.sort_values('Total', ascending=False)
            
            # Add quarterly columns (Q1, Q2, Q3, Q4) based on months
            quarter_mapping = {
                'Jan': 'Q1', 'Feb': 'Q1', 'Mar': 'Q1',
                'Apr': 'Q2', 'May': 'Q2', 'Jun': 'Q2', 
                'Jul': 'Q3', 'Aug': 'Q3', 'Sep': 'Q3',
                'Oct': 'Q4', 'Nov': 'Q4', 'Dec': 'Q4'
            }
            
            # Sort months in proper chronological order
            month_order = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                          'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
            
            # Sort month columns by actual month order, not alphabetically
            def month_sort_key(col):
                if '-' in col:
                    month_abbr = col.split('-')[0]
                    if month_abbr in month_order:
                        return month_order.index(month_abbr)
                return 999  # Put unrecognized formats at end
            
            sorted_month_columns = sorted(month_columns, key=month_sort_key)
            
            quarters = {}
            for month_col in sorted_month_columns:
                if '-' in month_col:
                    month_abbr = month_col.split('-')[0]
                    quarter = quarter_mapping.get(month_abbr, 'Q1')
                    if quarter not in quarters:
                        quarters[quarter] = []
                    quarters[quarter].append(month_col)
            
            # Add quarter total columns
            quarter_columns = []
            for quarter in ['Q1', 'Q2', 'Q3', 'Q4']:
                if quarter in quarters:
                    pivot_df[quarter] = pivot_df[quarters[quarter]].sum(axis=1)
                    quarter_columns.append(quarter)
            
            # Reorder columns: AE, Sector, Customer, Months (chronological), Quarters, Total
            final_columns = ['AE', 'Sector', 'Customer'] + sorted_month_columns + quarter_columns + ['Total']
            pivot_df = pivot_df[[col for col in final_columns if col in pivot_df.columns]]
            
            # Ensure all revenue values are numeric (not formatted strings)
            numeric_columns = [col for col in pivot_df.columns if col not in ['AE', 'Sector', 'Customer']]
            for col in numeric_columns:
                pivot_df[col] = pd.to_numeric(pivot_df[col], errors='coerce').fillna(0)
            
            # Add TOTALS row at the bottom
            totals_row = {
                'AE': '',
                'Sector': '',
                'Customer': 'TOTAL'
            }
            
            for col in numeric_columns:
                if col in pivot_df.columns:
                    totals_row[col] = pivot_df[col].sum()
            
            # Convert totals row to DataFrame and append
            totals_df = pd.DataFrame([totals_row])
            pivot_df = pd.concat([pivot_df, totals_df], ignore_index=True)
            
            # Write to Excel with proper numeric formatting
            pivot_df.to_excel(writer, sheet_name='Raw Data 2025', index=False)
            
            logger.info(f"   📊 Raw Data 2025 sheet: {len(pivot_df)-1} customer records with {len(month_columns)} months + totals")
            
        else:
            # Fallback: create detailed spot-level data if no months available
            # Select key columns matching your format
            key_columns = [
                'ae_name', 'sector', 'customer', 'broadcast_month', 
                'gross_rate', 'station_net', 'air_date', 'spot_type',
                'language_code', 'market_code', 'agency_name'
            ]
            
            available_columns = [col for col in key_columns if col in raw_df.columns]
            detail_df = raw_df[available_columns].copy()
            
            # Rename columns for clarity
            column_rename = {
                'ae_name': 'AE',
                'sector': 'Sector',
                'customer': 'Customer',
                'broadcast_month': 'Month',
                'gross_rate': 'Gross Rate',
                'station_net': 'Station Net',
                'air_date': 'Air Date',
                'spot_type': 'Type',
                'language_code': 'Language',
                'market_code': 'Market',
                'agency_name': 'Agency'
            }
            detail_df = detail_df.rename(columns=column_rename)
            
            # Ensure numeric columns are properly formatted
            for col in ['Gross Rate', 'Station Net']:
                if col in detail_df.columns:
                    detail_df[col] = pd.to_numeric(detail_df[col], errors='coerce').fillna(0)
            
            detail_df.to_excel(writer, sheet_name='Raw Data 2025', index=False)
            logger.info(f"   📊 Raw Data 2025 sheet: {len(detail_df)} spot records")
    
    def _create_monthly_breakdown_sheet(self, writer: pd.ExcelWriter, monthly: List[MonthlyRevenue]) -> None:
        """Create monthly breakdown sheet with numeric values and totals"""
        if not monthly:
            return
        
        monthly_data = []
        for month_data in monthly:
            avg_per_spot = month_data.revenue / max(month_data.spot_count, 1)
            monthly_data.append({
                'Month': month_data.broadcast_month,
                'Revenue': float(month_data.revenue),  # Numeric for Excel formatting
                'Spots': month_data.spot_count,
                'Avg per Spot': float(avg_per_spot)  # Numeric for Excel formatting
            })
        
        # Add totals row
        total_revenue = sum(m.revenue for m in monthly)
        total_spots = sum(m.spot_count for m in monthly)
        
        monthly_data.append({
            'Month': 'TOTAL',
            'Revenue': float(total_revenue),
            'Spots': total_spots,
            'Avg per Spot': float(total_revenue / max(total_spots, 1))
        })
        
        monthly_df = pd.DataFrame(monthly_data)
        monthly_df.to_excel(writer, sheet_name='Monthly Breakdown', index=False)
    
    def _create_sector_trends_sheet(self, writer: pd.ExcelWriter, sector_analysis: List[SectorSummary]) -> None:
        """Create sector trends sheet with numeric values"""
        if not sector_analysis:
            return
        
        # Get all years from the analysis
        all_years = set()
        for sector in sector_analysis:
            all_years.update(sector.revenue_by_year.keys())
        
        years = sorted(all_years)
        
        # Create sector trends data
        sector_data = []
        for sector in sector_analysis:
            row = {'Sector': sector.sector_name}
            
            for year in years:
                revenue = sector.revenue_by_year.get(year, Decimal('0'))
                row[str(year)] = float(revenue)  # Numeric for Excel formatting
            
            row['Total'] = float(sector.total_revenue)  # Numeric for Excel formatting
            row['Customers'] = sector.customer_count
            sector_data.append(row)
        
        # Add totals row
        totals_row = {'Sector': 'TOTAL'}
        grand_total = Decimal('0')
        total_customers = 0
        
        for year in years:
            year_total = sum(s.revenue_by_year.get(year, Decimal('0')) for s in sector_analysis)
            totals_row[str(year)] = float(year_total)  # Numeric for Excel formatting
            grand_total += year_total
        
        totals_row['Total'] = float(grand_total)  # Numeric for Excel formatting
        totals_row['Customers'] = sum(s.customer_count for s in sector_analysis)
        sector_data.append(totals_row)
        
        sector_df = pd.DataFrame(sector_data)
        sector_df.to_excel(writer, sheet_name='Sector Trends', index=False)
    
    def _create_language_sheet(self, writer: pd.ExcelWriter, language_analysis: List[LanguageRevenue]) -> None:
        """Create language revenue analysis sheet with numeric values and totals"""
        if not language_analysis:
            # Create empty sheet with message
            empty_df = pd.DataFrame({'Message': ['No language data available for analysis']})
            empty_df.to_excel(writer, sheet_name='Language Analysis', index=False)
            return
        
        # Create the language analysis data
        language_data = []
        
        for item in language_analysis:
            avg_per_spot = item.revenue / max(item.spot_count, 1)
            avg_per_customer = item.revenue / max(item.customer_count, 1)
            
            language_data.append({
                'Language Code': item.language_code,
                'Language Name': item.language_name,
                'Revenue': float(item.revenue),  # Numeric for Excel formatting
                'Spots': item.spot_count,
                'Customers': item.customer_count,
                'Avg per Spot': float(avg_per_spot),  # Numeric for Excel formatting
                'Avg per Customer': float(avg_per_customer)  # Numeric for Excel formatting
            })
        
        # Add totals row
        total_revenue = sum(item.revenue for item in language_analysis)
        total_spots = sum(item.spot_count for item in language_analysis)
        total_customers = sum(item.customer_count for item in language_analysis)
        
        language_data.append({
            'Language Code': '',
            'Language Name': 'TOTAL',
            'Revenue': float(total_revenue),  # Numeric for Excel formatting
            'Spots': total_spots,
            'Customers': total_customers,
            'Avg per Spot': float(total_revenue / max(total_spots, 1)),  # Numeric for Excel formatting
            'Avg per Customer': float(total_revenue / max(total_customers, 1))  # Numeric for Excel formatting
        })
        
        language_df = pd.DataFrame(language_data)
        language_df.to_excel(writer, sheet_name='Language Analysis', index=False)
        
        logger.info(f"   📊 Language Analysis sheet: {len(language_analysis)} languages analyzed")
    
    def _create_new_vs_returning_sheet(self, writer: pd.ExcelWriter, new_customers: List[Customer], returning_customers: List[Customer]) -> None:
        """Create new vs returning customers analysis sheet with numeric values and proper totals"""
        
        # Calculate summary metrics
        total_customers = len(new_customers) + len(returning_customers)
        new_revenue = sum(c.revenue_2025 for c in new_customers)
        returning_revenue = sum(c.revenue_2025 for c in returning_customers)
        total_revenue = new_revenue + returning_revenue
        
        new_spots = sum(c.spot_count for c in new_customers)
        returning_spots = sum(c.spot_count for c in returning_customers)
        total_spots = new_spots + returning_spots
        
        # Calculate percentages
        new_customer_pct = (len(new_customers) / max(total_customers, 1)) * 100
        returning_customer_pct = (len(returning_customers) / max(total_customers, 1)) * 100
        new_revenue_pct = (new_revenue / max(total_revenue, 1)) * 100
        returning_revenue_pct = (returning_revenue / max(total_revenue, 1)) * 100
        
        # Create summary data at the top
        summary_data = [
            ['NEW vs RETURNING CUSTOMERS ANALYSIS (2025)', '', '', '', '', ''],
            ['Comparison Base: 2024 vs 2025', '', '', '', '', ''],
            ['', '', '', '', '', ''],
            ['SUMMARY METRICS', '', '', '', '', ''],
            ['', 'Count', '% of Total', 'Revenue', '% of Revenue', 'Avg per Customer'],
            ['New Customers (2025)', len(new_customers), f'{new_customer_pct:.1f}%', 
             float(new_revenue), f'{new_revenue_pct:.1f}%', 
             float(new_revenue / max(len(new_customers), 1))],
            ['Returning Customers', len(returning_customers), f'{returning_customer_pct:.1f}%', 
             float(returning_revenue), f'{returning_revenue_pct:.1f}%', 
             float(returning_revenue / max(len(returning_customers), 1))],
            ['TOTAL', total_customers, '100.0%', 
             float(total_revenue), '100.0%', 
             float(total_revenue / max(total_customers, 1))],
            ['', '', '', '', '', ''],
        ]
        
        # Add NEW CUSTOMERS section
        summary_data.extend([
            ['NEW CUSTOMERS (First-Time in 2025)', '', '', '', '', ''],
            ['Customer', 'Sector', '2025 Revenue', 'Spots', 'Avg per Spot', '']
        ])
        
        new_total_revenue = Decimal('0')
        new_total_spots = 0
        
        for customer in new_customers:
            avg_per_spot = customer.revenue_2025 / max(customer.spot_count, 1)
            summary_data.append([
                customer.name,
                customer.sector, 
                float(customer.revenue_2025),  # Numeric for Excel formatting
                customer.spot_count,
                float(avg_per_spot),  # Numeric for Excel formatting
                ''
            ])
            new_total_revenue += customer.revenue_2025
            new_total_spots += customer.spot_count
        
        # Add new customers subtotal
        summary_data.append([
            'NEW CUSTOMERS TOTAL',
            '',
            float(new_total_revenue),
            new_total_spots,
            float(new_total_revenue / max(new_total_spots, 1)),
            ''
        ])
        
        # Add spacing and RETURNING CUSTOMERS section
        summary_data.extend([
            ['', '', '', '', '', ''],
            ['RETURNING CUSTOMERS (Active in Both 2024 & 2025)', '', '', '', '', ''],
            ['Customer', 'Sector', '2025 Revenue', 'Spots', 'Avg per Spot', '']
        ])
        
        returning_total_revenue = Decimal('0')
        returning_total_spots = 0
        
        for customer in returning_customers:
            avg_per_spot = customer.revenue_2025 / max(customer.spot_count, 1)
            summary_data.append([
                customer.name,
                customer.sector,
                float(customer.revenue_2025),  # Numeric for Excel formatting
                customer.spot_count,
                float(avg_per_spot),  # Numeric for Excel formatting
                ''
            ])
            returning_total_revenue += customer.revenue_2025
            returning_total_spots += customer.spot_count
        
        # Add returning customers subtotal
        summary_data.append([
            'RETURNING CUSTOMERS TOTAL',
            '',
            float(returning_total_revenue),
            returning_total_spots,
            float(returning_total_revenue / max(returning_total_spots, 1)),
            ''
        ])
        
        # Add grand total
        summary_data.extend([
            ['', '', '', '', '', ''],
            ['GRAND TOTAL',
             '',
             float(new_total_revenue + returning_total_revenue),
             new_total_spots + returning_total_spots,
             float((new_total_revenue + returning_total_revenue) / max(new_total_spots + returning_total_spots, 1)),
             ''
            ]
        ])
        
        # Create DataFrame and write to Excel
        analysis_df = pd.DataFrame(summary_data, columns=[
            'Customer/Metric', 'Sector/Count', 'Revenue/Value', 'Spots', 'Extra1', 'Extra2'
        ])
        
        analysis_df.to_excel(writer, sheet_name='New vs Returning', index=False, header=False)
        
        logger.info(f"   👥 New vs Returning sheet: {len(new_customers)} new, {len(returning_customers)} returning customers")

    def create_summary_report(self, ae_performances: List[AEPerformance], output_dir: Path) -> str:
        """Create summary report for sales manager"""
        output_file = output_dir / "Sales_Manager_Summary_2026_Planning.xlsx"
        
        logger.info("Creating summary report for sales manager")
        
        # Calculate totals
        total_revenue = sum(ae.total_revenue for ae in ae_performances)
        total_customers = sum(ae.customer_count for ae in ae_performances)
        total_spots = sum(ae.spot_count for ae in ae_performances)
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Company summary
            company_data = [
                ['SALES MANAGER - 2026 BUDGET PLANNING SUMMARY', '', '', ''],
                ['Report Date:', datetime.now().strftime('%B %d, %Y'), '', ''],
                ['Data Source:', 'CTV BookedBiz Database', '', ''],
                ['', '', '', ''],
                ['2025 COMPANY PERFORMANCE', '', '', ''],
                ['Total Company Revenue:', f"${total_revenue:,.2f}", '', ''],
                ['Total Active Customers:', str(total_customers), '', ''],
                ['Total Spots:', str(total_spots), '', ''],
                ['Average per AE:', f"${total_revenue / max(len(ae_performances), 1):,.2f}", '', ''],
                ['', '', '', ''],
                ['AE PERFORMANCE SUMMARY', '', '', ''],
                ['Account Executive', '2025 Revenue', 'Customers', 'Spots', 'Avg per Customer'],
            ]
            
            for ae in ae_performances:
                avg_per_customer = ae.total_revenue / max(ae.customer_count, 1)
                company_data.append([
                    ae.ae_name,
                    f"${ae.total_revenue:,.2f}",
                    str(ae.customer_count),
                    str(ae.spot_count),
                    f"${avg_per_customer:,.2f}"
                ])
            
            # Add totals row
            company_data.append([
                'TOTAL',
                f"${total_revenue:,.2f}",
                str(total_customers),
                str(total_spots),
                f"${total_revenue / max(total_customers, 1):,.2f}"
            ])
            
            company_df = pd.DataFrame(company_data, columns=['Metric', 'Value', 'Extra1', 'Extra2', 'Extra3'])
            company_df.to_excel(writer, sheet_name='Company Summary', index=False, header=False)
        
        logger.info(f"Created summary report: {output_file}")
        return str(output_file)

# Application Layer
def find_database() -> Optional[str]:
    """Find the production database file"""
    possible_paths = [
        './data/database/production.db',
        './data/database/production_dev.db',
        './production.db',
        './data/production.db'
    ]
    
    for path in possible_paths:
        if Path(path).exists():
            return path
    
    # Look for any .db files in data/database
    db_dir = Path('./data/database')
    if db_dir.exists():
        db_files = list(db_dir.glob('*.db'))
        if db_files:
            # Return the largest database file (likely production)
            return str(max(db_files, key=lambda x: x.stat().st_size))
    
    return None

def main() -> int:
    """Main execution function with dependency injection and file transfer"""
    try:
        logger.info("🚀 Starting Database-Driven AE Budget Sheet Generation for 2026 Planning")
        logger.info("=" * 70)
        
        # Find database
        db_path = find_database()
        if not db_path:
            logger.error("❌ No database file found")
            logger.error("Looking for database files in:")
            for path in ['./data/database/', './']:
                logger.error(f"  - {path}*.db")
            return 1
        
        logger.info(f"📁 Using database: {db_path}")
        
        # Initialize repositories and services (dependency injection)
        revenue_repo = RevenueRepository(db_path)
        revenue_service = RevenueService(revenue_repo)
        report_generator = ExcelReportGenerator()
        file_transfer = TailscaleFileTransfer("desktop-7402tkp")  # Your desktop name
        
        # Test Tailscale connection first
        logger.info("🔗 Testing Tailscale connection to desktop...")
        if not file_transfer.test_connection():
            logger.warning("⚠️  Desktop not reachable - files will be generated locally only")
            transfer_enabled = False
        else:
            transfer_enabled = True
        
        # Get database info
        db_info = revenue_repo.get_database_info()
        logger.info(f"📊 Database contains: {db_info['spots_2025']:,} spots in 2025 from {db_info['total_aes']} AEs")
        
        # Get available AEs
        available_aes = revenue_repo.get_available_aes()
        logger.info(f"👥 Found {len(available_aes)} AEs with 2025 revenue data")
        
        if not available_aes:
            logger.error("❌ No AEs found with 2025 revenue data")
            return 1
        
        # Create output directory
        output_dir = Path("budget_reports_2026_db")
        output_dir.mkdir(exist_ok=True)
        
        # Generate reports for each AE
        ae_performances = []
        generated_files = []
        
        for ae_name in available_aes:
            logger.info(f"📊 Processing {ae_name}...")
            
            # Calculate performance
            ae_performance = revenue_service.calculate_ae_performance(ae_name)
            
            if ae_performance.total_revenue > 0:
                # Get supporting data
                sector_analysis = revenue_service.get_sector_analysis(ae_name, [2021, 2022, 2023, 2024, 2025])
                monthly_breakdown = revenue_service.get_monthly_performance(ae_name)
                language_analysis = revenue_service.get_language_analysis(ae_name)
                new_customers, returning_customers = revenue_service.get_new_vs_returning_customers(ae_name)
                raw_data_2025 = revenue_service.get_raw_data_2025(ae_name)
                
                # Generate report
                output_file = report_generator.create_ae_worksheet(
                    ae_performance, 
                    sector_analysis, 
                    monthly_breakdown,
                    language_analysis,
                    new_customers,
                    returning_customers,
                    raw_data_2025,
                    output_dir
                )
                
                ae_performances.append(ae_performance)
                generated_files.append(output_file)
                
                logger.info(f"   ✅ {ae_name}: ${ae_performance.total_revenue:,.2f} from {ae_performance.customer_count} customers ({ae_performance.spot_count} spots)")
            else:
                logger.warning(f"   ⚠️  {ae_name}: No revenue data found")
        
        # Create summary report
        if ae_performances:
            summary_file = report_generator.create_summary_report(ae_performances, output_dir)
            generated_files.append(summary_file)
        
        # Transfer files to desktop via Tailscale
        if transfer_enabled and generated_files:
            logger.info("\n" + "=" * 70)
            logger.info("📤 TRANSFERRING FILES TO DESKTOP VIA TAILSCALE")
            logger.info("=" * 70)
            
            success_count, total_count = file_transfer.copy_directory_contents(output_dir)
            
            if success_count == total_count:
                logger.info(f"✅ All {total_count} files successfully transferred to desktop!")
            elif success_count > 0:
                logger.warning(f"⚠️  {success_count}/{total_count} files transferred successfully")
            else:
                logger.error(f"❌ Failed to transfer any files to desktop")
        elif not transfer_enabled:
            logger.info(f"\n📁 Files generated locally in: {output_dir}")
            logger.info("💡 Files not transferred due to Tailscale connection issue")
        
        # Final summary
        total_revenue = sum(ae.total_revenue for ae in ae_performances)
        total_customers = sum(ae.customer_count for ae in ae_performances)
        total_spots = sum(ae.spot_count for ae in ae_performances)
        
        logger.info("\n" + "=" * 70)
        logger.info("✅ DATABASE-DRIVEN BUDGET SHEET GENERATION COMPLETE!")
        logger.info("=" * 70)
        logger.info(f"📁 Output Directory: {output_dir}")
        logger.info(f"📊 Generated Files: {len(generated_files)}")
        logger.info(f"👥 Active AEs: {len(ae_performances)}")
        logger.info(f"💰 Total Company Revenue: ${total_revenue:,.2f}")
        logger.info(f"🏢 Total Customers: {total_customers}")
        logger.info(f"📺 Total Spots: {total_spots:,}")
        
        if transfer_enabled:
            logger.info(f"📤 Files transferred to: desktop-7402tkp")
        
        logger.info(f"\n🎯 Ready for 2026 Budget Planning Sessions!")
        logger.info("📈 Data sourced directly from CTV BookedBiz Database")
        
        return 0
        
    except Exception as e:
        logger.error(f"❌ Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        return 1

if __name__ == "__main__":
    exit(main())