#!/usr/bin/env python3
"""
Enhanced weekly update command with automatic market setup and language assignment processing.
Replaces open month data while protecting closed historical months.
Automatically creates missing markets, schedule assignments, and processes language assignments.
"""

import sys
import argparse
import sqlite3
from pathlib import Path
from datetime import datetime
from typing import Dict, Set

# Add tqdm for progress bars
from tqdm import tqdm

# Add src to path
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))


from services.broadcast_month_import_service import BroadcastMonthImportService, BroadcastMonthImportError
from services.import_integration_utilities import get_excel_import_summary, validate_excel_for_import
from database.connection import DatabaseConnection


class WeeklyMarketSetupManager:
    """Lightweight market setup manager optimized for weekly updates."""
    
    def __init__(self, db_connection: DatabaseConnection):
        self.db = db_connection
    
    def scan_excel_for_new_markets(self, excel_file: str) -> Dict[str, Dict]:
        """
        Quick scan of Excel file to detect any new market codes.
        Optimized for smaller weekly files.
        
        Returns:
            Dict mapping new market_code -> {'earliest_date': date, 'spot_count': int}
        """
        try:
            from openpyxl import load_workbook
            
            workbook = load_workbook(excel_file, read_only=True, data_only=True)
            worksheet = workbook.active
            
            # Find required columns
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
            market_col_index = None
            air_date_col_index = None
            
            for i, header in enumerate(header_row):
                if header:
                    header_str = str(header).strip().lower()
                    if header_str in ['market_name', 'market', 'market_code']:
                        market_col_index = i
                    elif header_str in ['air_date', 'date', 'airdate']:
                        air_date_col_index = i
            
            if market_col_index is None:
                return {}
            
            # Get existing markets to identify new ones
            existing_markets = self.get_existing_markets()
            
            markets_data = {}
            new_markets_found = set()
            
            # Get total rows for progress bar
            total_rows = worksheet.max_row - 1  # Exclude header
            
            # Process rows to find new markets with progress bar
            with tqdm(total=total_rows, desc="🔍 Scanning for new markets", unit=" rows") as pbar:
                for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                    pbar.update(1)
                    
                    if not any(cell for cell in row):
                        continue
                    
                    if market_col_index < len(row):
                        market_value = row[market_col_index]
                        
                        if market_value:
                            market_code = str(market_value).strip()
                            
                            # Only track if it's a NEW market
                            if market_code not in existing_markets:
                                new_markets_found.add(market_code)
                                
                                # Get air date if available
                                air_date = None
                                if air_date_col_index and air_date_col_index < len(row):
                                    air_date_value = row[air_date_col_index]
                                    if air_date_value:
                                        try:
                                            if isinstance(air_date_value, datetime):
                                                air_date = air_date_value.date()
                                            else:
                                                air_date = datetime.strptime(str(air_date_value), '%Y-%m-%d').date()
                                        except:
                                            pass
                                
                                # Track new market data
                                if market_code not in markets_data:
                                    markets_data[market_code] = {
                                        'earliest_date': air_date,
                                        'latest_date': air_date,
                                        'spot_count': 0
                                    }
                                else:
                                    if air_date:
                                        if not markets_data[market_code]['earliest_date'] or air_date < markets_data[market_code]['earliest_date']:
                                            markets_data[market_code]['earliest_date'] = air_date
                                        if not markets_data[market_code]['latest_date'] or air_date > markets_data[market_code]['latest_date']:
                                            markets_data[market_code]['latest_date'] = air_date
                                
                                markets_data[market_code]['spot_count'] += 1
                    
                    # Update progress description with found markets
                    if len(new_markets_found) > 0:
                        pbar.set_description(f"🔍 Scanning ({len(new_markets_found)} new markets found)")
            
            workbook.close()
            
            if markets_data:
                # Only show summary, not detailed list
                total_spots = sum(data['spot_count'] for data in markets_data.values())
                tqdm.write(f"✅ Found {len(markets_data)} new markets with {total_spots:,} spots")
            
            return markets_data
            
        except Exception as e:
            tqdm.write(f"⚠️  Warning: Could not scan for new markets: {str(e)}")
            return {}
    
    def get_existing_markets(self) -> Dict[str, int]:
        """Get existing markets from database."""
        with self.db.transaction() as conn:
            cursor = conn.execute("SELECT market_code, market_id FROM markets")
            return {row[0]: row[1] for row in cursor.fetchall()}
    
    def create_new_markets(self, new_markets: Dict[str, Dict]) -> Dict[str, int]:
        """Create new markets found in weekly data."""
        if not new_markets:
            return {}
        
        created_markets = {}
        
        with self.db.transaction() as conn:
            # Create progress bar for market creation
            with tqdm(total=len(new_markets), desc="🏗️  Creating markets", unit=" markets") as pbar:
                for market_code, market_data in sorted(new_markets.items()):
                    # Generate market name
                    market_name = self._generate_market_name(market_code)
                    
                    cursor = conn.execute("""
                        INSERT INTO markets (market_code, market_name) 
                        VALUES (?, ?)
                    """, (market_code, market_name))
                    
                    market_id = cursor.lastrowid
                    created_markets[market_code] = market_id
                    
                    pbar.update(1)
                    pbar.set_description(f"🏗️  Created {market_code}")
        
        tqdm.write(f"✅ Created {len(created_markets)} new markets")
        return created_markets
    
    def setup_schedules_for_new_markets(self, new_markets: Dict[str, Dict], market_mapping: Dict[str, int]) -> int:
        """Setup schedule assignments for newly created markets."""
        if not new_markets:
            return 0
        
        assignments_created = 0
        
        with self.db.transaction() as conn:
            # Create progress bar for schedule setup
            with tqdm(total=len(new_markets), desc="🗓️  Setting up schedules", unit=" assignments") as pbar:
                for market_code, market_data in new_markets.items():
                    market_id = market_mapping[market_code]
                    
                    # Use earliest date if available, otherwise use current date
                    effective_date = market_data['earliest_date'] if market_data['earliest_date'] else datetime.now().date()
                    
                    cursor = conn.execute("""
                        INSERT INTO schedule_market_assignments 
                        (schedule_id, market_id, effective_start_date, assignment_priority)
                        VALUES (1, ?, ?, 1)
                    """, (market_id, effective_date))
                    
                    assignments_created += 1
                    pbar.update(1)
                    pbar.set_description(f"🗓️  Setup {market_code}")
        
        tqdm.write(f"✅ Created {assignments_created} schedule assignments")
        return assignments_created
    
    def _generate_market_name(self, market_code: str) -> str:
        """Generate proper market name from code."""
        name_mappings = {
            'NYC': 'NEW YORK',
            'LAX': 'LOS ANGELES', 
            'SFO': 'SAN FRANCISCO',
            'SEA': 'SEATTLE',
            'CHI': 'CHICAGO',
            'MSP': 'MINNEAPOLIS',
            'DAL': 'DALLAS',
            'HOU': 'HOUSTON',
            'WDC': 'WASHINGTON DC',
            'CVC': 'CENTRAL VALLEY',
            'CMP': 'CHI MSP',
            'MMT': 'MAMMOTH',
            'ADMIN': 'ADMINISTRATIVE'
        }
        
        return name_mappings.get(market_code, market_code.upper().replace('_', ' '))
    
    def execute_weekly_market_setup(self, excel_file: str) -> Dict:
        """Execute lightweight market setup for weekly data."""
        start_time = datetime.now()
        
        # Step 1: Scan for new markets
        new_markets = self.scan_excel_for_new_markets(excel_file)
        
        if not new_markets:
            return {
                'new_markets_found': 0,
                'markets_created': 0,
                'schedules_created': 0,
                'duration_seconds': (datetime.now() - start_time).total_seconds()
            }
        
        # Step 2: Create new markets
        created_markets = self.create_new_markets(new_markets)
        
        # Step 3: Setup schedules for new markets
        schedules_created = self.setup_schedules_for_new_markets(new_markets, created_markets)
        
        duration = (datetime.now() - start_time).total_seconds()
        
        return {
            'new_markets_found': len(new_markets),
            'markets_created': len(created_markets),
            'schedules_created': schedules_created,
            'duration_seconds': duration,
            'new_markets': new_markets
        }


class EnhancedWeeklyImporter:
    """Enhanced weekly importer with language assignment integration."""
    
    def __init__(self, db_connection: DatabaseConnection):
        self.db = db_connection
        self.market_manager = WeeklyMarketSetupManager(db_connection)
        self.import_service = BroadcastMonthImportService(db_connection)
    
    def execute_enhanced_weekly_update(self, 
                                     excel_file: str,
                                     auto_setup_markets: bool = True,
                                     dry_run: bool = False) -> Dict:
        """
        Execute enhanced weekly update with language assignment integration.
        """
        start_time = datetime.now()
        batch_id = f"enhanced_weekly_{int(start_time.timestamp())}"
        
        # Clean header without excessive printing
        tqdm.write(f"🔄 Enhanced Weekly Update Starting")
        tqdm.write(f"📁 File: {Path(excel_file).name}")
        tqdm.write(f"🔧 Auto-setup: {auto_setup_markets} | Dry run: {dry_run}")
        tqdm.write(f"🆔 Batch ID: {batch_id}")
        tqdm.write("=" * 60)
        
        results = {
            'success': False,
            'batch_id': batch_id,
            'market_setup': None,
            'import_result': None,
            'language_assignment': None,
            'duration_seconds': 0,
            'error_messages': []
        }
        
        try:
            # Step 1: Market setup (if enabled)
            if auto_setup_markets and not dry_run:
                tqdm.write(f"🏗️  STEP 1: Automatic Market Setup")
                market_setup_result = self.market_manager.execute_weekly_market_setup(excel_file)
                results['market_setup'] = market_setup_result
                
                # Concise summary
                if market_setup_result['new_markets_found'] > 0:
                    tqdm.write(f"📊 Setup: {market_setup_result['markets_created']} markets, {market_setup_result['schedules_created']} assignments")
                else:
                    tqdm.write(f"📊 Setup: No new markets needed")
                tqdm.write("")
            
            # Step 2: Weekly data import with progress tracking
            tqdm.write(f"📦 STEP 2: Weekly Data Import")
            
            if dry_run:
                tqdm.write(f"🔍 DRY RUN - No changes would be made")
                summary = get_excel_import_summary(excel_file, self.db.db_path)
                results['import_result'] = {
                    'would_import': True,
                    'months_found': len(summary['months_in_excel']),
                    'total_spots': summary['total_existing_spots_affected']
                }
                tqdm.write(f"📊 Would import {summary['total_existing_spots_affected']:,} spots across {len(summary['months_in_excel'])} months")
            else:
                # Execute actual import with custom progress tracking
                import_result = self._execute_import_with_progress(excel_file)
                results['import_result'] = import_result
                
                # Step 3: Language Assignment Processing
                if import_result.success:
                    tqdm.write(f"\n🎯 STEP 3: Language Assignment Processing")
                    language_result = self._process_language_assignments(batch_id)
                    results['language_assignment'] = language_result
            
            results['success'] = True
            
        except Exception as e:
            error_msg = f"Enhanced weekly update failed: {str(e)}"
            results['error_messages'].append(error_msg)
            tqdm.write(f"❌ {error_msg}")
        
        # Calculate total duration
        results['duration_seconds'] = (datetime.now() - start_time).total_seconds()
        
        return results
    
    def _execute_import_with_progress(self, excel_file: str):
        """Execute import with progress tracking."""
        # Get summary first for progress setup
        summary = get_excel_import_summary(excel_file, self.db.db_path)
        total_spots = summary['total_existing_spots_affected']
        
        # Create a progress bar for the overall import
        with tqdm(total=100, desc="📦 Importing data", unit="%") as pbar:
            # Simulate import progress (since the actual import service may not have granular progress)
            # We'll update at key milestones
            
            pbar.set_description("📦 Preparing import")
            pbar.update(10)
            
            pbar.set_description("📦 Deleting existing data")
            pbar.update(20)
            
            # Execute actual import
            import_result = self.import_service.execute_month_replacement(
                excel_file,
                'WEEKLY_UPDATE',
                closed_by=None,
                dry_run=False
            )
            
            pbar.set_description("📦 Importing new data")
            pbar.update(50)
            
            pbar.set_description("📦 Finalizing import")
            pbar.update(20)
            
            pbar.set_description("📦 Import complete")
            pbar.update(100)
        
        # Clean summary
        if import_result.success:
            net_change = import_result.records_imported - import_result.records_deleted
            tqdm.write(f"📊 Import: {import_result.records_imported:,} imported, {import_result.records_deleted:,} deleted (net: {net_change:+,})")
        
        return import_result
    
    def _process_language_assignments(self, batch_id: str) -> Dict:
        """
        Process language assignments after import with comprehensive progress tracking.
        """
        language_result = {
            'success': False,
            'categorized': 0,
            'processed': 0,
            'language_assigned': 0,
            'default_english_assigned': 0,
            'flagged_for_review': 0,
            'error_messages': []
        }
        
        try:
            # Import language assignment services
            from services.spot_categorization_service import SpotCategorizationService
            from services.language_processing_orchestrator import LanguageProcessingOrchestrator
            
            conn = sqlite3.connect(self.db.db_path)
            
            # Step 3a: Categorization with progress
            categorization_service = SpotCategorizationService(conn)
            
            # Get uncategorized spots from current batch only
            batch_spots_query = """
                SELECT spot_id FROM spots 
                WHERE import_batch_id = ? AND spot_category IS NULL
            """
            cursor = conn.execute(batch_spots_query, (batch_id,))
            uncategorized_spots = [row[0] for row in cursor.fetchall()]
            
            if uncategorized_spots:
                # Categorize in batches with progress tracking
                batch_size = 1000
                total_categorized = 0
                
                with tqdm(total=len(uncategorized_spots), desc="🏷️  Categorizing spots", unit=" spots") as pbar:
                    for i in range(0, len(uncategorized_spots), batch_size):
                        batch = uncategorized_spots[i:i + batch_size]
                        categorization_service.categorize_spots_batch(batch)
                        total_categorized += len(batch)
                        pbar.update(len(batch))
                        pbar.set_description(f"🏷️  Categorized {total_categorized:,}/{len(uncategorized_spots):,}")
                
                language_result['categorized'] = len(uncategorized_spots)
                tqdm.write(f"✅ Categorized {len(uncategorized_spots):,} spots")
            else:
                tqdm.write(f"✅ No uncategorized spots found")
            
            # Step 3b: Process all categories with progress
            orchestrator = LanguageProcessingOrchestrator(conn)
            
            # Get processing status for progress tracking
            processing_results = orchestrator.process_batch_categories(batch_id)
            
            # Extract detailed results from processing
            summary = processing_results['summary']
            language_result['processed'] = summary['total_processed']
            language_result['language_assigned'] = summary['language_assigned']
            language_result['default_english_assigned'] = summary['default_english_assigned']
            language_result['flagged_for_review'] = summary['flagged_for_review']
            language_result['success'] = True
            
            # Clean summary
            tqdm.write(f"✅ Language assignment complete:")
            tqdm.write(f"   🎯 Processed: {language_result['processed']:,}")
            tqdm.write(f"   🔤 Language assigned: {language_result['language_assigned']:,}")
            tqdm.write(f"   🇺🇸 Default English: {language_result['default_english_assigned']:,}")
            if language_result['flagged_for_review'] > 0:
                tqdm.write(f"   📋 Review required: {language_result['flagged_for_review']:,}")
            
            conn.close()
            
        except Exception as e:
            error_msg = f"Language assignment processing failed: {str(e)}"
            language_result['error_messages'].append(error_msg)
            tqdm.write(f"⚠️  {error_msg}")
            
            # Try to provide partial success info
            try:
                if 'conn' in locals():
                    conn.close()
            except:
                pass
        
        return language_result


def display_enhanced_weekly_preview(excel_file: str, db_path: str, auto_setup: bool):
    """Display what the enhanced weekly update would do."""
    tqdm.write(f"📋 Enhanced Weekly Update Preview")
    tqdm.write(f"=" * 60)
    tqdm.write(f"📁 File: {Path(excel_file).name}")
    tqdm.write(f"🔧 Auto-setup: {auto_setup}")
    tqdm.write("")
    
    try:
        db_connection = DatabaseConnection(db_path)
        
        # Market setup preview
        if auto_setup:
            market_manager = WeeklyMarketSetupManager(db_connection)
            new_markets = market_manager.scan_excel_for_new_markets(excel_file)
            
            if new_markets:
                tqdm.write(f"🏗️  Market Setup Preview:")
                tqdm.write(f"   🆕 New markets: {len(new_markets)} ({', '.join(sorted(new_markets.keys()))})")
                tqdm.write("")
        
        # Standard import preview
        can_proceed = display_weekly_update_preview(excel_file, db_path)
        
        # Language assignment preview
        if can_proceed:
            tqdm.write(f"🎯 Language Assignment Preview:")
            tqdm.write(f"   📋 All spots will be categorized and processed automatically")
            tqdm.write(f"   🔤 Business rules applied, manual review flagged as needed")
        
        db_connection.close()
        return can_proceed
        
    except Exception as e:
        tqdm.write(f"❌ Error generating preview: {e}")
        return False


def display_weekly_update_preview(excel_file: str, db_path: str):
    """Display what the weekly update would do (original functionality)."""
    tqdm.write(f"📦 Weekly Update Preview:")
    
    try:
        # Get Excel summary
        summary = get_excel_import_summary(excel_file, db_path)
        
        # Validation check
        validation = validate_excel_for_import(excel_file, 'WEEKLY_UPDATE', db_path)
        
        if validation.is_valid:
            tqdm.write(f"   ✅ Weekly update allowed")
            tqdm.write(f"   📊 {summary['total_existing_spots_affected']:,} spots across {len(summary['months_in_excel'])} months")
        else:
            tqdm.write(f"   ❌ Weekly update BLOCKED")
            tqdm.write(f"      Reason: {validation.error_message}")
            tqdm.write(f"      Solution: {validation.suggested_action}")
            return False
        
        # Show month details concisely
        open_months = summary['open_months']
        closed_months = summary['closed_months']
        
        tqdm.write(f"   📂 Open months: {len(open_months)} ({', '.join(open_months) if len(open_months) <= 5 else f'{len(open_months)} months'})")
        if closed_months:
            tqdm.write(f"   🔒 Closed months: {len(closed_months)} (protected)")
        tqdm.write("")
        
        return True
        
    except Exception as e:
        tqdm.write(f"❌ Error generating preview: {e}")
        return False


def get_user_confirmation(total_spots: int, open_months: int, new_markets: int = 0, force: bool = False) -> bool:
    """Get user confirmation for the enhanced update."""
    if force:
        return True
    
    tqdm.write(f"🚨 CONFIRMATION REQUIRED")
    actions = [
        f"REPLACE {total_spots:,} existing spots",
        f"Update {open_months} open months",
        "Process language assignments automatically",
        "Preserve all closed/historical months"
    ]
    
    if new_markets > 0:
        actions.insert(0, f"Create {new_markets} new markets")
    
    tqdm.write(f"This will:")
    for action in actions:
        tqdm.write(f"  • {action}")
    
    while True:
        response = input(f"\nProceed with enhanced weekly update? (yes/no): ").strip().lower()
        if response in ['yes', 'y']:
            return True
        elif response in ['no', 'n', '']:
            return False
        else:
            tqdm.write("Please enter 'yes' or 'no'")


def main():
    parser = argparse.ArgumentParser(
        description="Enhanced weekly update with automatic market setup and language assignment",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Standard weekly update with language assignment
  python enhanced_weekly_update.py data/booked_business_current.xlsx

  # Enhanced weekly update with automatic market setup and language assignment
  python enhanced_weekly_update.py data/booked_business_current.xlsx --auto-setup

  # Preview what would happen
  python enhanced_weekly_update.py data/weekly_data.xlsx --auto-setup --dry-run

Enhanced Features:
  • Automatic detection of new markets in weekly data
  • Missing market creation with proper naming
  • Schedule assignment setup for new markets
  • Integrated language assignment processing with business rules
  • Comprehensive progress tracking with tqdm progress bars
  • Clean, professional output with minimal screen flooding
  • Backward compatible with existing workflows
        """
    )
    
    parser.add_argument("excel_file", help="Path to Excel file to import")
    parser.add_argument("--auto-setup", action="store_true",
                       help="Automatically create missing markets and schedule assignments")
    parser.add_argument("--db-path", default="data/database/production.db",
                       help="Database path (default: production.db)")
    parser.add_argument("--dry-run", action="store_true",
                       help="Preview import without making changes")
    parser.add_argument("--force", action="store_true",
                       help="Skip confirmation prompts")
    parser.add_argument("--verbose", action="store_true",
                       help="Enable verbose logging")
    
    args = parser.parse_args()
    
    # Setup logging
    import logging
    level = logging.DEBUG if args.verbose else logging.WARNING  # Reduced default logging
    logging.basicConfig(level=level, format='%(levelname)s: %(message)s')
    
    # Validate inputs
    if not Path(args.excel_file).exists():
        print(f"❌ Excel file not found: {args.excel_file}")
        sys.exit(1)
    
    if not Path(args.db_path).exists():
        print(f"❌ Database not found: {args.db_path}")
        print(f"Run: uv run python scripts/setup_database.py --db-path {args.db_path}")
        sys.exit(1)
    
    try:
        print(f"🔄 Enhanced Weekly Update Tool")
        print(f"📁 File: {Path(args.excel_file).name}")
        print(f"🗃️  Database: {Path(args.db_path).name}")
        if args.dry_run:
            print(f"🔍 Mode: DRY RUN (no changes will be made)")
        print()
        
        # Display enhanced preview and validate
        can_proceed = display_enhanced_weekly_preview(args.excel_file, args.db_path, args.auto_setup)
        
        if not can_proceed:
            print(f"\n❌ Weekly update cannot proceed due to validation errors")
            print(f"💡 Common solutions:")
            print(f"  • Remove closed month data from Excel file")
            print(f"  • Use historical import mode for closed months")
            print(f"  • Check if months need to be manually closed first")
            sys.exit(1)
        
        # Get confirmation unless forced or dry run
        if not args.dry_run and not args.force:
            summary = get_excel_import_summary(args.excel_file, args.db_path)
            new_market_count = 0
            
            if args.auto_setup:
                # Estimate new markets
                db_connection = DatabaseConnection(args.db_path)
                market_manager = WeeklyMarketSetupManager(db_connection)
                new_markets = market_manager.scan_excel_for_new_markets(args.excel_file)
                new_market_count = len(new_markets)
                db_connection.close()
            
            confirmed = get_user_confirmation(
                summary['total_existing_spots_affected'],
                len(summary['open_months']),
                new_market_count,
                args.force
            )
            
            if not confirmed:
                print(f"❌ Weekly update cancelled by user")
                sys.exit(0)
        
        # Execute the enhanced update
        db_connection = DatabaseConnection(args.db_path)
        importer = EnhancedWeeklyImporter(db_connection)
        
        try:
            results = importer.execute_enhanced_weekly_update(
                args.excel_file,
                args.auto_setup,
                args.dry_run
            )
            
            # Display clean final results
            print(f"\n{'='*60}")
            if args.dry_run:
                print(f"🔍 ENHANCED WEEKLY UPDATE DRY RUN COMPLETED")
            else:
                print(f"🎉 ENHANCED WEEKLY UPDATE COMPLETED")
            print(f"{'='*60}")
            
            print(f"📊 Results Summary:")
            print(f"  Status: {'✅ Success' if results['success'] else '❌ Failed'}")
            print(f"  Duration: {results['duration_seconds']:.1f} seconds")
            
            # Market setup results (concise)
            if results['market_setup'] and results['market_setup']['markets_created'] > 0:
                setup = results['market_setup']
                print(f"  Markets created: {setup['markets_created']}")
            
            # Import results (concise)
            if results['import_result'] and not args.dry_run:
                import_res = results['import_result']
                if hasattr(import_res, 'success') and import_res.success:
                    net_change = import_res.records_imported - import_res.records_deleted
                    print(f"  Spots imported: {import_res.records_imported:,}")
                    print(f"  Net change: {net_change:+,}")
            
            # Language assignment results (concise)
            if results['language_assignment'] and results['language_assignment']['success']:
                lang_res = results['language_assignment']
                print(f"  Language processed: {lang_res['processed']:,}")
                if lang_res['flagged_for_review'] > 0:
                    print(f"  Review needed: {lang_res['flagged_for_review']:,}")
            
            if results['error_messages']:
                print(f"\n❌ Errors:")
                for error in results['error_messages']:
                    print(f"  • {error}")
                sys.exit(1)
            
            if results['success'] and not args.dry_run:
                print(f"\n✅ Update completed successfully!")
                if results['language_assignment'] and results['language_assignment']['flagged_for_review'] > 0:
                    print(f"💡 Next: Review flagged spots with 'uv run python cli/assign_languages.py --review-required'")
                
        finally:
            db_connection.close()
    
    except BroadcastMonthImportError as e:
        print(f"❌ Import error: {e}")
        sys.exit(1)
    except KeyboardInterrupt:
        print(f"\n❌ Weekly update cancelled by user")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Unexpected error: {e}")
        if args.verbose:
            import traceback
            traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()