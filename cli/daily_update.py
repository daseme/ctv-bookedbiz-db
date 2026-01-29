#!/usr/bin/env python3
"""
Enhanced Daily Update Command with Multi-Sheet Support
Clean Architecture implementation optimized for automated multi-sheet processing.
Enhanced with unattended mode for automated daily processing of combined Commercial + WorldLink data.
"""

from __future__ import annotations
import sys
import os
import argparse
import sqlite3
import logging
from pathlib import Path
from datetime import datetime, date
from typing import Dict, Set, List, Optional, Any, Protocol, Tuple
from dataclasses import dataclass, field
from enum import Enum
import logging.handlers

# FIXED PATH SETUP - More robust approach
script_dir = Path(__file__).parent.absolute()
project_root = script_dir.parent
src_dir = project_root / "src"

# Add multiple possible paths
sys.path.insert(0, str(src_dir))
sys.path.insert(0, str(project_root))

# Debug path setup (remove this after confirming it works)
print(f"Script dir: {script_dir}")
print(f"Project root: {project_root}")
print(f"Src dir: {src_dir}")
print(f"Src dir exists: {src_dir.exists()}")

# Add tqdm for progress bars
try:
    from tqdm import tqdm
except ImportError:
    print("Warning: tqdm not available, progress bars will be limited")

    # Create a dummy tqdm class
    class tqdm:
        def __init__(self, *args, **kwargs):
            self.total = kwargs.get("total", 100)
            self.current = 0

        def update(self, n):
            self.current += n

        def set_description(self, desc):
            pass

        def __enter__(self):
            return self

        def __exit__(self, *args):
            pass

        @staticmethod
        def write(msg):
            print(msg)


# Try imports with better error messages
try:
    from src.services.broadcast_month_import_service import (
        BroadcastMonthImportService,
        BroadcastMonthImportError,
    )

    print("✅ Successfully imported BroadcastMonthImportService")
except ImportError as e:
    print(f"❌ Failed to import BroadcastMonthImportService: {e}")
    print(f"   Make sure the src directory exists at: {src_dir}")
    sys.exit(1)

try:
    from src.services.import_integration_utilities import (
        get_excel_import_summary,
        validate_excel_for_import,
    )

    print("✅ Successfully imported import utilities")
except ImportError as e:
    print(f"❌ Failed to import import utilities: {e}")
    sys.exit(1)

try:
    from src.database.connection import DatabaseConnection

    print("✅ Successfully imported DatabaseConnection")
except ImportError as e:
    print(f"❌ Failed to import DatabaseConnection: {e}")
    sys.exit(1)

# ============================================================================
# Domain Models (Enhanced for Multi-Sheet)
# ============================================================================


class BatchType(Enum):
    ENHANCED_DAILY = ("enhanced_daily", "Enhanced Daily Update (Multi-Sheet)")
    HISTORICAL_IMPORT = ("historical_import", "Historical Import")
    MANUAL_UPDATE = ("manual_update", "Manual Update")

    def __init__(self, code: str, display_name: str):
        self.code = code
        self.display_name = display_name


@dataclass
class MultiSheetPreview:
    """Preview information for multi-sheet imports"""

    total_spots: int
    sheet_breakdown: Dict[str, int]
    months_affected: List[str]
    new_markets_detected: int

    @property
    def has_worldlink_data(self) -> bool:
        """Check if WorldLink data is present"""
        return any(
            "worldlink" in sheet.lower() for sheet in self.sheet_breakdown.keys()
        )

    @property
    def summary_line(self) -> str:
        """One-line summary of the preview"""
        sources = []
        for sheet, count in self.sheet_breakdown.items():
            sources.append(f"{sheet}: {count:,}")
        return f"{self.total_spots:,} spots ({', '.join(sources)})"


@dataclass
class MarketData:
    market_code: str
    earliest_date: Optional[date]
    latest_date: Optional[date]
    spot_count: int = 0

    @classmethod
    def create_new(cls, market_code: str) -> MarketData:
        return cls(
            market_code=market_code, earliest_date=None, latest_date=None, spot_count=0
        )

    def add_spot_date(self, air_date: Optional[date]) -> None:
        """Add a spot date to this market data"""
        self.spot_count += 1
        if air_date:
            if not self.earliest_date or air_date < self.earliest_date:
                self.earliest_date = air_date
            if not self.latest_date or air_date > self.latest_date:
                self.latest_date = air_date


@dataclass
class DailyUpdateConfig:
    excel_file: Path
    auto_setup_markets: bool = True
    dry_run: bool = False
    force: bool = False
    verbose: bool = False
    unattended: bool = False
    log_file: Optional[Path] = None
    db_path: Path = Path("data/database/production.db")

    @classmethod
    def from_args(cls, args) -> DailyUpdateConfig:
        """Create configuration from CLI arguments"""
        return cls(
            excel_file=Path(args.excel_file),
            auto_setup_markets=args.auto_setup,
            dry_run=args.dry_run,
            force=args.force or args.unattended,  # Unattended implies force
            verbose=args.verbose,
            unattended=args.unattended,
            log_file=Path(args.log_file) if args.log_file else None,
            db_path=Path(args.db_path),
        )


@dataclass
class MarketSetupResult:
    new_markets_found: int
    markets_created: int
    schedules_created: int
    duration_seconds: float
    new_markets: Dict[str, MarketData] = field(default_factory=dict)

    @property
    def summary(self) -> str:
        """Human readable summary"""
        if self.new_markets_found == 0:
            return "No new markets needed"
        return f"{self.markets_created} markets, {self.schedules_created} assignments created"


@dataclass
class ImportResult:
    success: bool
    records_imported: int
    records_deleted: int
    batch_id: str
    duration_seconds: float
    months_processed: List[str] = field(default_factory=list)
    error_message: Optional[str] = None
    sheet_breakdown: Dict[str, int] = field(
        default_factory=dict
    )  # NEW: Sheet source tracking

    @property
    def net_change(self) -> int:
        """Calculate net change in records"""
        return self.records_imported - self.records_deleted

    @property
    def has_multisheet_data(self) -> bool:
        """Check if import included multiple sheet sources"""
        return len(self.sheet_breakdown) > 1


@dataclass
class LanguageAssignmentResult:
    success: bool
    categorized: int
    processed: int
    language_assigned: int
    default_english_assigned: int
    flagged_for_review: int
    duration_seconds: float = 0.0
    error_messages: List[str] = field(default_factory=list)

    @property
    def summary(self) -> str:
        """Human readable summary"""
        return f"Processed: {self.processed:,}, Language assigned: {self.language_assigned:,}, Review needed: {self.flagged_for_review:,}"


@dataclass
class DailyUpdateResult:
    success: bool
    batch_id: str
    market_setup: Optional[MarketSetupResult] = None
    import_result: Optional[ImportResult] = None
    language_assignment: Optional[LanguageAssignmentResult] = None
    duration_seconds: float = 0.0
    error_messages: List[str] = field(default_factory=list)

    @property
    def summary_line(self) -> str:
        """One-line summary for logging"""
        status = "SUCCESS" if self.success else "FAILED"
        duration = f"{self.duration_seconds:.1f}s"
        multisheet = ""
        if self.import_result and self.import_result.has_multisheet_data:
            multisheet = " (Multi-sheet)"
        return f"{status}{multisheet} | Duration: {duration} | Batch: {self.batch_id}"


# ============================================================================
# Value Objects and Business Rules (Enhanced)
# ============================================================================


class MarketNameGenerator:
    """Value object for generating market names from codes"""

    NAME_MAPPINGS = {
        "NYC": "NEW YORK",
        "LAX": "LOS ANGELES",
        "SFO": "SAN FRANCISCO",
        "SEA": "SEATTLE",
        "CHI": "CHICAGO",
        "MSP": "MINNEAPOLIS",
        "DAL": "DALLAS",
        "HOU": "HOUSTON",
        "WDC": "WASHINGTON DC",
        "CVC": "CENTRAL VALLEY",
        "CMP": "CHI MSP",
        "MMT": "MAMMOTH",
        "ADMIN": "ADMINISTRATIVE",
    }

    @classmethod
    def generate_name(cls, market_code: str) -> str:
        """Generate proper market name from code"""
        return cls.NAME_MAPPINGS.get(market_code, market_code.upper().replace("_", " "))


class BatchIdGenerator:
    """Value object for generating standardized batch IDs"""

    @staticmethod
    def generate(batch_type: BatchType, timestamp: Optional[datetime] = None) -> str:
        """Generate standardized batch ID"""
        if timestamp is None:
            timestamp = datetime.now()
        return f"{batch_type.code}_{int(timestamp.timestamp())}"


class MultiSheetPreviewGenerator:
    """Service for generating enhanced previews of multi-sheet data"""

    @staticmethod
    def generate_preview(
        excel_file: Path, db_connection: DatabaseConnection
    ) -> MultiSheetPreview:
        """Generate preview information for multi-sheet Excel file"""
        try:
            import pandas as pd

            # FIXED: Read all required sheets instead of just "Commercials"
            total_spots = 0
            sheet_breakdown = {}

            for sheet_name in CommercialLogConfig.REQUIRED_SHEETS:
                try:
                    df = pd.read_excel(excel_file, sheet_name=sheet_name)
                    sheet_breakdown[sheet_name] = len(df)
                    total_spots += len(df)
                except Exception:
                    # Sheet doesn't exist, skip it
                    continue

            # Get months from the import utility
            from src.services.import_integration_utilities import (
                get_excel_import_summary,
            )

            summary = get_excel_import_summary(str(excel_file), db_connection.db_path)

            return MultiSheetPreview(
                total_spots=total_spots,
                sheet_breakdown=sheet_breakdown,
                months_affected=summary.get("months_in_excel", []),
                new_markets_detected=0,  # Could enhance this later
            )

        except Exception as e:
            # Return empty preview on error
            return MultiSheetPreview(
                total_spots=0,
                sheet_breakdown={},
                months_affected=[],
                new_markets_detected=0,
            )


# ============================================================================
# Data Access Layer (Unchanged but Enhanced Comments)
# ============================================================================


class MarketRepository:
    """Repository for market-related database operations"""

    def __init__(self, db_connection: DatabaseConnection):
        self.db = db_connection

    def get_existing_markets(self) -> Dict[str, int]:
        """Get existing markets from database"""
        with self.db.transaction() as conn:
            cursor = conn.execute("SELECT market_code, market_id FROM markets")
            return {row[0]: row[1] for row in cursor.fetchall()}

    def create_market(self, market_code: str, market_name: str) -> int:
        """Create new market and return its ID"""
        with self.db.transaction() as conn:
            cursor = conn.execute(
                "INSERT INTO markets (market_code, market_name) VALUES (?, ?)",
                (market_code, market_name),
            )
            return cursor.lastrowid

    def create_schedule_assignment(self, market_id: int, effective_date: date) -> None:
        """Create schedule assignment for market"""
        with self.db.transaction() as conn:
            conn.execute(
                """
                INSERT INTO schedule_market_assignments 
                (schedule_id, market_id, effective_start_date, assignment_priority)
                VALUES (1, ?, ?, 1)
            """,
                (market_id, effective_date),
            )


class SpotRepository:
    """Repository for spot-related database operations"""

    def __init__(self, db_connection: DatabaseConnection):
        self.db = db_connection

    def get_uncategorized_spots_by_batch(self, batch_id: str) -> List[int]:
        """Get uncategorized spot IDs for a specific batch"""
        with self.db.transaction() as conn:
            cursor = conn.execute(
                """
                SELECT spot_id FROM spots 
                WHERE import_batch_id = ? AND spot_category IS NULL
            """,
                (batch_id,),
            )
            return [row[0] for row in cursor.fetchall()]

    def get_sheet_source_breakdown(self, batch_id: str) -> Dict[str, int]:
        """Get breakdown of spots by sheet source for a batch"""
        with self.db.transaction() as conn:
            cursor = conn.execute(
                """
                SELECT 
                    CASE 
                        WHEN source_file LIKE '%:Worldlink Lines' THEN 'Worldlink Lines'
                        WHEN source_file LIKE '%:Commercials' THEN 'Commercials'  
                        ELSE 'Other'
                    END as sheet_source,
                    COUNT(*) as count
                FROM spots 
                WHERE import_batch_id = ?
                GROUP BY 1
                ORDER BY count DESC
            """,
                (batch_id,),
            )
            return {row[0]: row[1] for row in cursor.fetchall()}


# ============================================================================
# Progress Reporting Adapters (Enhanced)
# ============================================================================


class ProgressReporter(Protocol):
    """Interface for progress reporting - can be implemented for CLI, web, logging"""

    def create_progress(self, description: str, total: int) -> ProgressContext:
        """Create a progress tracking context"""
        ...

    def write(self, message: str) -> None:
        """Write a message to output"""
        ...


class ProgressContext(Protocol):
    """Context manager for progress tracking"""

    def update(self, increment: int) -> None:
        """Update progress by increment"""
        ...

    def set_description(self, description: str) -> None:
        """Update progress description"""
        ...


class TqdmProgressReporter:
    """Tqdm implementation of progress reporting for interactive use"""

    def create_progress(self, description: str, total: int):
        """Create tqdm progress bar context"""
        return tqdm(total=total, desc=description, unit=" items")

    def write(self, message: str) -> None:
        """Write message using tqdm.write to avoid conflicts with progress bars"""
        tqdm.write(message)


class LoggingProgressReporter:
    """Enhanced logging implementation for unattended operation"""

    def __init__(self, logger: logging.Logger):
        self.logger = logger

    def create_progress(self, description: str, total: int):
        """Create logging-based progress context"""
        return LoggingProgressContext(self.logger, description, total)

    def write(self, message: str) -> None:
        """Write message to logger with enhanced formatting"""
        clean_message = self._clean_message_for_logs(message)

        # Determine log level based on message content
        if "ERROR" in message or "FAILED" in message:
            self.logger.error(clean_message)
        elif "WARNING" in message or "WARN" in message:
            self.logger.warning(clean_message)
        else:
            self.logger.info(clean_message)

    def _clean_message_for_logs(self, message: str) -> str:
        """Enhanced message cleaning for multi-sheet logs"""
        # Remove common emojis and replace with meaningful tags
        emoji_map = {
            "🔄": "[UPDATE]",
            "📁": "[FILE]",
            "🔧": "[CONFIG]",
            "🆔": "[BATCH]",
            "🏗️": "[SETUP]",
            "✅": "[SUCCESS]",
            "❌": "[ERROR]",
            "⚠️": "[WARNING]",
            "📦": "[IMPORT]",
            "🎯": "[LANGUAGE]",
            "🔍": "[SCAN]",
            "🗓️": "[SCHEDULE]",
            "📊": "[STATS]",
            "🎉": "[COMPLETE]",
            "💡": "[INFO]",
            "🚨": "[CONFIRM]",
            "📋": "[PREVIEW]",
            "🏷️": "[CATEGORIZE]",
            "🔤": "[PROCESS]",
            # NEW: Multi-sheet specific
            "📄": "[SHEET]",
            "🔗": "[COMBINE]",
            "📈": "[BREAKDOWN]",
        }

        clean_msg = message
        for emoji, replacement in emoji_map.items():
            clean_msg = clean_msg.replace(emoji, replacement)

        # Enhance multi-sheet specific messages
        if "Commercials:" in clean_msg and "Worldlink Lines:" in clean_msg:
            clean_msg = "[MULTI-SHEET] " + clean_msg

        return clean_msg


class LoggingProgressContext:
    """Enhanced progress context with multi-sheet awareness"""

    def __init__(self, logger: logging.Logger, description: str, total: int):
        self.logger = logger
        self.description = description
        self.total = total
        self.current = 0
        self.last_logged_percent = 0

        # Enhanced start logging
        clean_desc = self._clean_message(description)
        self.logger.info(f"{clean_desc} - Starting (0/{total:,})")

    def update(self, increment: int) -> None:
        """Update progress and log at milestones"""
        self.current += increment
        percent = int((self.current / self.total) * 100) if self.total > 0 else 100

        # Log at 25%, 50%, 75%, 100% milestones + every 1000 for large imports
        milestone_hit = (
            percent in [25, 50, 75, 100] and percent > self.last_logged_percent
        )
        large_batch_milestone = self.total > 5000 and self.current % 1000 == 0

        if milestone_hit or large_batch_milestone:
            clean_desc = self._clean_message(self.description)
            self.logger.info(
                f"{clean_desc} - Progress {percent}% ({self.current:,}/{self.total:,})"
            )
            if milestone_hit:
                self.last_logged_percent = percent

    def set_description(self, description: str) -> None:
        """Update description"""
        self.description = description

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type is None:
            clean_desc = self._clean_message(self.description)
            self.logger.info(
                f"{clean_desc} - Completed ({self.current:,}/{self.total:,})"
            )

    def _clean_message(self, message: str) -> str:
        """Remove emojis for clean logging"""
        return message.encode("ascii", "ignore").decode("ascii").strip()


# ============================================================================
# Business Logic Layer (Enhanced for Multi-Sheet)
# ============================================================================


class ExcelMarketScanner:
    """Enhanced service for scanning Excel files for new markets"""

    def __init__(self, progress_reporter: ProgressReporter):
        self.progress_reporter = progress_reporter

    def scan_for_new_markets(
        self, excel_file: Path, existing_markets: Set[str]
    ) -> Dict[str, MarketData]:
        """
        Scan Excel file to detect any new market codes not in existing set.
        Enhanced to work with multi-sheet combined files.
        """
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(excel_file, read_only=True, data_only=True)
            worksheet = workbook.active

            # Find required columns
            market_col_index, air_date_col_index = self._find_columns(worksheet)
            if market_col_index is None:
                return {}

            new_markets = {}
            total_rows = worksheet.max_row - 1  # Exclude header

            # Process rows to find new markets with progress bar
            with self.progress_reporter.create_progress(
                "Scanning for new markets", total_rows
            ) as pbar:
                for row_num, row in enumerate(
                    worksheet.iter_rows(min_row=2, values_only=True), start=2
                ):
                    pbar.update(1)

                    if not any(cell for cell in row):
                        continue

                    market_code = self._extract_market_code(row, market_col_index)
                    if not market_code or market_code in existing_markets:
                        continue

                    # New market found - create or update data
                    if market_code not in new_markets:
                        new_markets[market_code] = MarketData.create_new(market_code)

                    air_date = self._extract_air_date(row, air_date_col_index)
                    new_markets[market_code].add_spot_date(air_date)

                    # Update progress description with found markets
                    if len(new_markets) > 0:
                        pbar.set_description(
                            f"Scanning ({len(new_markets)} new markets found)"
                        )

            workbook.close()

            if new_markets:
                total_spots = sum(data.spot_count for data in new_markets.values())
                self.progress_reporter.write(
                    f"Found {len(new_markets)} new markets with {total_spots:,} spots"
                )

            return new_markets

        except Exception as e:
            self.progress_reporter.write(
                f"Warning: Could not scan for new markets: {str(e)}"
            )
            return {}

    def _find_columns(self, worksheet) -> Tuple[Optional[int], Optional[int]]:
        """Find market and date column indices"""
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        market_col_index = None
        air_date_col_index = None

        for i, header in enumerate(header_row):
            if header:
                header_str = str(header).strip().lower()
                if header_str in ["market_name", "market", "market_code"]:
                    market_col_index = i
                elif header_str in ["air_date", "date", "airdate", "start date"]:
                    air_date_col_index = i

        return market_col_index, air_date_col_index

    def _extract_market_code(self, row: tuple, market_col: int) -> Optional[str]:
        """Extract and clean market code from row"""
        if market_col < len(row) and row[market_col]:
            return str(row[market_col]).strip()
        return None

    def _extract_air_date(self, row: tuple, date_col: Optional[int]) -> Optional[date]:
        """Extract air date from row if available"""
        if date_col is None or date_col >= len(row):
            return None

        air_date_value = row[date_col]
        if not air_date_value:
            return None

        try:
            if isinstance(air_date_value, datetime):
                return air_date_value.date()
            else:
                return datetime.strptime(str(air_date_value), "%Y-%m-%d").date()
        except:
            return None


class MarketSetupService:
    """Service for setting up new markets and schedules"""

    def __init__(
        self, market_repository: MarketRepository, progress_reporter: ProgressReporter
    ):
        self.market_repository = market_repository
        self.progress_reporter = progress_reporter

    def execute_daily_market_setup(self, excel_file: Path) -> MarketSetupResult:
        """Execute lightweight market setup for daily data"""
        start_time = datetime.now()

        # Step 1: Scan for new markets
        existing_markets = set(self.market_repository.get_existing_markets().keys())
        scanner = ExcelMarketScanner(self.progress_reporter)
        new_markets = scanner.scan_for_new_markets(excel_file, existing_markets)

        if not new_markets:
            return MarketSetupResult(
                new_markets_found=0,
                markets_created=0,
                schedules_created=0,
                duration_seconds=(datetime.now() - start_time).total_seconds(),
            )

        # Step 2: Create new markets and schedules
        created_markets = self._create_markets(new_markets)
        schedules_created = self._create_schedule_assignments(
            new_markets, created_markets
        )

        duration = (datetime.now() - start_time).total_seconds()

        return MarketSetupResult(
            new_markets_found=len(new_markets),
            markets_created=len(created_markets),
            schedules_created=schedules_created,
            duration_seconds=duration,
            new_markets=new_markets,
        )

    def _create_markets(self, new_markets: Dict[str, MarketData]) -> Dict[str, int]:
        """Create new markets with progress tracking"""
        created_markets = {}

        with self.progress_reporter.create_progress(
            "Creating markets", len(new_markets)
        ) as pbar:
            for market_code, market_data in sorted(new_markets.items()):
                market_name = MarketNameGenerator.generate_name(market_code)
                market_id = self.market_repository.create_market(
                    market_code, market_name
                )
                created_markets[market_code] = market_id

                pbar.update(1)
                pbar.set_description(f"Created {market_code}")

        self.progress_reporter.write(f"Created {len(created_markets)} new markets")
        return created_markets

    def _create_schedule_assignments(
        self, new_markets: Dict[str, MarketData], market_mapping: Dict[str, int]
    ) -> int:
        """Setup schedule assignments for newly created markets"""
        assignments_created = 0

        with self.progress_reporter.create_progress(
            "Setting up schedules", len(new_markets)
        ) as pbar:
            for market_code, market_data in new_markets.items():
                market_id = market_mapping[market_code]

                # Use earliest date if available, otherwise use current date
                effective_date = (
                    market_data.earliest_date
                    if market_data.earliest_date
                    else datetime.now().date()
                )

                self.market_repository.create_schedule_assignment(
                    market_id, effective_date
                )
                assignments_created += 1

                pbar.update(1)
                pbar.set_description(f"Setup {market_code}")

        self.progress_reporter.write(
            f"Created {assignments_created} schedule assignments"
        )
        return assignments_created


class ImportService:
    """Enhanced service for handling multi-sheet data imports with normalization repair"""

    def __init__(
        self,
        broadcast_import_service: BroadcastMonthImportService,
        spot_repository: SpotRepository,
        progress_reporter: ProgressReporter,
    ):
        self.broadcast_service = broadcast_import_service
        self.spot_repository = spot_repository
        self.progress_reporter = progress_reporter

    def ensure_bill_codes_in_raw_inputs(self, excel_file: Path) -> None:
        """
        Ensure all bill_code values from Excel are added to raw_customer_inputs
        This prevents normalization gaps from occurring.
        """
        try:
            import pandas as pd
            
            # Try to read the main data sheet - handle multiple possible sheet names
            sheet_names_to_try = ["Commercials", "Commercial Lines", "Sheet1", 0]  # 0 = first sheet
            df = None
            sheet_used = None
            
            for sheet_name in sheet_names_to_try:
                try:
                    df = pd.read_excel(excel_file, sheet_name=sheet_name)
                    sheet_used = sheet_name
                    self.progress_reporter.write(f"📄 Reading bill codes from sheet: {sheet_name}")
                    break
                except Exception:
                    continue
            
            if df is None:
                self.progress_reporter.write("⚠️ Warning: Could not read any sheet from Excel file")
                return

            # Get unique bill codes - try multiple possible column names
            bill_code_columns = ['bill_code', 'Bill Code', 'Customer', 'Client', 'Advertiser', 'customer']
            bill_codes = []
            column_used = None
            
            for col in bill_code_columns:
                if col in df.columns:
                    bill_codes = df[col].dropna().unique().tolist()
                    column_used = col
                    break
            
            if not bill_codes:
                self.progress_reporter.write("⚠️ Warning: No bill code column found in Excel file")
                available_columns = list(df.columns)[:10]  # Show first 10 columns
                self.progress_reporter.write(f"Available columns: {available_columns}")
                return
            
            # Clean and filter bill codes
            clean_bill_codes = []
            for bill_code in bill_codes:
                if bill_code and str(bill_code).strip() and str(bill_code).strip().upper() != 'NAN':
                    clean_bill_codes.append(str(bill_code).strip())
            
            if clean_bill_codes:
                # Add to raw_customer_inputs
                current_time = datetime.now().isoformat()
                added_count = 0
                
                with self.broadcast_service.db_connection.transaction() as conn:
                    for bill_code in clean_bill_codes:
                        try:
                            cursor = conn.execute("""
                                INSERT OR IGNORE INTO raw_customer_inputs (raw_text, created_at)
                                VALUES (?, ?)
                            """, (bill_code, current_time))
                            
                            if cursor.rowcount > 0:
                                added_count += 1
                                
                        except Exception as e:
                            # Continue with other bill codes if one fails
                            continue
                
                self.progress_reporter.write(f"✅ Normalization system updated:")
                self.progress_reporter.write(f"   Sheet: {sheet_used}, Column: {column_used}")
                self.progress_reporter.write(f"   Total bill codes found: {len(clean_bill_codes)}")
                self.progress_reporter.write(f"   New bill codes added: {added_count}")
                
                if added_count == 0:
                    self.progress_reporter.write(f"   (All bill codes were already in system)")
            else:
                self.progress_reporter.write("⚠️ Warning: No valid bill codes found after cleaning")
        
        except Exception as e:
            self.progress_reporter.write(f"⚠️ Warning: Could not update raw_customer_inputs: {e}")
            # Don't fail the entire import if this step fails

    def execute_import_with_progress(
        self, excel_file: Path, batch_id: str
    ) -> ImportResult:
        """Execute import with enhanced multi-sheet progress tracking + normalization repair"""
        start_time = datetime.now()

        # CRITICAL FIX: Ensure bill codes are in raw_customer_inputs BEFORE import
        self.progress_reporter.write("🔧 Step 1: Updating normalization system with new bill codes...")
        self.ensure_bill_codes_in_raw_inputs(excel_file)

        # Get summary first for progress setup
        self.progress_reporter.write("🔍 Step 2: Analyzing Excel file for import...")
        try:
            summary = get_excel_import_summary(
                str(excel_file), self.broadcast_service.db_connection.db_path
            )
        except Exception as e:
            self.progress_reporter.write(f"⚠️ Warning: Could not get import summary: {e}")
            summary = {"months_in_excel": [], "total_existing_spots_affected": 0}

        # Create a progress bar for the overall import
        self.progress_reporter.write("📦 Step 3: Executing data import...")
        with self.progress_reporter.create_progress(
            "Importing multi-sheet data", 100
        ) as pbar:
            pbar.set_description("Preparing multi-sheet import")
            pbar.update(10)

            pbar.set_description("Deleting existing data")
            pbar.update(20)

            # Execute actual import
            try:
                import_result = self.broadcast_service.execute_month_replacement(
                    str(excel_file),
                    "WEEKLY_UPDATE",  # Use WEEKLY_UPDATE mode for daily operations
                    closed_by=None,
                    dry_run=False,
                )
            except Exception as e:
                self.progress_reporter.write(f"❌ Import failed: {e}")
                return ImportResult(
                    success=False,
                    records_imported=0,
                    records_deleted=0,
                    batch_id=batch_id,
                    duration_seconds=(datetime.now() - start_time).total_seconds(),
                    error_message=f"Import failed: {e}",
                    months_processed=[],
                    sheet_breakdown={},
                )

            pbar.set_description("Importing combined data")
            pbar.update(50)

            pbar.set_description("Finalizing import")
            pbar.update(20)

            pbar.set_description("Import complete")
            pbar.update(100)

        # Enhanced: Get sheet source breakdown from database
        try:
            sheet_breakdown = self.spot_repository.get_sheet_source_breakdown(batch_id)
        except Exception as e:
            self.progress_reporter.write(f"⚠️ Warning: Could not get sheet breakdown: {e}")
            sheet_breakdown = {}

        # Convert to our domain model
        result = ImportResult(
            success=import_result.success,
            records_imported=import_result.records_imported,
            records_deleted=import_result.records_deleted,
            batch_id=batch_id,
            duration_seconds=(datetime.now() - start_time).total_seconds(),
            months_processed=summary.get("months_in_excel", []),
            sheet_breakdown=sheet_breakdown,
        )

        if result.success:
            self.progress_reporter.write(f"✅ Import completed successfully:")
            self.progress_reporter.write(f"   Records imported: {result.records_imported:,}")
            self.progress_reporter.write(f"   Records deleted: {result.records_deleted:,}")
            self.progress_reporter.write(f"   Net change: {result.net_change:+,}")
            self.progress_reporter.write(f"   Duration: {result.duration_seconds:.1f} seconds")
            
            # Enhanced: Show multi-sheet breakdown
            if result.has_multisheet_data:
                self.progress_reporter.write(f"📊 Multi-sheet breakdown:")
                for sheet_source, count in result.sheet_breakdown.items():
                    self.progress_reporter.write(f"   {sheet_source}: {count:,} spots")
            
            # Show months processed
            if result.months_processed:
                self.progress_reporter.write(f"📅 Months processed: {', '.join(result.months_processed)}")
                
        else:
            self.progress_reporter.write(f"❌ Import failed!")
            if hasattr(import_result, 'error_message') and import_result.error_message:
                self.progress_reporter.write(f"   Error: {import_result.error_message}")

        return result

    def simulate_import(self, excel_file: Path) -> ImportResult:
        """Simulate import for dry run with multi-sheet preview"""
        try:
            summary = get_excel_import_summary(
                str(excel_file), self.broadcast_service.db_connection.db_path
            )
        except Exception as e:
            self.progress_reporter.write(f"⚠️ Warning: Could not get import summary: {e}")
            summary = {"months_in_excel": [], "total_existing_spots_affected": 0}

        # Try to get sheet breakdown for dry run
        sheet_breakdown = {}
        try:
            import pandas as pd

            # Try multiple sheet names
            sheet_names_to_try = ["Commercials", "Commercial Lines", "Sheet1", 0]
            for sheet_name in sheet_names_to_try:
                try:
                    df = pd.read_excel(excel_file, sheet_name=sheet_name)
                    sheet_breakdown[str(sheet_name)] = len(df)
                    break
                except Exception:
                    continue
                    
            # If we found data, also try to simulate the normalization update
            if sheet_breakdown:
                self.progress_reporter.write(f"🔍 DRY RUN: Would update normalization system")
                
                # Show what bill codes would be added
                df = None
                for sheet_name in sheet_names_to_try:
                    try:
                        df = pd.read_excel(excel_file, sheet_name=sheet_name)
                        break
                    except Exception:
                        continue
                
                if df is not None:
                    bill_code_columns = ['bill_code', 'Bill Code', 'Customer', 'Client', 'Advertiser']
                    for col in bill_code_columns:
                        if col in df.columns:
                            unique_codes = df[col].dropna().nunique()
                            self.progress_reporter.write(f"   Found {unique_codes} unique bill codes in column '{col}'")
                            break
                    
        except Exception as e:
            self.progress_reporter.write(f"⚠️ Warning: Could not analyze Excel for dry run: {e}")

        return ImportResult(
            success=True,
            records_imported=summary.get("total_existing_spots_affected", 0),
            records_deleted=summary.get("total_existing_spots_affected", 0),
            batch_id="dry_run",
            duration_seconds=0.0,
            months_processed=summary.get("months_in_excel", []),
            sheet_breakdown=sheet_breakdown,
        )

    def ensure_bill_codes_in_raw_inputs(self, excel_file: Path) -> None:
        """
        Ensure all bill_code values from Excel are added to raw_customer_inputs
        This prevents normalization gaps from occurring.
        """
        try:
            import pandas as pd
            
            # Read the main sheet (adjust sheet name as needed)
            df = pd.read_excel(excel_file, sheet_name="Commercials")  # or your main sheet
            
            # Get unique bill codes
            if 'bill_code' in df.columns:
                bill_codes = df['bill_code'].dropna().unique().tolist()
            else:
                # Try alternative column names
                bill_code_columns = ['Bill Code', 'Customer', 'Client', 'Advertiser']
                bill_codes = []
                for col in bill_code_columns:
                    if col in df.columns:
                        bill_codes = df[col].dropna().unique().tolist()
                        break
            
            if bill_codes:
                # Add to raw_customer_inputs
                current_time = datetime.now().isoformat()
                
                with self.broadcast_service.db_connection.transaction() as conn:
                    for bill_code in bill_codes:
                        conn.execute("""
                            INSERT OR IGNORE INTO raw_customer_inputs (raw_text, created_at)
                            VALUES (?, ?)
                        """, (str(bill_code).strip(), current_time))
                
                self.progress_reporter.write(f"Added {len(bill_codes)} bill codes to normalization system")
            
        except Exception as e:
            self.progress_reporter.write(f"Warning: Could not update raw_customer_inputs: {e}")


    def simulate_import(self, excel_file: Path) -> ImportResult:
        """Simulate import for dry run with multi-sheet preview"""
        summary = get_excel_import_summary(
            str(excel_file), self.broadcast_service.db_connection.db_path
        )

        # Try to get sheet breakdown for dry run
        try:
            import pandas as pd

            df = pd.read_excel(excel_file, sheet_name="Commercials")
            sheet_breakdown = {}
            if "sheet_source" in df.columns:
                breakdown_series = df["sheet_source"].value_counts()
                sheet_breakdown = dict(breakdown_series)
        except:
            sheet_breakdown = {}

        return ImportResult(
            success=True,
            records_imported=summary["total_existing_spots_affected"],
            records_deleted=summary["total_existing_spots_affected"],
            batch_id="dry_run",
            duration_seconds=0.0,
            months_processed=summary["months_in_excel"],
            sheet_breakdown=sheet_breakdown,
        )


# ============================================================================
# Presentation Layer (Enhanced for Multi-Sheet)
# ============================================================================


class DailyUpdatePreviewService:
    """Enhanced service for displaying multi-sheet preview"""

    def __init__(
        self, db_connection: DatabaseConnection, progress_reporter: ProgressReporter
    ):
        self.db = db_connection
        self.progress_reporter = progress_reporter

    def display_enhanced_daily_preview(self, config: DailyUpdateConfig) -> bool:
        """Display what the enhanced multi-sheet daily update would do"""
        if not config.unattended:
            self.progress_reporter.write(f"Enhanced Multi-Sheet Daily Update Preview")
            self.progress_reporter.write(f"=" * 60)
            self.progress_reporter.write(f"File: {config.excel_file.name}")
            self.progress_reporter.write(f"Auto-setup: {config.auto_setup_markets}")
            self.progress_reporter.write("")

        try:
            # Enhanced multi-sheet preview
            preview = MultiSheetPreviewGenerator.generate_preview(
                config.excel_file, self.db
            )

            if preview.has_worldlink_data:
                self.progress_reporter.write(f"Multi-Sheet Data Detected:")
                self.progress_reporter.write(f"   {preview.summary_line}")
                self.progress_reporter.write("")

            # Market setup preview
            if config.auto_setup_markets:
                self._display_market_setup_preview(config.excel_file)

            # Standard import preview with multi-sheet awareness
            can_proceed = self._display_import_preview(config.excel_file, preview)

            # Language assignment preview (only for interactive mode)
            if can_proceed and not config.unattended:
                self.progress_reporter.write(f"Language Assignment Preview:")
                self.progress_reporter.write(
                    f"   Languages assigned directly from Excel language column"
                )
                self.progress_reporter.write(
                    f"   Simple and reliable - no complex categorization needed"
                )
                if preview.has_worldlink_data:
                    self.progress_reporter.write(
                        f"   WorldLink and Commercial data processed identically"
                    )

            return can_proceed

        except Exception as e:
            self.progress_reporter.write(f"Error generating preview: {e}")
            return False

    def _display_market_setup_preview(self, excel_file: Path) -> None:
        """Display market setup preview"""
        try:
            market_repo = MarketRepository(self.db)
            existing_markets = set(market_repo.get_existing_markets().keys())

            scanner = ExcelMarketScanner(self.progress_reporter)
            new_markets = scanner.scan_for_new_markets(excel_file, existing_markets)

            if new_markets:
                self.progress_reporter.write(f"Market Setup Preview:")
                self.progress_reporter.write(
                    f"   New markets: {len(new_markets)} ({', '.join(sorted(new_markets.keys()))})"
                )
                self.progress_reporter.write("")
        except Exception as e:
            self.progress_reporter.write(f"Could not preview market setup: {e}")

    def _display_import_preview(
        self, excel_file: Path, preview: MultiSheetPreview
    ) -> bool:
        """Display enhanced import preview with multi-sheet information"""
        self.progress_reporter.write(f"Daily Update Preview:")

        try:
            # Get Excel summary
            summary = get_excel_import_summary(str(excel_file), self.db.db_path)

            # Validation check
            validation = validate_excel_for_import(
                str(excel_file), "WEEKLY_UPDATE", self.db.db_path
            )

            if validation.is_valid:
                self.progress_reporter.write(f"   Daily update allowed")
                self.progress_reporter.write(
                    f"   {summary['total_existing_spots_affected']:,} spots across {len(summary['months_in_excel'])} months"
                )

                # Enhanced: Show multi-sheet breakdown in preview
                if preview.sheet_breakdown:
                    self.progress_reporter.write(
                        f"   Sheet breakdown: {preview.summary_line}"
                    )
            else:
                self.progress_reporter.write(f"   Daily update BLOCKED")
                self.progress_reporter.write(
                    f"      Reason: {validation.error_message}"
                )
                self.progress_reporter.write(
                    f"      Solution: {validation.suggested_action}"
                )
                return False

            # Show month details concisely
            open_months = summary["open_months"]
            closed_months = summary["closed_months"]

            self.progress_reporter.write(
                f"   Open months: {len(open_months)} ({', '.join(open_months) if len(open_months) <= 5 else f'{len(open_months)} months'})"
            )
            if closed_months:
                self.progress_reporter.write(
                    f"   Closed months: {len(closed_months)} (protected)"
                )
            self.progress_reporter.write("")

            return True

        except Exception as e:
            self.progress_reporter.write(f"Error generating preview: {e}")
            return False


# ============================================================================
# Application Layer (Enhanced CLI Interface)
# ============================================================================


class DailyUpdateCLI:
    """Enhanced CLI interface optimized for multi-sheet automated processing"""

    def __init__(self, db_connection: DatabaseConnection):
        self.db_connection = db_connection
        self.progress_reporter = None  # Will be set based on mode
        self.logger = None  # Will be set if logging configured

    def execute_from_args(self, args) -> int:
        """Execute enhanced daily update from command line arguments"""
        try:
            config = DailyUpdateConfig.from_args(args)

            # Configure logging and progress reporting
            self._setup_logging_and_progress(config)

            # Validate inputs
            if not config.excel_file.exists():
                self._log_error(f"Excel file not found: {config.excel_file}")
                return 1

            if not config.db_path.exists():
                self._log_error(f"Database not found: {config.db_path}")
                self._log_error(
                    f"Run: uv run python scripts/setup_database.py --db-path {config.db_path}"
                )
                return 1

            # Enhanced header for multi-sheet processing
            if config.unattended:
                self.progress_reporter.write(
                    f"Enhanced Multi-Sheet Daily Update Tool - UNATTENDED MODE"
                )
                self.progress_reporter.write(f"File: {config.excel_file.name}")
                self.progress_reporter.write(f"Database: {config.db_path.name}")
                self.progress_reporter.write(f"Sheets: Commercials + Worldlink Lines")
            else:
                print(f"Enhanced Multi-Sheet Daily Update Tool")
                print(f"File: {config.excel_file.name}")
                print(f"Database: {config.db_path.name}")
                print(f"Processing: Commercials + Worldlink Lines sheets")
                if config.dry_run:
                    print(f"Mode: DRY RUN (no changes will be made)")
                print()

            # Display enhanced preview and validate
            preview_service = DailyUpdatePreviewService(
                self.db_connection, self.progress_reporter
            )
            can_proceed = preview_service.display_enhanced_daily_preview(config)

            if not can_proceed:
                self._display_validation_error_help(config.unattended)
                return 1

            # Get confirmation unless forced, dry run, or unattended
            if not config.dry_run and not config.force and not config.unattended:
                confirmed = self._get_confirmation(config)
                if not confirmed:
                    self.progress_reporter.write(f"Daily update cancelled by user")
                    return 0

            # Create enhanced services with dependency injection
            orchestrator = self._create_enhanced_orchestrator()

            # Execute the enhanced update
            result = orchestrator.execute_daily_update(config)

            # Display enhanced results
            presenter = DailyUpdateResultsPresenter(self.progress_reporter)
            presenter.display_final_results(result, config.dry_run, config.unattended)

            # Enhanced logging for automated monitoring
            if config.unattended and self.logger:
                self.logger.info(
                    f"Enhanced daily update completed: {result.summary_line}"
                )
                if result.import_result and result.import_result.has_multisheet_data:
                    breakdown_str = ", ".join(
                        [
                            f"{k}: {v}"
                            for k, v in result.import_result.sheet_breakdown.items()
                        ]
                    )
                    self.logger.info(f"Multi-sheet processing: {breakdown_str}")
                if not result.success:
                    self.logger.error(
                        f"Enhanced daily update failed: {'; '.join(result.error_messages)}"
                    )

            return 0 if result.success else 1

        except BroadcastMonthImportError as e:
            error_msg = f"Import error: {e}"
            self._log_error(error_msg)
            return 1
        except KeyboardInterrupt:
            self._log_error(f"Daily update cancelled by user")
            return 1
        except Exception as e:
            error_msg = f"Unexpected error: {e}"
            self._log_error(error_msg)
            if hasattr(args, "verbose") and args.verbose:
                import traceback

                if self.logger:
                    self.logger.error(f"Stack trace: {traceback.format_exc()}")
                else:
                    traceback.print_exc()
            return 1

    def _setup_logging_and_progress(self, config: DailyUpdateConfig) -> None:
        """Configure logging and progress reporting based on mode"""
        if config.unattended:
            # Setup enhanced file logging for multi-sheet operations
            self.logger = LoggingConfigService.setup_unattended_logging(
                config.log_file, config.verbose
            )
            self.progress_reporter = LoggingProgressReporter(self.logger)
        else:
            # Use interactive tqdm progress
            self.progress_reporter = TqdmProgressReporter()

    def _log_error(self, message: str) -> None:
        """Log error message to appropriate output"""
        if self.logger:
            self.logger.error(message)
        else:
            print(f"ERROR: {message}")

    def _create_enhanced_orchestrator(self) -> DailyUpdateOrchestrator:
        """Create orchestrator with enhanced multi-sheet dependencies"""
        # Create repositories
        market_repo = MarketRepository(self.db_connection)
        spot_repo = SpotRepository(self.db_connection)

        # Create enhanced services
        market_setup_service = MarketSetupService(market_repo, self.progress_reporter)
        import_service = ImportService(
            BroadcastMonthImportService(self.db_connection),
            spot_repo,
            self.progress_reporter,
        )
        language_service = LanguageAssignmentService(
            self.db_connection, self.progress_reporter
        )

        # Create orchestrator
        return DailyUpdateOrchestrator(
            market_setup_service,
            import_service,
            language_service,
            self.progress_reporter,
        )

    def _get_confirmation(self, config: DailyUpdateConfig) -> bool:
        """Get user confirmation with multi-sheet information"""
        summary = get_excel_import_summary(str(config.excel_file), config.db_path)

        # Enhanced: Get multi-sheet preview for confirmation
        preview = MultiSheetPreviewGenerator.generate_preview(
            config.excel_file, self.db_connection
        )

        new_market_count = 0
        if config.auto_setup_markets:
            # Estimate new markets
            market_repo = MarketRepository(self.db_connection)
            existing_markets = set(market_repo.get_existing_markets().keys())
            scanner = ExcelMarketScanner(self.progress_reporter)
            new_markets = scanner.scan_for_new_markets(
                config.excel_file, existing_markets
            )
            new_market_count = len(new_markets)

        # Enhanced confirmation with multi-sheet info
        return self._get_enhanced_user_confirmation(
            summary["total_existing_spots_affected"],
            len(summary["open_months"]),
            new_market_count,
            preview,
            config.force,
            config.unattended,
        )

    def _get_enhanced_user_confirmation(
        self,
        total_spots: int,
        open_months: int,
        new_markets: int,
        preview: MultiSheetPreview,
        force: bool,
        unattended: bool,
    ) -> bool:
        """Enhanced user confirmation with multi-sheet information"""
        if force or unattended:
            return True

        tqdm.write(f"CONFIRMATION REQUIRED")
        actions = [
            f"REPLACE {total_spots:,} existing spots",
            f"Update {open_months} open months",
            "Process language assignments automatically",
            "Preserve all closed/historical months",
        ]

        # Add multi-sheet specific information
        if preview.has_worldlink_data:
            actions.insert(0, f"Process combined data: {preview.summary_line}")

        if new_markets > 0:
            actions.insert(-1, f"Create {new_markets} new markets")

        tqdm.write(f"This will:")
        for action in actions:
            tqdm.write(f"  • {action}")

        while True:
            response = (
                input(f"\nProceed with enhanced multi-sheet daily update? (yes/no): ")
                .strip()
                .lower()
            )
            if response in ["yes", "y"]:
                return True
            elif response in ["no", "n", ""]:
                return False
            else:
                tqdm.write("Please enter 'yes' or 'no'")

    def _display_validation_error_help(self, unattended: bool) -> None:
        """Display validation error help"""
        if unattended:
            self.progress_reporter.write(
                f"Enhanced daily update cannot proceed due to validation errors"
            )
            self.progress_reporter.write(
                f"Solutions: Remove closed month data from Excel file OR use historical import mode"
            )
        else:
            print(f"\nEnhanced daily update cannot proceed due to validation errors")
            print(f"Common solutions:")
            print(f"  • Remove closed month data from Excel file")
            print(f"  • Use historical import mode for closed months")
            print(f"  • Check if months need to be manually closed first")


# Import other classes unchanged but with enhanced logging
# (LanguageAssignmentService, DailyUpdateOrchestrator, etc.)
# ============================================================================
# Import Other Services (Unchanged Core Logic)
# ============================================================================


class LanguageAssignmentService:
    """Service for processing language assignments after import"""

    def __init__(
        self, db_connection: DatabaseConnection, progress_reporter: ProgressReporter
    ):
        self.db = db_connection
        self.progress_reporter = progress_reporter

    def process_languages_directly(self, batch_id: str) -> LanguageAssignmentResult:
        """Process languages directly from imported spots without categorization"""
        start_time = datetime.now()

        result = LanguageAssignmentResult(
            success=False,
            categorized=0,
            processed=0,
            language_assigned=0,
            default_english_assigned=0,
            flagged_for_review=0,
        )

        try:
            # Find the actual batch ID that was used (handle batch ID mismatch)
            with self.db.transaction() as conn:
                # First try the provided batch_id
                cursor = conn.execute(
                    "SELECT COUNT(*) FROM spots WHERE import_batch_id = ?", (batch_id,)
                )
                if cursor.fetchone()[0] > 0:
                    actual_batch_id = batch_id
                else:
                    # Fall back to today's most recent batch
                    cursor = conn.execute("""
                        SELECT import_batch_id FROM spots 
                        WHERE DATE(load_date) = DATE('now')
                        ORDER BY load_date DESC LIMIT 1
                    """)
                    row = cursor.fetchone()
                    actual_batch_id = row[0] if row else batch_id

                # Get all spots that need language assignment
                cursor = conn.execute(
                    """
                    SELECT spot_id FROM spots 
                    WHERE import_batch_id = ?
                    AND spot_id NOT IN (SELECT spot_id FROM spot_language_assignments)
                """,
                    (actual_batch_id,),
                )
                spot_ids = [row[0] for row in cursor.fetchall()]

            if not spot_ids:
                self.progress_reporter.write("No spots need language assignment")
                result.success = True
                return result

            # Process spots directly - import only what we need
            from src.services.language_assignment_service import (
                LanguageAssignmentService,
            )
            from src.database.language_assignment_queries import (
                LanguageAssignmentQueries,
            )

            # Create the service using the proper database connection
            language_queries = LanguageAssignmentQueries(self.db)

            processed = 0
            with self.progress_reporter.create_progress(
                "Processing languages directly", len(spot_ids)
            ) as pbar:
                with self.db.transaction() as conn:
                    for spot_id in spot_ids:
                        try:
                            # Get the language code directly from the database
                            cursor = conn.execute(
                                """
                                SELECT language_code FROM spots WHERE spot_id = ?
                            """,
                                (spot_id,),
                            )
                            row = cursor.fetchone()

                            if row and row[0]:
                                language_code = row[0]

                                # Insert directly into spot_language_assignments table
                                conn.execute(
                                    """
                                    INSERT INTO spot_language_assignments 
                                    (spot_id, language_code, language_status, confidence, assignment_method, assigned_date, requires_review)
                                    VALUES (?, ?, 'determined', 1.0, 'direct_from_excel', CURRENT_TIMESTAMP, 0)
                                """,
                                    (spot_id, language_code),
                                )

                                processed += 1

                            if processed % 100 == 0:
                                pbar.update(100)

                        except Exception:
                            continue

            result.processed = processed
            result.language_assigned = processed
            result.success = True

            self.progress_reporter.write(
                f"Direct language processing complete: {processed:,} spots processed"
            )

        except Exception as e:
            result.error_messages.append(f"Direct processing failed: {str(e)}")
            self.progress_reporter.write(f"ERROR: Direct processing failed: {str(e)}")

        result.duration_seconds = (datetime.now() - start_time).total_seconds()
        return result


class DailyUpdateOrchestrator:
    """Enhanced orchestrator for multi-sheet daily update process"""

    def __init__(
        self,
        market_setup_service: MarketSetupService,
        import_service: ImportService,
        language_service: LanguageAssignmentService,
        progress_reporter: ProgressReporter,
    ):
        self.market_setup_service = market_setup_service
        self.import_service = import_service
        self.language_service = language_service
        self.progress_reporter = progress_reporter

    def execute_daily_update(self, config: DailyUpdateConfig) -> DailyUpdateResult:
        """Execute the complete enhanced multi-sheet daily update process"""
        start_time = datetime.now()
        batch_id = BatchIdGenerator.generate(BatchType.ENHANCED_DAILY, start_time)

        self._display_update_header(config, batch_id)

        result = DailyUpdateResult(success=False, batch_id=batch_id)

        try:
            # Step 1: Market setup (if enabled)
            if config.auto_setup_markets and not config.dry_run:
                self.progress_reporter.write(f"STEP 1: Automatic Market Setup")
                result.market_setup = (
                    self.market_setup_service.execute_daily_market_setup(
                        config.excel_file
                    )
                )
                self.progress_reporter.write(f"Setup: {result.market_setup.summary}")
                self.progress_reporter.write("")

            # Step 2: Enhanced multi-sheet data import
            self.progress_reporter.write(f"STEP 2: Enhanced Multi-Sheet Data Import")

            if config.dry_run:
                self.progress_reporter.write(f"DRY RUN - No changes would be made")
                result.import_result = self.import_service.simulate_import(
                    config.excel_file
                )
                if result.import_result.has_multisheet_data:
                    self.progress_reporter.write(
                        f"Would process: {result.import_result.sheet_breakdown}"
                    )
            else:
                result.import_result = self.import_service.execute_import_with_progress(
                    config.excel_file, batch_id
                )

            # Step 3: Language Assignment Processing (if import succeeded)
            if (
                result.import_result
                and result.import_result.success
                and not config.dry_run
            ):
                self.progress_reporter.write(f"STEP 3: Language Assignment Processing")
                result.language_assignment = (
                    self.language_service.process_languages_directly(batch_id)
                )

            result.success = True

        except Exception as e:
            error_msg = f"Enhanced multi-sheet daily update failed: {str(e)}"
            result.error_messages.append(error_msg)
            self.progress_reporter.write(f"ERROR: {error_msg}")

        # Calculate total duration
        result.duration_seconds = (datetime.now() - start_time).total_seconds()
        return result

    def _display_update_header(self, config: DailyUpdateConfig, batch_id: str) -> None:
        """Display enhanced header for multi-sheet processing"""
        self.progress_reporter.write(f"Enhanced Multi-Sheet Daily Update Starting")
        self.progress_reporter.write(f"File: {config.excel_file.name}")
        self.progress_reporter.write(f"Processing: Commercials + Worldlink Lines data")
        self.progress_reporter.write(
            f"Auto-setup: {config.auto_setup_markets} | Dry run: {config.dry_run}"
        )
        self.progress_reporter.write(f"Batch ID: {batch_id}")
        self.progress_reporter.write("=" * 60)


class DailyUpdateResultsPresenter:
    """Enhanced service for presenting multi-sheet daily update results"""

    def __init__(self, progress_reporter: ProgressReporter):
        self.progress_reporter = progress_reporter

    def display_final_results(
        self, result: DailyUpdateResult, dry_run: bool, unattended: bool = False
    ) -> None:
        """Display enhanced final results with multi-sheet information"""
        if not unattended:
            self.progress_reporter.write(f"\n{'=' * 60}")

        if dry_run:
            self.progress_reporter.write(
                f"ENHANCED MULTI-SHEET DAILY UPDATE DRY RUN COMPLETED"
            )
        else:
            self.progress_reporter.write(f"ENHANCED MULTI-SHEET DAILY UPDATE COMPLETED")

        if not unattended:
            self.progress_reporter.write(f"{'=' * 60}")

        self.progress_reporter.write(f"Results Summary:")
        self.progress_reporter.write(
            f"  Status: {'SUCCESS' if result.success else 'FAILED'}"
        )
        self.progress_reporter.write(
            f"  Duration: {result.duration_seconds:.1f} seconds"
        )

        # Market setup results (concise)
        if result.market_setup and result.market_setup.markets_created > 0:
            self.progress_reporter.write(
                f"  Markets created: {result.market_setup.markets_created}"
            )

        # Enhanced import results with multi-sheet breakdown
        if result.import_result and not dry_run:
            import_res = result.import_result
            if import_res.success:
                self.progress_reporter.write(
                    f"  Spots imported: {import_res.records_imported:,}"
                )
                self.progress_reporter.write(
                    f"  Net change: {import_res.net_change:+,}"
                )

                # Show multi-sheet breakdown if available
                if import_res.has_multisheet_data:
                    self.progress_reporter.write(f"  Sheet sources:")
                    for sheet_source, count in import_res.sheet_breakdown.items():
                        self.progress_reporter.write(
                            f"    {sheet_source}: {count:,} spots"
                        )

        # Language assignment results (concise)
        if result.language_assignment and result.language_assignment.success:
            lang_res = result.language_assignment
            self.progress_reporter.write(
                f"  Language processed: {lang_res.processed:,}"
            )
            if lang_res.flagged_for_review > 0:
                self.progress_reporter.write(
                    f"  Review needed: {lang_res.flagged_for_review:,}"
                )

        if result.error_messages:
            self.progress_reporter.write(f"ERRORS:")
            for error in result.error_messages:
                self.progress_reporter.write(f"  • {error}")

        if result.success and not dry_run:
            self.progress_reporter.write(
                f"Enhanced multi-sheet update completed successfully!"
            )
            if (
                result.language_assignment
                and result.language_assignment.flagged_for_review > 0
            ):
                self.progress_reporter.write(
                    f"Next: Review flagged spots with 'uv run python cli/assign_languages.py --review-required'"
                )


# ============================================================================
# Logging Configuration (Enhanced)
# ============================================================================


class LoggingConfigService:
    """Enhanced service for configuring logging for multi-sheet automated operations"""

    @staticmethod
    def setup_unattended_logging(
        log_file: Optional[Path], verbose: bool
    ) -> logging.Logger:
        """Configure enhanced logging for unattended multi-sheet operation"""
        logger = logging.getLogger("enhanced_daily_update")
        logger.setLevel(logging.DEBUG if verbose else logging.INFO)

        # Remove existing handlers
        for handler in logger.handlers[:]:
            logger.removeHandler(handler)

        # Configure file handler if log file specified
        if log_file:
            log_file.parent.mkdir(parents=True, exist_ok=True)

            # Rotating file handler (20MB max, keep 10 files for multi-sheet operations)
            file_handler = logging.handlers.RotatingFileHandler(
                log_file,
                maxBytes=20 * 1024 * 1024,  # 20MB (increased for multi-sheet logs)
                backupCount=10,  # Keep more logs for better debugging
                encoding="utf-8",
            )

            # Enhanced formatter for multi-sheet operations
            file_formatter = logging.Formatter(
                "%(asctime)s - %(levelname)s - [%(name)s] - %(message)s",
                datefmt="%Y-%m-%d %H:%M:%S",
            )
            file_handler.setFormatter(file_formatter)
            logger.addHandler(file_handler)

        # Always add console handler for immediate feedback
        console_handler = logging.StreamHandler()
        console_formatter = logging.Formatter("%(levelname)s: %(message)s")
        console_handler.setFormatter(console_formatter)
        logger.addHandler(console_handler)

        return logger


# ============================================================================
# Main Entry Point (Enhanced)
# ============================================================================


def main():
    parser = argparse.ArgumentParser(
        description="Enhanced multi-sheet daily update with automatic market setup and language assignment",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Enhanced multi-sheet daily update with language assignment
  python daily_update.py data/raw/daily/Commercial\ Log\ 250904.xlsx

  # Enhanced daily update with automatic market setup (recommended for automation)
  python daily_update.py data/raw/daily/Commercial\ Log\ 250904.xlsx --auto-setup

  # Unattended mode for automated multi-sheet processing
  python daily_update.py data/raw/daily/Commercial\ Log\ 250904.xlsx --auto-setup --unattended --log-file /var/log/ctv-daily-update/update.log

  # Preview what would happen with multi-sheet data
  python daily_update.py data/raw/daily/Commercial\ Log\ 250904.xlsx --auto-setup --dry-run

Enhanced Multi-Sheet Features:
  • Automatic processing of Commercials + Worldlink Lines data
  • Enhanced source tracking with filename:sheet_name format
  • Sheet-specific breakdown in logging and reporting
  • Optimized for automated daily processing with detailed audit trail
  • Clean Architecture implementation with proper separation of concerns
  • Backward compatibility with single-sheet files
        """,
    )

    parser.add_argument("excel_file", help="Path to multi-sheet Excel file to import")
    parser.add_argument(
        "--auto-setup",
        action="store_true",
        help="Automatically create missing markets and schedule assignments",
    )
    parser.add_argument(
        "--db-path",
        default="data/database/production.db",
        help="Database path (default: production.db)",
    )
    parser.add_argument(
        "--dry-run", action="store_true", help="Preview import without making changes"
    )
    parser.add_argument(
        "--force", action="store_true", help="Skip confirmation prompts"
    )
    parser.add_argument("--verbose", action="store_true", help="Enable verbose logging")
    parser.add_argument(
        "--unattended",
        action="store_true",
        help="Run in unattended mode (no prompts, suitable for automation)",
    )
    parser.add_argument(
        "--log-file",
        type=str,
        help="Log file path for unattended mode (enables file logging with rotation)",
    )

    args = parser.parse_args()

    # Setup logging for non-unattended mode
    if not args.unattended:
        import logging

        level = logging.DEBUG if args.verbose else logging.WARNING
        logging.basicConfig(level=level, format="%(levelname)s: %(message)s")

    # Create database connection
    try:
        db_connection = DatabaseConnection(args.db_path)

        # Create and execute enhanced CLI
        cli = DailyUpdateCLI(db_connection)
        exit_code = cli.execute_from_args(args)

        db_connection.close()
        sys.exit(exit_code)

    except Exception as e:
        if args.unattended and args.log_file:
            # Emergency logging if main logger setup failed
            try:
                with open(args.log_file, "a", encoding="utf-8") as f:
                    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    f.write(
                        f"{timestamp} - ERROR - Failed to initialize enhanced daily update: {e}\n"
                    )
            except:
                pass
        print(f"Failed to initialize enhanced daily update: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
