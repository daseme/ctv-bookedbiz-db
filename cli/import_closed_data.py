#!/usr/bin/env python3
"""
Production Sales Data Import Tool

The primary import solution for closed sales data, handling both historical setup and 
ongoing monthly operations with automatic market configuration, progress tracking, 
and language assignment processing.

CORE FUNCTIONALITY:
• Imports closed sales data from Excel files (1K-400K+ records)
• Auto-creates missing markets and schedule assignments as needed
• Processes language assignments using business rules
• Closes imported months permanently (closed month protection)
• Provides real-time progress bars for all operations

TYPICAL USE CASES:
• Monthly closed month imports (primary ongoing workflow)
• Initial database setup with historical annual data
• Quarterly or periodic bulk data imports
• Any closed data requiring permanent month protection

KEY FEATURES:
• Memory-efficient streaming import handles 400K+ records
• Intelligent market detection and automatic database setup
• Integrated language assignment workflow (categorization → processing)
• Transaction safety with automatic rollback on failure
• Comprehensive audit trail with batch tracking

WORKFLOW:
1. Analyzes Excel file structure and market requirements
2. Creates missing markets and schedule assignments (if needed)
3. Imports closed sales data with automatic month closure protection
4. Categorizes spots for language assignment processing
5. Applies business rules and flags spots requiring manual review

USAGE EXAMPLES:
    # Monthly closed month import (typical workflow)
    python bulk_import_closed_data.py data/monthly/May-2025.xlsx --year 2025 --closed-by "Kurt"
    
    # Historical annual import (initial setup)
    python bulk_import_closed_data.py data/historical/2024-complete.xlsx --year 2024 --closed-by "Kurt" --auto-setup
    
    # Preview before importing
    python bulk_import_closed_data.py data/March-2025.xlsx --year 2025 --closed-by "Kurt" --dry-run

CAUTION:
This tool permanently closes imported months and cannot be undone. Always use --dry-run
first to preview changes. Once months are closed, they become read-only and protected 
from modification. Use for closed months only - never for open/current month data.

Dependencies: openpyxl, tqdm, psutil, database services

Dependencies: openpyxl, tqdm, psutil, database services
"""

import sys
import argparse
import gc
import psutil
import os
from pathlib import Path
from datetime import datetime
from typing import Optional, Generator, Set, Dict, List, Tuple
import sqlite3

# Add tqdm for progress bars
from tqdm import tqdm

# Add src to path
sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'src'))

from services.broadcast_month_import_service import BroadcastMonthImportService
from services.import_integration_utilities import get_excel_import_summary
from utils.broadcast_month_utils import BroadcastMonthParser
from database.connection import DatabaseConnection


class EnhancedMarketSetupManager:
    """Enhanced market setup manager with tqdm progress bars."""
    
    def __init__(self, db_connection: DatabaseConnection):
        self.db = db_connection
        self.parser = BroadcastMonthParser()
    
    def scan_excel_for_markets(self, excel_file: str) -> Dict[str, Dict]:
        """
        Scan Excel file to extract all market codes with tqdm progress bar.
        """
        print(f"🔍 Scanning Excel file for market codes: {excel_file}")
        
        try:
            from openpyxl import load_workbook
            
            workbook = load_workbook(excel_file, read_only=True, data_only=True)
            worksheet = workbook.active
            
            # Find required columns
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
            market_col_index = None
            air_date_col_index = None
            
            for i, header in enumerate(header_row):
                if header:
                    header_str = str(header).strip().lower()
                    if header_str in ['market_name', 'market', 'market_code', 'market name']:
                        market_col_index = i
                    elif header_str in ['start date', 'air date', 'date', 'airdate', 'air_date']:
                        air_date_col_index = i
            
            if market_col_index is None:
                raise ValueError("Market column not found in Excel file")
            if air_date_col_index is None:
                raise ValueError("Air date column not found in Excel file")
            
            print(f"📍 Found Market column at index {market_col_index}")
            print(f"📍 Found Air Date column at index {air_date_col_index}")
            
            # ENHANCED: Get total row count for tqdm
            total_rows = worksheet.max_row - 1  # Subtract header row
            
            markets_data = {}
            
            # ENHANCED: Use tqdm for progress tracking
            with tqdm(total=total_rows, desc="📊 Scanning Excel rows", unit=" rows") as pbar:
                for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                    pbar.update(1)
                    
                    if not any(cell for cell in row):
                        continue
                    
                    if market_col_index < len(row) and air_date_col_index < len(row):
                        market_value = row[market_col_index]
                        air_date_value = row[air_date_col_index]
                        
                        if market_value and air_date_value:
                            market_code = str(market_value).strip()
                            
                            # Parse air date
                            try:
                                if isinstance(air_date_value, datetime):
                                    air_date = air_date_value.date()
                                else:
                                    air_date = datetime.strptime(str(air_date_value), '%Y-%m-%d').date()
                            except:
                                continue
                            
                            # Track market data
                            if market_code not in markets_data:
                                markets_data[market_code] = {
                                    'earliest_date': air_date,
                                    'latest_date': air_date,
                                    'spot_count': 0
                                }
                            else:
                                if air_date < markets_data[market_code]['earliest_date']:
                                    markets_data[market_code]['earliest_date'] = air_date
                                if air_date > markets_data[market_code]['latest_date']:
                                    markets_data[market_code]['latest_date'] = air_date
                            
                            markets_data[market_code]['spot_count'] += 1
                    
                    # Update progress description periodically
                    if row_num % 5000 == 0:
                        pbar.set_description(f"📊 Scanning ({len(markets_data)} markets found)")
            
            workbook.close()
            
            print(f"\n✅ Excel scan complete:")
            print(f"   📊 {total_rows:,} total spots analyzed")
            print(f"   🎯 {len(markets_data)} unique markets found")
            
            # Display market summary with tqdm
            print(f"\n📋 Market Summary:")
            for market_code, data in tqdm(sorted(markets_data.items()), desc="📋 Displaying markets", leave=False):
                print(f"   📋 {market_code}: {data['spot_count']:,} spots ({data['earliest_date']} to {data['latest_date']})")
            
            return markets_data
            
        except Exception as e:
            raise RuntimeError(f"Failed to scan Excel for markets: {str(e)}")
    
    def create_missing_markets(self, excel_markets: Dict[str, Dict]) -> Dict[str, int]:
        """Create missing markets with tqdm progress bar."""
        existing_markets = self.get_existing_markets()
        missing_markets = set(excel_markets.keys()) - set(existing_markets.keys())
        
        if not missing_markets:
            print("✅ All markets already exist in database")
            return existing_markets
        
        print(f"🏗️  Creating {len(missing_markets)} missing markets...")
        
        with self.db.transaction() as conn:
            # ENHANCED: Use tqdm for market creation
            for market_code in tqdm(sorted(missing_markets), desc="🏗️  Creating markets", unit=" markets"):
                market_name = self._generate_market_name(market_code)
                
                cursor = conn.execute("""
                    INSERT INTO markets (market_code, market_name) 
                    VALUES (?, ?)
                """, (market_code, market_name))
                
                market_id = cursor.lastrowid
                existing_markets[market_code] = market_id
                
                tqdm.write(f"   ✅ Created market: {market_code} ({market_name}) - ID: {market_id}")
        
        print(f"✅ Market creation complete")
        return existing_markets
    
    def setup_schedule_assignments(self, excel_markets: Dict[str, Dict], market_mapping: Dict[str, int]) -> int:
        """Create schedule assignments with tqdm progress bar."""
        markets_with_schedules = self.get_markets_with_schedules()
        assignments_created = 0
        
        print(f"🗓️  Setting up schedule assignments...")
        
        # Filter markets that need assignments
        markets_needing_assignments = [
            (market_code, market_data) for market_code, market_data in excel_markets.items()
            if market_mapping[market_code] not in markets_with_schedules
        ]
        
        if not markets_needing_assignments:
            print("✅ All markets already have schedule assignments")
            return 0
        
        with self.db.transaction() as conn:
            # ENHANCED: Use tqdm for schedule assignments
            for market_code, market_data in tqdm(markets_needing_assignments, desc="🗓️  Creating assignments", unit=" assignments"):
                market_id = market_mapping[market_code]
                earliest_date = market_data['earliest_date']
                
                cursor = conn.execute("""
                    INSERT INTO schedule_market_assignments 
                    (schedule_id, market_id, effective_start_date, assignment_priority)
                    VALUES (1, ?, ?, 1)
                """, (market_id, earliest_date))
                
                assignments_created += 1
                tqdm.write(f"   ✅ {market_code}: Created schedule assignment (effective from {earliest_date})")
        
        print(f"✅ Schedule assignment setup complete: {assignments_created} assignments created")
        return assignments_created
    
    # Keep existing helper methods...
    def get_existing_markets(self) -> Dict[str, int]:
        """Get existing markets from database."""
        with self.db.transaction() as conn:
            cursor = conn.execute("SELECT market_code, market_id FROM markets")
            return {row[0]: row[1] for row in cursor.fetchall()}
    
    def get_markets_with_schedules(self) -> Set[int]:
        """Get market IDs that already have schedule assignments."""
        with self.db.transaction() as conn:
            cursor = conn.execute("SELECT DISTINCT market_id FROM schedule_market_assignments")
            return {row[0] for row in cursor.fetchall()}
    
    def _generate_market_name(self, market_code: str) -> str:
        """Generate a proper market name from market code."""
        name_mappings = {
            'NYC': 'NEW YORK', 'LAX': 'LOS ANGELES', 'SFO': 'SAN FRANCISCO',
            'SEA': 'SEATTLE', 'CHI': 'CHICAGO', 'MSP': 'MINNEAPOLIS',
            'DAL': 'DALLAS', 'HOU': 'HOUSTON', 'WDC': 'WASHINGTON DC',
            'CVC': 'CENTRAL VALLEY', 'CMP': 'CHI MSP', 'MMT': 'MAMMOTH',
            'ADMIN': 'ADMINISTRATIVE'
        }
        
        return name_mappings.get(market_code, market_code.upper().replace('_', ' '))


class EnhancedHistoricalImporter:
    """Enhanced historical importer with tqdm progress bars and language assignment integration."""
    
    def __init__(self, db_connection: DatabaseConnection):
        self.db = db_connection
        self.market_manager = EnhancedMarketSetupManager(db_connection)
        self.import_service = BroadcastMonthImportService(db_connection)
        self.process = psutil.Process(os.getpid())
    
    def execute_enhanced_import(self, 
                              excel_file: str,
                              expected_year: int,
                              closed_by: str,
                              auto_setup_markets: bool = True,
                              dry_run: bool = False) -> Dict:
        """
        Execute enhanced historical import with tqdm progress bars and language assignment.
        """
        start_time = datetime.now()
        batch_id = f"enhanced_historical_{int(start_time.timestamp())}"
        
        print(f"Production Sales Data Import Starting...")
        print(f"Excel file: {excel_file}")
        print(f"Expected year: {expected_year}")
        print(f"Closed by: {closed_by}")
        print(f"Auto-setup markets: {auto_setup_markets}")
        print(f"Dry run: {dry_run}")
        print(f"Batch ID: {batch_id}")
        print("=" * 70)
        
        results = {
            'success': False,
            'batch_id': batch_id,
            'market_setup': None,
            'import_result': None,
            'language_assignment': None,  # NEW
            'duration_seconds': 0,
            'error_messages': []
        }
        
        try:
            # Step 1: Market setup (if enabled)
            if auto_setup_markets and not dry_run:
                print(f"🏗️  STEP 1: Automatic Market Setup")
                market_setup_result = self.market_manager.execute_market_setup(excel_file)
                results['market_setup'] = market_setup_result
                
                print(f"📊 Market Setup Results:")
                print(f"   🎯 Markets found in Excel: {market_setup_result['markets_found']}")
                print(f"   🏗️  New markets created: {market_setup_result.get('markets_created', 0)}")
                print(f"   🗓️  Schedule assignments created: {market_setup_result['assignments_created']}")
                print(f"   📅 Schedule dates updated: {market_setup_result['dates_updated']}")
                print()
            
            # Step 2: Historical data import
            print(f"📦 STEP 2: Historical Data Import")
            
            if dry_run:
                print(f"🔍 DRY RUN - No changes would be made")
                summary = get_excel_import_summary(excel_file, self.db.db_path)
                results['import_result'] = {
                    'would_import': True,
                    'months_found': len(summary['months_in_excel']),
                    'total_spots': summary['total_existing_spots_affected']
                }
            else:
                # Execute actual import with progress tracking
                import_result = self.import_service.execute_month_replacement(
                    excel_file,
                    'HISTORICAL',
                    closed_by,
                    dry_run=False
                )
                results['import_result'] = import_result
                
                # NEW: Step 3: Language Assignment Processing
                if import_result.success:
                    print(f"\n🎯 STEP 3: Language Assignment Processing")
                    language_result = self._process_language_assignments(batch_id)
                    results['language_assignment'] = language_result
            
            results['success'] = True
            
        except Exception as e:
            error_msg = f"Production import failed: {str(e)}"
            results['error_messages'].append(error_msg)
            print(f"❌ {error_msg}")
        
        # Calculate total duration
        results['duration_seconds'] = (datetime.now() - start_time).total_seconds()
        
        return results
    
    def _process_language_assignments(self, batch_id: str) -> Dict:
        """
        NEW: Process language assignments after import with tqdm progress bars.
        """
        language_result = {
            'success': False,
            'categorized': 0,
            'processed': 0,
            'flagged_for_review': 0,
            'error_messages': []
        }
        
        try:
            # Import language assignment services
            from src.services.spot_categorization_service import SpotCategorizationService
            from src.services.language_processing_orchestrator import LanguageProcessingOrchestrator
            
            conn = sqlite3.connect(self.db.db_path)
            
            # Step 3a: Categorization with progress
            print("🏷️  Categorizing spots for language assignment...")
            categorization_service = SpotCategorizationService(conn)
            
            # Get uncategorized spots for progress tracking
            uncategorized_spots = categorization_service.get_uncategorized_spots()
            
            if uncategorized_spots:
                with tqdm(total=len(uncategorized_spots), desc="🏷️  Categorizing spots", unit=" spots") as pbar:
                    # Categorize in batches for better progress reporting
                    batch_size = 1000
                    for i in range(0, len(uncategorized_spots), batch_size):
                        batch = uncategorized_spots[i:i + batch_size]
                        categorization_service.categorize_spots_batch(batch)
                        pbar.update(len(batch))
                        pbar.set_description(f"🏷️  Categorized {min(i + batch_size, len(uncategorized_spots)):,}/{len(uncategorized_spots):,}")
                
                language_result['categorized'] = len(uncategorized_spots)
            
            # Step 3b: Process all categories with progress
            print("🎯 Processing language assignments...")
            orchestrator = LanguageProcessingOrchestrator(conn)
            
            # Get processing status for progress tracking
            processing_results = orchestrator.process_all_categories()
            
            language_result['processed'] = processing_results['summary']['total_processed']
            language_result['flagged_for_review'] = processing_results['summary']['flagged_for_review']
            language_result['success'] = True
            
            print(f"✅ Language assignment complete:")
            print(f"   📊 Spots categorized: {language_result['categorized']:,}")
            print(f"   🔤 Spots processed: {language_result['processed']:,}")
            print(f"   📋 Flagged for review: {language_result['flagged_for_review']:,}")
            
            conn.close()
            
        except Exception as e:
            error_msg = f"Language assignment processing failed: {str(e)}"
            language_result['error_messages'].append(error_msg)
            print(f"⚠️  {error_msg}")
        
        return language_result
    
    def execute_market_setup(self, excel_file: str) -> Dict:
        """Execute complete market setup process with enhanced progress tracking."""
        print(f"🚀 Starting automatic market setup...")
        start_time = datetime.now()
        
        # Step 1: Scan Excel for markets (now with tqdm)
        excel_markets = self.market_manager.scan_excel_for_markets(excel_file)
        
        # Step 2: Create missing markets (now with tqdm)
        market_mapping = self.market_manager.create_missing_markets(excel_markets)
        
        # Step 3: Setup schedule assignments (now with tqdm)
        assignments_created = self.market_manager.setup_schedule_assignments(excel_markets, market_mapping)
        
        # Step 4: Update existing schedule dates if needed
        dates_updated = self.market_manager.update_existing_schedule_dates(excel_markets, market_mapping)
        
        duration = (datetime.now() - start_time).total_seconds()
        
        summary = {
            'markets_found': len(excel_markets),
            'markets_created': len([m for m in excel_markets.keys() if m not in self.market_manager.get_existing_markets()]),
            'assignments_created': assignments_created,
            'dates_updated': dates_updated,
            'duration_seconds': duration,
            'excel_markets': excel_markets
        }
        
        print(f"✅ Market setup complete in {duration:.2f} seconds!")
        return summary


def display_production_preview(excel_file: str, expected_year: int, db_path: str, auto_setup: bool):
    """Display comprehensive preview with progress estimation."""
    print(f"Production Import Preview")
    print(f"=" * 70)
    print(f"Excel file: {excel_file}")
    print(f"Expected year: {expected_year}")
    print(f"Auto-setup markets: {auto_setup}")
    print()
    
    try:
        db_connection = DatabaseConnection(db_path)
        
        # Preview with progress estimation
        print("🔍 Analyzing Excel file for preview...")
        
        if auto_setup:
            # Market setup preview with progress
            market_manager = EnhancedMarketSetupManager(db_connection)
            
            # Quick scan for preview (without full tqdm)
            from openpyxl import load_workbook
            workbook = load_workbook(excel_file, read_only=True, data_only=True)
            
            total_rows = workbook.max_row - 1
            print(f"📊 File contains ~{total_rows:,} rows to analyze")
            
            # Quick market analysis
            excel_markets = market_manager.scan_excel_for_markets(excel_file)
            existing_markets = market_manager.get_existing_markets()
            
            missing_markets = set(excel_markets.keys()) - set(existing_markets.keys())
            
            print(f"🏗️  Market Setup Preview:")
            print(f"   📊 Markets in Excel: {len(excel_markets)}")
            print(f"   ✅ Already exist: {len(existing_markets)}")
            print(f"   🏗️  Will create: {len(missing_markets)}")
            
            if missing_markets:
                print(f"   📋 New markets to create: {sorted(missing_markets)}")
            print()
            
            workbook.close()
        
        # Show import preview
        summary = get_excel_import_summary(excel_file, db_path)
        
        print(f"📦 Import Preview:")
        print(f"   📅 Months found: {len(summary['months_in_excel'])}")
        print(f"   📊 Total spots to process: {summary['total_existing_spots_affected']:,}")
        print(f"   🔒 Closed months: {len(summary['closed_months'])}")
        print(f"   📂 Open months: {len(summary['open_months'])}")
        
        # NEW: Language assignment preview
        print(f"")
        print(f"🎯 Language Assignment Preview:")
        print(f"   📋 All imported spots will be categorized for language assignment")
        print(f"   🔤 Business rules will be applied automatically")
        print(f"   📋 Spots requiring manual review will be flagged")
        
        db_connection.close()
        
    except Exception as e:
        print(f"❌ Error generating preview: {e}")


def main():
    """Enhanced main function with language assignment integration."""
    parser = argparse.ArgumentParser(
        description="Production Sales Data Import Tool",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python bulk_import_closed_data.py data/monthly/May-2025.xlsx --year 2025 --closed-by "Kurt"
  python bulk_import_closed_data.py data/historical/2024-complete.xlsx --year 2024 --closed-by "Kurt" --auto-setup
  python bulk_import_closed_data.py data/March-2025.xlsx --year 2025 --closed-by "Kurt" --dry-run

Features:
  - Real-time progress tracking
  - Automatic market detection and creation  
  - Integrated language assignment processing
  - Permanent month closure protection
  - Comprehensive audit trail
        """
    )
    
    parser.add_argument("excel_file", help="Path to Excel file to import")
    parser.add_argument("--year", type=int, required=True, help="Expected year for validation")
    parser.add_argument("--closed-by", required=True, help="Name/ID of person performing the import")
    parser.add_argument("--auto-setup", action="store_true", help="Automatically create missing markets and schedule assignments")
    parser.add_argument("--db-path", default="data/database/production.db", help="Database path")
    parser.add_argument("--dry-run", action="store_true", help="Preview import without making changes")
    parser.add_argument("--force", action="store_true", help="Skip confirmation prompts")
    parser.add_argument("--verbose", action="store_true", help="Enable verbose logging")
    
    args = parser.parse_args()
    
    # Validation and execution logic (same as before but with enhanced progress reporting)...
    
    try:
        print(f"Production Sales Data Import Tool")
        print(f"=" * 60)
        
        # Display production preview
        display_production_preview(args.excel_file, args.year, args.db_path, args.auto_setup)
        
        # Get confirmation unless forced or dry run
        if not args.dry_run and not args.force:
            print(f"\n🚨 CONFIRMATION REQUIRED")
            action_list = ["Import historical data and close months", "Process language assignments automatically"]
            if args.auto_setup:
                action_list.insert(0, "Create missing markets and schedule assignments")
            
            print(f"This will:")
            for i, action in enumerate(action_list, 1):
                print(f"  {i}. {action}")
            
            response = input(f"\nProceed with import? (type 'yes' to confirm): ").strip().lower()
            if response != 'yes':
                print(f"❌ Import cancelled by user")
                sys.exit(0)
        
        # Execute enhanced import
        db_connection = DatabaseConnection(args.db_path)
        importer = EnhancedHistoricalImporter(db_connection)
        
        try:
            results = importer.execute_enhanced_import(
                args.excel_file,
                args.year,
                args.closed_by,
                args.auto_setup,
                args.dry_run
            )
            
            # Display enhanced results
            print(f"\n" + "=" * 70)
            print(f"PRODUCTION IMPORT {'PREVIEW' if args.dry_run else 'COMPLETED'}")
            print(f"=" * 70)
            
            print(f"📊 Overall Results:")
            print(f"  Success: {'✅' if results['success'] else '❌'}")
            print(f"  Duration: {results['duration_seconds']:.2f} seconds")
            print(f"  Batch ID: {results['batch_id']}")
            
            # Market setup results
            if results['market_setup']:
                setup = results['market_setup']
                print(f"\n🏗️  Market Setup Results:")
                print(f"  Markets found: {setup['markets_found']}")
                print(f"  New markets created: {setup.get('markets_created', 0)}")
                print(f"  Schedule assignments created: {setup['assignments_created']}")
                print(f"  Schedule dates updated: {setup['dates_updated']}")
            
            # Import results
            if results['import_result']:
                import_res = results['import_result']
                print(f"\n📦 Import Results:")
                if not args.dry_run and hasattr(import_res, 'success'):
                    print(f"  Records deleted: {import_res.records_deleted:,}")
                    print(f"  Records imported: {import_res.records_imported:,}")
                    print(f"  Months affected: {len(import_res.broadcast_months_affected)}")
                    if hasattr(import_res, 'closed_months') and import_res.closed_months:
                        print(f"  Months closed: {import_res.closed_months}")
                else:
                    print(f"  Would import: {import_res.get('total_spots', 0):,} spots")
                    print(f"  Months found: {import_res.get('months_found', 0)}")
            
            # NEW: Language assignment results
            if results['language_assignment']:
                lang_res = results['language_assignment']
                print(f"\n🎯 Language Assignment Results:")
                print(f"  Spots categorized: {lang_res['categorized']:,}")
                print(f"  Spots processed: {lang_res['processed']:,}")
                print(f"  Flagged for review: {lang_res['flagged_for_review']:,}")
                print(f"  Success: {'✅' if lang_res['success'] else '❌'}")
                
                if lang_res['error_messages']:
                    print(f"  Errors: {len(lang_res['error_messages'])}")
            
            if results['error_messages']:
                print(f"\n❌ Errors:")
                for error in results['error_messages']:
                    print(f"  • {error}")
                sys.exit(1)
            
            if results['success'] and not args.dry_run:
                print(f"\nClosed data import completed successfully.")
                print(f"Next steps:")
                print(f"  - All closed data is imported and protected")
                print(f"  - Markets and schedules are configured")
                print(f"  - Language assignments have been processed")
                print(f"  - Check for spots requiring manual review:")
                print(f"    uv run python cli_01_language_assignment.py --review-required")
                print(f"  - Database is ready for reporting and analysis")
        
        finally:
            db_connection.close()
    
    except Exception as e:
        print(f"❌ Production import error: {e}")
        if args.verbose:
            import traceback
            traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()